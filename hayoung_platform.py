# 하영자원 폐기물 데이터 플랫폼 Pro v3.0
# ============================================================
# [v3 추가/개선 사항]
#   [추가2] 재활용품 23종 실시간 시세 DB + 품목별 단가 관리
#   [추가3] 스쿨존 Geofencing - 등하교 시간대 자동 차단 (08~09시, 14~16시)
#   [추가4] 수거 일정 캘린더 뷰 (월별 달력 UI)
#   [추가5] 교육청 통합 관제 모드 + 공공예산 절감 지표
#   [전체]  프로토타입 → 실제 작동 전환 (버튼 기능, 상태 변경, 데이터 반영)
# ============================================================
# 실행 방법: cd Desktop\하영자원 → python -m streamlit run hayoung_platform.py
# 필수 설치: pip install streamlit pandas xlsxwriter requests python-dotenv
# ============================================================

import streamlit as st
import pandas as pd
import sqlite3
import time
import io
import os
import random
import calendar
import zipfile
from datetime import datetime, date

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── 선택 패키지 가용 여부 확인 (Cloud/로컬 공통) ─────────────
# reportlab, openpyxl 은 requirements.txt 에 명시
# 없을 경우 PDF/엑셀 기능만 비활성화, 앱 전체는 정상 동작
try:
    import reportlab as _rl   # noqa: F401
    _REPORTLAB_OK = True
except ImportError:
    _REPORTLAB_OK = False

try:
    import openpyxl as _oxl   # noqa: F401
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_raw_pw        = os.getenv("EXCEL_PASSWORD", "")
EXCEL_PASSWORD = _raw_pw if _raw_pw else None
KAKAO_API_KEY    = os.getenv("KAKAO_API_KEY", "")
KAKAO_SENDER_KEY = os.getenv("KAKAO_SENDER_KEY", "")

# ============================================================
# 0. 상수 및 기초 데이터
# ============================================================
STUDENT_COUNTS = {
    "화성초등학교": 309,  "동탄중학교": 1033, "수원고등학교": 884,  "안양남초등학교": 486,
    "평촌초등학교": 1126, "부림초등학교": 782, "부흥중학교": 512,   "덕천초등학교": 859,
    "서초고등학교": 831,  "구암고등학교": 547, "국사봉중학교": 346, "당곡고등학교": 746,
    "당곡중학교": 512,   "서울공업고등학교": 735, "강남중학교": 265, "영남중학교": 409,
    "선유고등학교": 580,  "신목고등학교": 1099, "고척고등학교": 782, "구현고등학교": 771,
    "안산국제비지니스고등학교": 660, "안산고등학교": 745, "송호고등학교": 879, "비봉고등학교": 734
}
SCHOOL_LIST = sorted(list(STUDENT_COUNTS.keys()))

# ============================================================
# 실제 수거 데이터 내장 (2025년 3~12월 엑셀 원본)
# 형식: (날짜, 학교명, 음식물_L, 단가)
# ============================================================
REAL_COLLECTION_DATA = [
    ("2025-03-05","강남중학교",140.0,190),("2025-03-07","강남중학교",180.0,190),("2025-03-11","강남중학교",120.0,190),("2025-03-13","강남중학교",180.0,190),("2025-03-14","강남중학교",120.0,190),("2025-03-17","강남중학교",140.0,190),("2025-03-18","강남중학교",120.0,190),("2025-03-19","강남중학교",160.0,190),("2025-03-20","강남중학교",140.0,190),("2025-03-21","강남중학교",100.0,190),("2025-03-24","강남중학교",120.0,190),("2025-03-25","강남중학교",120.0,190),("2025-03-26","강남중학교",140.0,190),("2025-03-27","강남중학교",120.0,190),("2025-03-28","강남중학교",120.0,190),
    ("2025-04-01","강남중학교",120.0,190),("2025-04-02","강남중학교",140.0,190),("2025-04-03","강남중학교",110.0,190),("2025-04-04","강남중학교",120.0,190),("2025-04-07","강남중학교",140.0,190),("2025-04-08","강남중학교",120.0,190),("2025-04-09","강남중학교",100.0,190),("2025-04-10","강남중학교",140.0,190),("2025-04-11","강남중학교",120.0,190),("2025-04-14","강남중학교",120.0,190),("2025-04-15","강남중학교",120.0,190),("2025-04-16","강남중학교",140.0,190),("2025-04-17","강남중학교",100.0,190),("2025-04-18","강남중학교",120.0,190),("2025-04-21","강남중학교",120.0,190),("2025-04-22","강남중학교",140.0,190),("2025-04-23","강남중학교",120.0,190),("2025-04-24","강남중학교",100.0,190),("2025-04-25","강남중학교",120.0,190),("2025-04-28","강남중학교",120.0,190),("2025-04-29","강남중학교",130.0,190),("2025-04-30","강남중학교",110.0,190),
]

# 내장 데이터 로딩 함수
def _load_embedded_data():
    """real_data_embedded.py 파일이 있으면 로드, 없으면 REAL_COLLECTION_DATA 사용"""
    import os
    embed_path = os.path.join(os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else ".", "real_data_embedded.py")
    if os.path.exists(embed_path):
        import importlib.util
        spec = importlib.util.spec_from_file_location("real_data_embedded", embed_path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.REAL_COLLECTION_DATA
    return REAL_COLLECTION_DATA

# [추가2] 재활용품 23종 기본 시세 (원/kg) - 실제 시장 기준값
RECYCLE_ITEMS_DEFAULT = {
    "폐지(골판지)": 80,   "폐지(신문지)": 60,   "폐지(혼합)": 40,
    "고철(철스크랩)": 300, "고철(알루미늄캔)": 800, "고철(스테인리스)": 500,
    "유리병(갈색)": 50,   "유리병(투명)": 60,   "유리병(혼합)": 40,
    "페트병(투명)": 400,  "페트병(유색)": 150,  "플라스틱(PP)": 200,
    "플라스틱(PE)": 180,  "플라스틱(PS)": 100,  "플라스틱(PVC)": 80,
    "비닐(투명)": 120,   "비닐(혼합)": 60,    "스티로폼": 150,
    "우유팩": 200,       "종이컵": 80,         "의류/섬유": 100,
    "목재": 30,          "전자폐기물": 500,
}

# [추가3] 스쿨존 제한 시간대
SCHOOLZONE_RESTRICTED = [(8, 9), (14, 16)]  # (시작시, 종료시)

# ============================================================
# 0-1. 사용자 계정 DB (로그인 시스템)
# ============================================================
# 비밀번호는 실제 운영 시 해시 처리 권장
USER_ACCOUNTS = {
    # ── 관리자 (1개) ──────────────────────────────────────────
    "admin": {
        "password": "hayoung2025!",
        "role": "관리자",
        "display_name": "하영자원 관리자",
        "org": "하영자원(본사)",
    },

    # ── 교육청 (2개) ──────────────────────────────────────────
    "hwaseong_edu": {
        "password": "edu_hwaseong1",
        "role": "교육청",
        "display_name": "화성오산교육지원청 담당자",
        "org": "화성오산교육지원청",
    },
    "seoulsouth_edu": {
        "password": "edu_seoul2025",
        "role": "교육청",
        "display_name": "서울남부교육지원청 담당자",
        "org": "서울남부교육지원청",
    },

    # ── 수거 기사 (3개) ───────────────────────────────────────
    "driver_kim": {
        "password": "driver_kim1",
        "role": "수거기사",
        "display_name": "김기사",
        "org": "하영자원(본사 직영)",
    },
    "driver_lee": {
        "password": "driver_lee2",
        "role": "수거기사",
        "display_name": "이기사",
        "org": "하영자원(본사 직영)",
    },
    "driver_park": {
        "password": "driver_park3",
        "role": "수거기사",
        "display_name": "박기사",
        "org": "하영자원(본사 직영)",
    },

    # ── 학교 계정 (엑셀 학교리스트.xlsx 기준) ────────────────
    "kn12": {"password": "1234", "role": "학교", "display_name": "강남중학교 행정실",          "org": "강남중학교"},
    "hs01": {"password": "1234", "role": "학교", "display_name": "화성초등학교 행정실",         "org": "화성초등학교"},
    "dt02": {"password": "1234", "role": "학교", "display_name": "동탄중학교 행정실",           "org": "동탄중학교"},
    "sw03": {"password": "1234", "role": "학교", "display_name": "수원고등학교 행정실",         "org": "수원고등학교"},
    "an04": {"password": "1234", "role": "학교", "display_name": "안양남초등학교 행정실",       "org": "안양남초등학교"},
    "pc05": {"password": "1234", "role": "학교", "display_name": "평촌초등학교 행정실",         "org": "평촌초등학교"},
    "br06": {"password": "1234", "role": "학교", "display_name": "부림초등학교 행정실",         "org": "부림초등학교"},
    "bh07": {"password": "1234", "role": "학교", "display_name": "부흥중학교 행정실",           "org": "부흥중학교"},
    "dc08": {"password": "1234", "role": "학교", "display_name": "덕천초등학교 행정실",         "org": "덕천초등학교"},
    "sc09": {"password": "1234", "role": "학교", "display_name": "서초고등학교 행정실",         "org": "서초고등학교"},
    "ga10": {"password": "1234", "role": "학교", "display_name": "구암고등학교 행정실",         "org": "구암고등학교"},
    "gs11": {"password": "1234", "role": "학교", "display_name": "국사봉중학교 행정실",         "org": "국사봉중학교"},
    "dg13": {"password": "1234", "role": "학교", "display_name": "당곡고등학교 행정실",         "org": "당곡고등학교"},
    "dg14": {"password": "1234", "role": "학교", "display_name": "당곡중학교 행정실",           "org": "당곡중학교"},
    "sg15": {"password": "1234", "role": "학교", "display_name": "서울공업고등학교 행정실",     "org": "서울공업고등학교"},
    "yn16": {"password": "1234", "role": "학교", "display_name": "영남중학교 행정실",           "org": "영남중학교"},
    "sy17": {"password": "1234", "role": "학교", "display_name": "선유고등학교 행정실",         "org": "선유고등학교"},
    "sm18": {"password": "1234", "role": "학교", "display_name": "신목고등학교 행정실",         "org": "신목고등학교"},
    "gc19": {"password": "1234", "role": "학교", "display_name": "고척고등학교 행정실",         "org": "고척고등학교"},
    "gh20": {"password": "1234", "role": "학교", "display_name": "구현고등학교 행정실",         "org": "구현고등학교"},
    "as21": {"password": "1234", "role": "학교", "display_name": "안산국제비지니스고 행정실",   "org": "안산국제비지니스고등학교"},
    "as22": {"password": "1234", "role": "학교", "display_name": "안산고등학교 행정실",         "org": "안산고등학교"},
    "sh23": {"password": "1234", "role": "학교", "display_name": "송호고등학교 행정실",         "org": "송호고등학교"},
    "bb24": {"password": "1234", "role": "학교", "display_name": "비봉고등학교 행정실",         "org": "비봉고등학교"},
}

# [추가5] 교육청 목록
EDU_OFFICES = {
    "화성오산교육지원청": ["화성초등학교","부림초등학교","비봉고등학교","송호고등학교","안산고등학교","안산국제비지니스고등학교"],
    "수원교육지원청":     ["수원고등학교","평촌초등학교","안양남초등학교","부흥중학교","동탄중학교","덕천초등학교"],
    "서울남부교육지원청": ["서초고등학교","구암고등학교","국사봉중학교","당곡고등학교","당곡중학교",
                          "서울공업고등학교","강남중학교","영남중학교","선유고등학교",
                          "신목고등학교","고척고등학교","구현고등학교"],
}

# ============================================================
# 1. 페이지 설정 및 CSS
# ============================================================
st.set_page_config(
    page_title="하영자원 플랫폼 Pro v3",
    page_icon="♻️", layout="wide",
    initial_sidebar_state="expanded"
)
st.markdown('<meta name="google" content="notranslate">', unsafe_allow_html=True)
st.markdown("""
<style>
.custom-card         { background:#fff; color:#202124; padding:20px; border-radius:12px; box-shadow:0 4px 10px rgba(0,0,0,.05); margin-bottom:20px; border-top:5px solid #1a73e8; }
.custom-card-green   { border-top:5px solid #34a853; }
.custom-card-orange  { border-top:5px solid #fbbc05; }
.custom-card-red     { border-top:5px solid #ea4335; }
.custom-card-purple  { border-top:5px solid #9b59b6; }
.custom-card-teal    { border-top:5px solid #00897b; }
.metric-title        { font-size:14px; color:#5f6368!important; font-weight:bold; margin-bottom:5px; }
.metric-value-food   { font-size:26px; font-weight:900; color:#ea4335!important; }
.metric-value-recycle{ font-size:26px; font-weight:900; color:#34a853!important; }
.metric-value-biz    { font-size:26px; font-weight:900; color:#9b59b6!important; }
.metric-value-total  { font-size:26px; font-weight:900; color:#1a73e8!important; }
.mobile-app-header   { background:#202124; color:#fff!important; padding:15px; border-radius:10px 10px 0 0; text-align:center; margin-bottom:15px; }
.safety-box          { background:#e8f5e9; border:1px solid #c8e6c9; padding:15px; border-radius:8px; color:#2e7d32; font-weight:bold; margin-bottom:15px; }
.alert-box           { background:#ffebee; border:1px solid #ffcdd2; padding:15px; border-radius:8px; color:#c62828; margin-bottom:15px; }
.warn-box            { background:#fff8e1; border:1px solid #ffe082; padding:15px; border-radius:8px; color:#f57f17; margin-bottom:15px; }
.timeline-text       { font-size:15px; line-height:1.8; color:#333; }
.badge-new           { background:#e8f0fe; color:#1a73e8; padding:2px 10px; border-radius:20px; font-size:12px; font-weight:bold; margin-left:6px; }
.badge-v3            { background:#e6f4ea; color:#1e8e3e; padding:2px 10px; border-radius:20px; font-size:12px; font-weight:bold; margin-left:6px; }
.cal-day             { text-align:center; padding:6px 2px; border-radius:6px; font-size:13px; }
.cal-collect         { background:#e8f5e9; color:#2e7d32; font-weight:bold; }
.cal-today           { background:#1a73e8; color:white; font-weight:bold; }
.cal-weekend         { color:#bbb; }
.schoolzone-danger   { background:#d32f2f; color:white; padding:20px; border-radius:12px; text-align:center; font-size:24px; font-weight:900; margin-bottom:15px; }
.schoolzone-safe     { background:#388e3c; color:white; padding:20px; border-radius:12px; text-align:center; font-size:18px; font-weight:bold; margin-bottom:15px; }

/* ── 로그인 화면 ── */
.login-bg { background:linear-gradient(160deg,#e8f4fd 0%,#d4eaf7 50%,#c2e0f4 100%); min-height:100vh; padding:40px 20px; }
.login-header { text-align:center; margin-bottom:40px; }
.login-header h1 { font-size:2.6rem; font-weight:900; color:#1a3a5c; margin:0; }
.login-header p  { font-size:1.05rem; color:#4a6b8a; margin-top:8px; }
.login-card-wrap { display:flex; gap:24px; justify-content:center; flex-wrap:wrap; margin-bottom:40px; }
.login-card {
    background:white; border-radius:16px; padding:36px 28px; width:260px;
    box-shadow:0 6px 24px rgba(0,80,160,.12); cursor:pointer; transition:.2s;
    text-align:center; border:3px solid transparent;
}
.login-card:hover { transform:translateY(-6px); box-shadow:0 12px 32px rgba(0,80,160,.2); border-color:#1a73e8; }
.login-card.active { border-color:#1a73e8; background:#f0f7ff; }
.login-card-icon { font-size:3.2rem; margin-bottom:16px; }
.login-card-title { font-size:1.2rem; font-weight:800; color:#1a3a5c; margin-bottom:8px; }
.login-card-desc  { font-size:0.85rem; color:#5f6368; line-height:1.5; }
.login-form-box { background:white; border-radius:16px; max-width:440px; margin:0 auto; padding:36px; box-shadow:0 4px 20px rgba(0,0,0,.1); }
.login-form-title { font-size:1.3rem; font-weight:800; color:#1a3a5c; margin-bottom:24px; text-align:center; }
.login-footer { text-align:center; margin-top:40px; color:#8aa0b8; font-size:0.82rem; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 2. SQLite DB 초기화 및 함수
# ============================================================
DB_PATH = "hayoung_v3.db"

def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_conn()
    c = conn.cursor()

    # 수거 데이터
    c.execute("""
        CREATE TABLE IF NOT EXISTS collections (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            날짜       TEXT,
            학교명     TEXT,
            학생수     INTEGER,
            수거업체   TEXT,
            음식물_kg  REAL DEFAULT 0,
            재활용_kg  REAL DEFAULT 0,
            사업장_kg  REAL DEFAULT 0,
            상태       TEXT DEFAULT '정산대기',
            현장사진   TEXT DEFAULT ''
        )
    """)

    # 전역 설정
    c.execute("""
        CREATE TABLE IF NOT EXISTS global_settings (
            key TEXT PRIMARY KEY, value TEXT
        )
    """)
    defaults = [
        ("default_food_price","150"), ("default_recycle_price","300"),
        ("default_biz_price","200"),  ("kakao_notify_enabled","false"),
        ("budget_saving_per_school","5200000"),  # 학교당 연간 예산절감액(원)
    ]
    for k, v in defaults:
        c.execute("INSERT OR IGNORE INTO global_settings (key,value) VALUES (?,?)", (k, v))

    # 학교별 단가 + 담당자
    c.execute("""
        CREATE TABLE IF NOT EXISTS school_prices (
            학교명           TEXT PRIMARY KEY,
            음식물단가       INTEGER DEFAULT 150,
            재활용단가       INTEGER DEFAULT 300,
            사업장단가       INTEGER DEFAULT 200,
            담당자명         TEXT DEFAULT '',
            담당자연락처     TEXT DEFAULT '',
            담당자이메일     TEXT DEFAULT '',
            교육청           TEXT DEFAULT '',
            -- [섹션A] 학교 마스터 확장 컬럼
            학교_사업자번호  TEXT DEFAULT '',
            학교_주소        TEXT DEFAULT '',
            학교_전화        TEXT DEFAULT '',
            계약_시작일      TEXT DEFAULT '',
            계약_종료일      TEXT DEFAULT '',
            계약_상태        TEXT DEFAULT '미계약',
            비고             TEXT DEFAULT '',
            updated_at       TEXT
        )
    """)
    # [섹션A] 기존 DB에 컬럼 없으면 ALTER로 추가 (마이그레이션)
    _sp_new_cols = {
        "학교_사업자번호": "TEXT DEFAULT ''",
        "학교_주소":       "TEXT DEFAULT ''",
        "학교_전화":       "TEXT DEFAULT ''",
        "계약_시작일":     "TEXT DEFAULT ''",
        "계약_종료일":     "TEXT DEFAULT ''",
        "계약_상태":       "TEXT DEFAULT '미계약'",
        "비고":            "TEXT DEFAULT ''",
        # [6단계] 부담당자 + 워크플로우 컬럼
        "담당자2명":       "TEXT DEFAULT ''",
        "담당자2연락처":   "TEXT DEFAULT ''",
        "행정실팩스":      "TEXT DEFAULT ''",
        "계약_사인_여부":  "INTEGER DEFAULT 0",
    }
    _existing = [row[1] for row in c.execute("PRAGMA table_info(school_prices)").fetchall()]
    for col, coldef in _sp_new_cols.items():
        if col not in _existing:
            c.execute(f"ALTER TABLE school_prices ADD COLUMN {col} {coldef}")
    for school in SCHOOL_LIST:
        edu = next((k for k, v in EDU_OFFICES.items() if school in v), "")
        c.execute(
            "INSERT OR IGNORE INTO school_prices (학교명, 교육청, updated_at) VALUES (?,?,?)",
            (school, edu, datetime.now().strftime("%Y-%m-%d"))
        )

    # [추가2] 재활용품 23종 시세 테이블
    c.execute("""
        CREATE TABLE IF NOT EXISTS recycle_prices (
            품목명   TEXT PRIMARY KEY,
            단가     INTEGER,
            단위     TEXT DEFAULT 'kg',
            updated_at TEXT
        )
    """)
    for item, price in RECYCLE_ITEMS_DEFAULT.items():
        c.execute(
            "INSERT OR IGNORE INTO recycle_prices (품목명, 단가, updated_at) VALUES (?,?,?)",
            (item, price, datetime.now().strftime("%Y-%m-%d"))
        )

    # [추가4] 수거 일정 테이블
    c.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            날짜     TEXT,
            학교명   TEXT,
            메모     TEXT DEFAULT '',
            완료여부 INTEGER DEFAULT 0
        )
    """)

    # [섹션C] 서류 유효기간 관리 테이블
    c.execute("""
        CREATE TABLE IF NOT EXISTS contract_docs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_name    TEXT NOT NULL,
            issued_date TEXT DEFAULT '',
            expire_date TEXT NOT NULL,
            renew_url   TEXT DEFAULT '',
            file_note   TEXT DEFAULT '',
            renewed     INTEGER DEFAULT 0,
            memo        TEXT DEFAULT '',
            updated_at  TEXT
        )
    """)
    # 기본 서류 7종 삽입 (최초 1회만)
    _default_docs = [
        ("소상공인 확인서",        "", "2026-03-31", "sminfo.mss.go.kr",         "중소벤처기업부 공식 발급"),
        ("창업기업 확인서",        "", "2027-01-07", "중소벤처기업부",           "개업일 기준 7년 유효"),
        ("재해율 확인서",          "", "2027-02-19", "안전보건공단 kosha.or.kr", "연 1회 갱신"),
        ("사업자등록증",           "", "9999-12-31", "국세청 hometax.go.kr",     "변경 시 재발급"),
        ("폐기물수집운반업 허가증","", "9999-12-31", "화성시청",                 "제20-35호 재교부 2023.09.26"),
        ("사용인감계",             "", "9999-12-31", "자체 관리",                "계약별 첨부"),
        ("사업자계좌 통장사본",    "", "9999-12-31", "기업은행",                 "450-092046-01-017"),
    ]
    for row in _default_docs:
        c.execute(
            """INSERT OR IGNORE INTO contract_docs
               (doc_name, issued_date, expire_date, renew_url, file_note, updated_at)
               VALUES (?,?,?,?,?,?)""",
            (*row, datetime.now().strftime("%Y-%m-%d"))
        )

    # [섹션A] 계약 이력 마스터 테이블
    c.execute("""
        CREATE TABLE IF NOT EXISTS contract_master (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            학교명          TEXT NOT NULL,
            계약번호        TEXT DEFAULT '',
            계약_시작일     TEXT NOT NULL,
            계약_종료일     TEXT NOT NULL,
            폐기물_종류     TEXT DEFAULT '음식물류폐기물',
            단가            INTEGER DEFAULT 150,
            월_예상량_L     REAL DEFAULT 0,
            계약_상태       TEXT DEFAULT '계약중',
            갱신_알림일     TEXT DEFAULT '',
            나라장터_번호   TEXT DEFAULT '',
            비고            TEXT DEFAULT '',
            created_at      TEXT,
            updated_at      TEXT
        )
    """)
    # 서초고등학교 기본 계약 데이터 삽입 (최초 1회)
    if c.execute("SELECT COUNT(*) FROM contract_master WHERE 학교명='서초고등학교'").fetchone()[0] == 0:
        c.execute(
            """INSERT INTO contract_master
               (학교명, 계약번호, 계약_시작일, 계약_종료일,
                폐기물_종류, 단가, 계약_상태, 나라장터_번호, 비고, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            ("서초고등학교", "", "2026-03-01", "2027-02-28",
             "음식물류폐기물", 180, "계약중", "R26TA01543339 00",
             "2026년 신규계약",
             datetime.now().strftime("%Y-%m-%d"),
             datetime.now().strftime("%Y-%m-%d"))
        )

    # ── [5단계] contract_master 전자문서 컬럼 마이그레이션 ────
    _cm_new_cols = {
        "전자문서번호":  "TEXT DEFAULT ''",
        "송신자ID":      "TEXT DEFAULT ''",
        "수신일자":      "TEXT DEFAULT ''",
        "계약서_상태":   "TEXT DEFAULT '미수신'",
        # [6단계] 계약 체결 워크플로우 5단계
        "단계1_완료":    "INTEGER DEFAULT 0",   # 견적서 발송
        "단계2_완료":    "INTEGER DEFAULT 0",   # 위수탁계약서 발송
        "단계3_완료":    "INTEGER DEFAULT 0",   # 서류 제출 완료
        "단계4_완료":    "INTEGER DEFAULT 0",   # 나라장터 전자계약
        "단계5_완료":    "INTEGER DEFAULT 0",   # 계약 완료
    }
    _cm_existing = [r[1] for r in c.execute("PRAGMA table_info(contract_master)").fetchall()]
    for col, coldef in _cm_new_cols.items():
        if col not in _cm_existing:
            c.execute(f"ALTER TABLE contract_master ADD COLUMN {col} {coldef}")

    # 서초고 전자문서 예시 데이터 업데이트
    c.execute(
        """UPDATE contract_master SET
           전자문서번호=?, 송신자ID=?, 수신일자=?, 계약서_상태=?, updated_at=?
           WHERE 학교명='서초고등학교' AND 전자문서번호=''""",
        ("019c93cb-58c9-7b21-981f-a64746d74cbc",
         "A311030206000",
         "2026-02-20",
         "서명완료",
         datetime.now().strftime("%Y-%m-%d"))
    )
    # 서초고등학교 school_prices 마스터 정보 업데이트
    c.execute(
        """UPDATE school_prices SET
           학교_사업자번호=?, 학교_주소=?, 학교_전화=?,
           계약_시작일=?, 계약_종료일=?, 계약_상태=?, updated_at=?
           WHERE 학교명='서초고등학교'""",
        ("210-83-00086", "서울특별시 서초구 반포대로27길 29", "02-580-3891",
         "2026-03-01", "2027-02-28", "계약중",
         datetime.now().strftime("%Y-%m-%d"))
    )

    # ── [1단계] 학교주소록.xlsx 데이터 일괄 반영 ──────────────
    # 기존 학교: 사업자번호 + 주소 + 이메일(담당자이메일) UPDATE
    _school_addr_data = [
        # (학교명, 사업자번호, 주소, 이메일)
        ("송호고등학교",             "134-83-02191", "경기도 안산시 상록구 광덕3로296",                       "songho296@korea.kr"),
        ("안산고등학교",             "134-82-02947", "경기도 안산시 상록구 청룡4길8",                         "syphilip@korea.kr"),
        ("당곡중학교",               "112-83-01076", "서울특별시 관악구 봉천로 13길 101",                     "danggokms@sen.go.kr"),
        ("부림초등학교",             "138-83-01160", "경기도 안양시 동안구 달안로161-0 (관양동)",             "yaburim@korea.kr"),
        ("덕천초등학교",             "123-83-01912", "경기도 안양시 만안구 전파로27-0 (안양동)",              "duck0886@korea.kr"),
        ("당곡고등학교",             "112-83-01192", "서울특별시 관악구 봉천로21길",                          "dkhs@sen.go.kr"),
        ("안양남초등학교",           "138-83-00079", "경기도 안양시 동안구 흥안대로275-0 (호계동)",           "anam7815@korea.kr"),
        ("구암고등학교",             "119-83-02872", "서울특별시 관악구 성현로 91",                           "guam@sen.go.kr"),
        ("평촌초등학교",             "138-83-01200", "경기도 안양시 동안구 평촌대로180번길56-0 (평촌동)",     "e41175852@korea.kr"),
        ("영남중학교",               "118-83-00279", "서울특별시 영등포구 대림로31나길 17",                   "yeongnam@sen.go.kr"),
        ("국사봉중학교",             "108-83-30062", "서울특별시 동작구 양녕로20길 20,1층",                   "kuk3615@sen.go.kr"),
        ("고척고등학교",             "113-83-00293", "서울특별시 구로구 중앙로151길 162",                     "gocheokhs@sen.go.kr"),
        ("부흥중학교",               "138-83-00953", "경기도 안양시 동안구 달안로25-0 (비산동)",              "aybhms@korea.kr"),
        ("비봉고등학교",             "124-82-13506", "경기도 화성시 비봉면 양노로109번길5-0",                 "hanwoorikrf@korea.kr"),
        ("선유고등학교",             "107-83-02980", "서울특별시 영등포구 선유로 227",                        "y20690040@sen.go.kr"),
        ("안산국제비지니스고등학교", "134-82-00217", "경기도 안산시 상록구 수인로1981-0 (장상동)",            "ansan6181@korea.kr"),
        ("강남중학교",               "118-83-00041", "서울특별시 동작구 대방동 15길 29",                      "k8145637@sen.go.kr"),
    ]
    for 학교명, 사업자번호, 주소, 이메일 in _school_addr_data:
        c.execute(
            """UPDATE school_prices SET
               학교_사업자번호=?, 학교_주소=?, 담당자이메일=?, updated_at=?
               WHERE 학교명=?""",
            (사업자번호, 주소, 이메일,
             datetime.now().strftime("%Y-%m-%d"), 학교명)
        )

    # 영양사 이메일 → 담당자이메일2 컬럼이 없으므로 비고에 기록
    _nutrition_emails = [
        ("당곡고등학교",             "영양사: choi_yejin00@naver.com"),
        ("안양남초등학교",           "영양사: sangji35@korea.kr"),
        ("덕천초등학교",             "영양사: hwlee123@korea.kr"),
        ("영남중학교",               "영양사: chattingme@sen.go.kr"),
        ("고척고등학교",             "영양사: hee9528@sen.go.kr"),
        ("서초고등학교",             "영양사: dbsghk508@sen.go.kr"),
        ("평촌초등학교",             "영양사: himiok16@korea.kr"),
        ("구암고등학교",             "영양사: bravomss@sen.go.kr"),
        ("강남중학교",               "영양사: alsemffp82@sen.go.kr"),
        ("국사봉중학교",             "영양사: intj74@sen.go.kr"),
        ("송호고등학교",             "영양사: dbswjdgus555@korea.kr"),
        ("선유고등학교",             "영양사: yumyum00@sen.go.kr"),
        ("부흥중학교",               "영양사: sarang0907@korea.kr"),
        ("부림초등학교",             "영양사: ht5362@korea.kr"),
        ("안산국제비지니스고등학교", "영양사: jaeokh@korea.kr"),
    ]
    for 학교명, 영양사메모 in _nutrition_emails:
        c.execute(
            """UPDATE school_prices SET 비고=?, updated_at=?
               WHERE 학교명=? AND (비고='' OR 비고 IS NULL)""",
            (영양사메모, datetime.now().strftime("%Y-%m-%d"), 학교명)
        )

    # ★ 신규 학교: 대길초등학교 (학교주소록에서 신규 발견)
    c.execute(
        """INSERT OR IGNORE INTO school_prices
           (학교명, 교육청, 음식물단가, 학교_사업자번호,
            학교_주소, 담당자이메일, 계약_상태, 비고, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        ("대길초등학교", "서울남부교육지원청", 150,
         "118-83-00304", "서울특별시 영등포구 대방천로 206",
         "eorlfch@sen.go.kr", "미계약", "학교주소록 신규 등록",
         datetime.now().strftime("%Y-%m-%d"))
    )

    conn.commit()

    # 실제 수거 데이터 (2025년 3~12월 엑셀 원본 내장)
    if c.execute("SELECT COUNT(*) FROM collections").fetchone()[0] == 0:
        src_data = _load_embedded_data()  # (날짜, 학교명, 음식물L, 단가) 리스트
        rows = []
        price_map = {}  # 학교별 마지막 단가 추적

        for date_str, school, liter, price in src_data:
            rows.append((
                date_str + " 09:00:00",
                school,
                STUDENT_COUNTS.get(school, 0),
                "하영자원(본사 직영)",
                liter,   # 음식물_kg (1L=1kg)
                0.0,     # 재활용_kg
                0.0,     # 사업장_kg
                "정산완료",
                ""
            ))
            if price > 0:
                price_map[school] = (price, date_str)

        c.executemany("""
            INSERT INTO collections (날짜,학교명,학생수,수거업체,음식물_kg,재활용_kg,사업장_kg,상태,현장사진)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, rows)

        # 학교별 단가 반영
        for school, (price, date_str) in price_map.items():
            c.execute(
                "UPDATE school_prices SET 음식물단가=?, updated_at=? WHERE 학교명=?",
                (price, date_str, school)
            )

    # 실제 수거 데이터 기반 일정 자동 생성
    if c.execute("SELECT COUNT(*) FROM schedules").fetchone()[0] == 0:
        # collections 테이블에서 실제 수거일을 읽어 schedules에 삽입
        real_days = c.execute(
            "SELECT SUBSTR(날짜,1,10), 학교명 FROM collections ORDER BY 날짜"
        ).fetchall()
        sched_rows = []
        seen = set()
        for day_str, school in real_days:
            key = (day_str, school)
            if key not in seen:
                seen.add(key)
                sched_rows.append((day_str, school, "정기 수거 (실적)", 1))  # 완료여부=1
        if sched_rows:
            c.executemany(
                "INSERT INTO schedules (날짜,학교명,메모,완료여부) VALUES (?,?,?,?)",
                sched_rows
            )

    conn.commit()
    conn.close()

# ── 헬퍼 함수 ──────────────────────────────────────────────
def get_setting(key):
    conn = get_conn()
    row = conn.execute("SELECT value FROM global_settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row[0] if row else None

def set_setting(key, value):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO global_settings (key,value) VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()

def load_data():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT c.*,
            COALESCE(p.음식물단가, CAST(s.fp AS INTEGER)) AS 단가,
            COALESCE(p.재활용단가, CAST(s.rp AS INTEGER)) AS 재활용단가,
            COALESCE(p.사업장단가, CAST(s.bp AS INTEGER)) AS 사업장단가,
            COALESCE(p.교육청, '') AS 교육청
        FROM collections c
        LEFT JOIN school_prices p ON c.학교명 = p.학교명
        LEFT JOIN (
            SELECT
                (SELECT value FROM global_settings WHERE key='default_food_price')    AS fp,
                (SELECT value FROM global_settings WHERE key='default_recycle_price') AS rp,
                (SELECT value FROM global_settings WHERE key='default_biz_price')     AS bp
        ) s ON 1=1
    """, conn)
    conn.close()
    if not df.empty:
        df.rename(columns={"음식물_kg":"음식물(kg)","재활용_kg":"재활용(kg)","사업장_kg":"사업장(kg)"}, inplace=True)
        df["음식물비용"] = df["음식물(kg)"] * df["단가"]
        df["사업장비용"] = df["사업장(kg)"] * df["사업장단가"]
        df["재활용수익"] = df["재활용(kg)"] * df["재활용단가"]
        df["최종정산액"] = df["음식물비용"] + df["사업장비용"] - df["재활용수익"]
        df["월별"] = df["날짜"].astype(str).str[:7]
        df["년도"] = df["날짜"].astype(str).str[:4]
        # 탄소감축량 산정 (환경부 기준, 2024)
        # 음식물 퇴비화 처리: 매립 대비 0.3 kgCO₂eq/kg 감축 (환경부 온실가스 배출계수)
        # 재활용 처리: 소각 대비 0.4 kgCO₂eq/kg 감축 (한국환경공단 자원순환 가이드라인)
        CO2_FOOD    = 0.3   # kgCO₂/kg - 음식물 퇴비화 감축계수
        CO2_RECYCLE = 0.4   # kgCO₂/kg - 재활용품 감축계수
        df["탄소감축량(kg)"] = (df["음식물(kg)"] * CO2_FOOD) + (df["재활용(kg)"] * CO2_RECYCLE)
    return df

def save_collection(row: dict):
    conn = get_conn()
    conn.execute("""
        INSERT INTO collections (날짜,학교명,학생수,수거업체,음식물_kg,재활용_kg,사업장_kg,상태,현장사진)
        VALUES (:날짜,:학교명,:학생수,:수거업체,:음식물_kg,:재활용_kg,:사업장_kg,:상태,:현장사진)
    """, row)
    conn.commit()
    conn.close()

def update_collection_status(ids: list, new_status: str):
    """수거 레코드 상태 일괄 변경"""
    conn = get_conn()
    conn.executemany(
        "UPDATE collections SET 상태=? WHERE id=?",
        [(new_status, i) for i in ids]
    )
    conn.commit()
    conn.close()

def get_school_prices():
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM school_prices ORDER BY 학교명", conn)
    conn.close()
    return df

def update_school_price(school, food, recycle, biz, name, tel, email):
    conn = get_conn()
    conn.execute("""
        UPDATE school_prices
        SET 음식물단가=?,재활용단가=?,사업장단가=?,담당자명=?,담당자연락처=?,담당자이메일=?,updated_at=?
        WHERE 학교명=?
    """, (food, recycle, biz, name, tel, email, datetime.now().strftime("%Y-%m-%d"), school))
    conn.commit()
    conn.close()

# [추가2] 재활용 시세 함수
def get_recycle_prices():
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM recycle_prices ORDER BY 품목명", conn)
    conn.close()
    return df

def update_recycle_price(item, price):
    conn = get_conn()
    conn.execute(
        "UPDATE recycle_prices SET 단가=?, updated_at=? WHERE 품목명=?",
        (price, datetime.now().strftime("%Y-%m-%d"), item)
    )
    conn.commit()
    conn.close()

def get_avg_recycle_price():
    """전체 재활용품 가중 평균 단가"""
    conn = get_conn()
    row = conn.execute("SELECT AVG(단가) FROM recycle_prices").fetchone()
    conn.close()
    return int(row[0]) if row and row[0] else 300

# [추가4] 일정 함수
def get_schedules_month(year, month):
    conn = get_conn()
    prefix = f"{year}-{month:02d}"
    df = pd.read_sql_query(
        "SELECT * FROM schedules WHERE 날짜 LIKE ? ORDER BY 날짜, 학교명",
        conn, params=(f"{prefix}%",)
    )
    conn.close()
    return df

def add_schedule(date_str, school, memo):
    conn = get_conn()
    conn.execute(
        "INSERT INTO schedules (날짜,학교명,메모,완료여부) VALUES (?,?,?,0)",
        (date_str, school, memo)
    )
    conn.commit()
    conn.close()

def toggle_schedule(sid, current):
    conn = get_conn()
    conn.execute("UPDATE schedules SET 완료여부=? WHERE id=?", (0 if current else 1, sid))
    conn.commit()
    conn.close()

def delete_schedule(sid):
    conn = get_conn()
    conn.execute("DELETE FROM schedules WHERE id=?", (sid,))
    conn.commit()
    conn.close()

# [추가3] 스쿨존 시간 체크
def is_schoolzone_restricted():
    h = datetime.now().hour
    for start, end in SCHOOLZONE_RESTRICTED:
        if start <= h < end:
            return True, f"{start}:00~{end}:00"
    return False, ""

# 엑셀 생성
def create_secure_excel(df, title):
    if not EXCEL_PASSWORD:
        st.warning("⚠️ .env 파일에 EXCEL_PASSWORD가 설정되지 않아 시트 보호가 적용되지 않습니다.")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="법정실적보고서", startrow=2)
        wb = writer.book
        ws = writer.sheets["법정실적보고서"]
        fmt = wb.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter"})
        ws.merge_range(0, 0, 1, len(df.columns)-1, f"■ {title} ■", fmt)
        for i in range(len(df.columns)):
            ws.set_column(i, i, 16)
        if EXCEL_PASSWORD:
            ws.protect(EXCEL_PASSWORD, {"objects":True,"scenarios":True,"format_cells":False,"sort":True})
    return output.getvalue()

# 알림톡 (시뮬레이션만)
def send_kakao_alimtalk(phone, school, food_kg, total_price):
    st.info(f"📱 [알림톡]\n▸ 수신: {school} 담당자 ({phone})\n▸ 내용: 음식물 {food_kg:,.0f}kg 수거 완료, 청구 예정액 {total_price:,}원")
    return True


# ============================================================
# [섹션 E] 계약서류 패키지 자동 생성 시스템
# ============================================================

# ── 하영자원 고정 정보 (전 서류 공통 사용) ─────────────────
HY = {
    "name":       "하영자원",
    "ceo":        "정석완",
    "biz_no":     "405-11-42991",
    "permit_no":  "제20-35호",
    "address":    "경기도 화성시 남양읍 남양성지로 219, 2층",
    "tel":        "031-414-3713",
    "mobile":     "010-3114-4030",
    "fax":        "031-356-3713",
    "email":      "hyrecycling@naver.com",
    "bank":       "기업은행",
    "account":    "450-092046-01-017",
    "biz_type":   "폐기물처리",
    "biz_item":   "지정외폐기물수집,운반업",
    "processor":  "주식회사 청명",   # 처리업체
}

# 서류 유효기간 마스터 (D-day 알림용)
DOC_EXPIRE = [
    {"name": "소상공인 확인서",        "expire": "2026-03-31", "renew_url": "sminfo.mss.go.kr"},
    {"name": "창업기업 확인서",         "expire": "2027-01-07", "renew_url": "중소벤처기업부"},
    {"name": "재해율 확인서",           "expire": "2027-02-19", "renew_url": "안전보건공단"},
    {"name": "사업자등록증",            "expire": "9999-12-31", "renew_url": "국세청"},
    {"name": "폐기물수집운반업 허가증", "expire": "9999-12-31", "renew_url": "화성시청"},
]


# ── [섹션C] contract_docs DB 헬퍼 함수 ─────────────────────

def c_get_all_docs() -> list[dict]:
    """전체 서류 목록 + D-day 계산해서 반환"""
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, doc_name, issued_date, expire_date, renew_url, "
        "file_note, renewed, memo FROM contract_docs ORDER BY expire_date ASC"
    ).fetchall()
    today_dt = date.today()
    result = []
    for r in rows:
        exp_str = r[3]
        if exp_str == "9999-12-31":
            dday, status = "무기한", "🟢 정상"
        else:
            exp     = date.fromisoformat(exp_str)
            diff    = (exp - today_dt).days
            dday    = f"D-{diff}" if diff >= 0 else f"만료 +{abs(diff)}일"
            if diff < 0:     status = "⛔ 만료됨"
            elif diff <= 30: status = "🔴 만료임박"
            elif diff <= 60: status = "🟡 주의"
            else:            status = "🟢 정상"
        result.append({
            "id": r[0], "서류명": r[1], "발급일": r[2],
            "만료일": exp_str if exp_str != "9999-12-31" else "무기한",
            "갱신처": r[4], "비고": r[5],
            "갱신완료": bool(r[6]), "메모": r[7],
            "D-day": dday, "상태": status,
        })
    # 우선순위 정렬: 만료/임박 → 주의 → 정상
    _sort_key = {"⛔ 만료됨": 0, "🔴 만료임박": 1, "🟡 주의": 2, "🟢 정상": 3}
    result.sort(key=lambda x: _sort_key.get(x["상태"], 9))
    return result


def c_update_doc(doc_id: int, issued: str, expire: str,
                 renew_url: str, file_note: str,
                 renewed: bool, memo: str):
    conn = get_conn()
    conn.execute(
        """UPDATE contract_docs SET issued_date=?, expire_date=?,
           renew_url=?, file_note=?, renewed=?, memo=?, updated_at=?
           WHERE id=?""",
        (issued, expire, renew_url, file_note, int(renewed), memo,
         datetime.now().strftime("%Y-%m-%d"), doc_id)
    )
    conn.commit()


def c_add_doc(doc_name: str, issued: str, expire: str,
              renew_url: str, file_note: str, memo: str):
    conn = get_conn()
    conn.execute(
        """INSERT INTO contract_docs
           (doc_name, issued_date, expire_date, renew_url, file_note, memo, updated_at)
           VALUES (?,?,?,?,?,?,?)""",
        (doc_name, issued, expire, renew_url, file_note, memo,
         datetime.now().strftime("%Y-%m-%d"))
    )
    conn.commit()


def c_delete_doc(doc_id: int):
    conn = get_conn()
    conn.execute("DELETE FROM contract_docs WHERE id=?", (doc_id,))
    conn.commit()


def c_toggle_renewed(doc_id: int):
    conn = get_conn()
    conn.execute(
        "UPDATE contract_docs SET renewed = CASE WHEN renewed=1 THEN 0 ELSE 1 END "
        "WHERE id=?", (doc_id,)
    )
    conn.commit()


# ── [섹션A] 학교 마스터 + 계약 이력 헬퍼 함수 ──────────────

def a_get_all_schools() -> list[dict]:
    """school_prices 전체 + 계약 만료 D-day 계산"""
    conn  = get_conn()
    today = date.today()
    rows  = conn.execute(
        """SELECT 학교명, 음식물단가, 교육청,
                  학교_사업자번호, 학교_주소, 학교_전화,
                  계약_시작일, 계약_종료일, 계약_상태, 비고, updated_at,
                  담당자명, 담당자연락처, 담당자이메일
           FROM school_prices ORDER BY 교육청, 학교명"""
    ).fetchall()
    result = []
    for r in rows:
        end = r[7] or ""
        if end and end != "9999-12-31":
            try:
                diff = (date.fromisoformat(end) - today).days
                if diff < 0:       contract_dday = f"⛔ 만료 +{abs(diff)}일"
                elif diff <= 30:   contract_dday = f"🔴 D-{diff}"
                elif diff <= 90:   contract_dday = f"🟡 D-{diff}"
                else:              contract_dday = f"🟢 D-{diff}"
            except Exception:
                contract_dday = end
        elif end == "9999-12-31":
            contract_dday = "무기한"
        else:
            contract_dday = "미설정"
        result.append({
            "학교명":       r[0],
            "음식물단가":   r[1],
            "교육청":       r[2] or "",
            "사업자번호":   r[3] or "",
            "주소":         r[4] or "",
            "전화":         r[5] or "",
            "계약시작":     r[6] or "",
            "계약종료":     r[7] or "",
            "계약상태":     r[8] or "미계약",
            "비고":         r[9] or "",
            "수정일":       r[10] or "",
            "계약D-day":   contract_dday,
            "담당자명":     r[11] or "",
            "담당자연락처": r[12] or "",
            "담당자이메일": r[13] or "",
        })
    return result


def a_update_school(학교명: str, 단가: int, 사업자번호: str,
                    주소: str, 전화: str, 시작일: str,
                    종료일: str, 상태: str, 비고: str,
                    담당자명: str, 담당자연락처: str, 담당자이메일: str):
    conn = get_conn()
    conn.execute(
        """UPDATE school_prices SET
           음식물단가=?, 학교_사업자번호=?, 학교_주소=?, 학교_전화=?,
           계약_시작일=?, 계약_종료일=?, 계약_상태=?,
           비고=?, 담당자명=?, 담당자연락처=?, 담당자이메일=?, updated_at=?
           WHERE 학교명=?""",
        (단가, 사업자번호, 주소, 전화, 시작일, 종료일, 상태,
         비고, 담당자명, 담당자연락처, 담당자이메일,
         datetime.now().strftime("%Y-%m-%d"), 학교명)
    )
    conn.commit()


def a_add_school(학교명: str, 교육청: str, 단가: int,
                 사업자번호: str, 주소: str, 전화: str,
                 시작일: str, 종료일: str, 상태: str, 비고: str):
    conn = get_conn()
    conn.execute(
        """INSERT OR IGNORE INTO school_prices
           (학교명, 교육청, 음식물단가, 학교_사업자번호,
            학교_주소, 학교_전화, 계약_시작일, 계약_종료일,
            계약_상태, 비고, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (학교명, 교육청, 단가, 사업자번호, 주소, 전화,
         시작일, 종료일, 상태, 비고,
         datetime.now().strftime("%Y-%m-%d"))
    )
    conn.commit()


def a_get_contracts(학교명: str = None) -> list[dict]:
    """계약 이력 목록 반환"""
    conn  = get_conn()
    today = date.today()
    query = "SELECT * FROM contract_master"
    params: tuple = ()
    if 학교명:
        query  += " WHERE 학교명=?"
        params  = (학교명,)
    query += " ORDER BY 계약_시작일 DESC"
    rows = conn.execute(query, params).fetchall()
    cols = ["id","학교명","계약번호","계약_시작일","계약_종료일",
            "폐기물_종류","단가","월_예상량_L","계약_상태",
            "갱신_알림일","나라장터_번호","비고","created_at","updated_at"]
    result = []
    for r in rows:
        d = dict(zip(cols, r))
        end = d.get("계약_종료일","")
        if end:
            try:
                diff = (date.fromisoformat(end) - today).days
                d["D-day"] = (f"D-{diff}" if diff >= 0 else f"만료 +{abs(diff)}일")
            except Exception:
                d["D-day"] = ""
        result.append(d)
    return result


def a_add_contract(학교명: str, 계약번호: str, 시작일: str,
                   종료일: str, 폐기물종류: str, 단가: int,
                   월예상량: float, 상태: str,
                   나라장터번호: str, 비고: str):
    conn = get_conn()
    conn.execute(
        """INSERT INTO contract_master
           (학교명, 계약번호, 계약_시작일, 계약_종료일,
            폐기물_종류, 단가, 월_예상량_L, 계약_상태,
            나라장터_번호, 비고, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (학교명, 계약번호, 시작일, 종료일, 폐기물종류,
         단가, 월예상량, 상태, 나라장터번호, 비고,
         datetime.now().strftime("%Y-%m-%d"),
         datetime.now().strftime("%Y-%m-%d"))
    )
    # school_prices 계약 정보도 동기화
    conn.execute(
        """UPDATE school_prices SET
           계약_시작일=?, 계약_종료일=?, 계약_상태=?, updated_at=?
           WHERE 학교명=?""",
        (시작일, 종료일, 상태, datetime.now().strftime("%Y-%m-%d"), 학교명)
    )
    conn.commit()


def a_delete_contract(contract_id: int):
    conn = get_conn()
    conn.execute("DELETE FROM contract_master WHERE id=?", (contract_id,))
    conn.commit()


# ── [5단계] 나라장터 계약번호 파서 + 전자문서 헬퍼 ───────────

def parse_g2b_contract_no(contract_no: str) -> dict:
    """
    나라장터 계약번호 파싱
    예시: R26TA01543339 00
    """
    s = contract_no.strip().replace(" ", "")
    if len(s) < 12:
        return {"원본": contract_no, "오류": "형식 불일치 (최소 12자)"}

    type_map  = {"R": "용역", "S": "물품", "C": "공사", "G": "기타"}
    order_map = {"00": "원계약"}

    contract_type = type_map.get(s[0], f"기타({s[0]})")
    year          = "20" + s[1:3]
    org_code      = s[3:5]
    serial        = s[5:13]
    order_code    = s[13:15] if len(s) >= 15 else "00"
    order_label   = order_map.get(order_code, f"{int(order_code)}차 변경계약")
    is_change     = order_code != "00"

    return {
        "원본":         contract_no,
        "계약유형":     contract_type,
        "계약연도":     year,
        "기관코드":     org_code,
        "일련번호":     serial,
        "차수코드":     order_code,
        "차수":         order_label,
        "변경계약여부": is_change,
    }


def g2b_update_edoc(contract_id: int, edoc_no: str, sender_id: str,
                    recv_date: str, edoc_status: str):
    """전자문서 정보 업데이트"""
    conn = get_conn()
    conn.execute(
        """UPDATE contract_master SET
           전자문서번호=?, 송신자ID=?, 수신일자=?, 계약서_상태=?, updated_at=?
           WHERE id=?""",
        (edoc_no, sender_id, recv_date, edoc_status,
         datetime.now().strftime("%Y-%m-%d"), contract_id)
    )
    # 서명완료 시 school_prices 계약상태 → 계약중 자동 변경
    if edoc_status == "서명완료":
        row = conn.execute(
            "SELECT 학교명, 계약_시작일, 계약_종료일 FROM contract_master WHERE id=?",
            (contract_id,)
        ).fetchone()
        if row:
            conn.execute(
                """UPDATE school_prices SET 계약_상태=?, updated_at=?
                   WHERE 학교명=?""",
                ("계약중", datetime.now().strftime("%Y-%m-%d"), row[0])
            )
    conn.commit()


def g2b_get_edoc_list() -> list[dict]:
    """전자문서 수신 현황 전체 조회"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT id, 학교명, 나라장터_번호, 전자문서번호,
                  송신자ID, 수신일자, 계약서_상태, 계약_시작일, 계약_종료일
           FROM contract_master
           ORDER BY 수신일자 DESC, id DESC"""
    ).fetchall()
    cols = ["id","학교명","나라장터_번호","전자문서번호",
            "송신자ID","수신일자","계약서_상태","계약_시작일","계약_종료일"]
    return [dict(zip(cols, r)) for r in rows]

# ── [신규 1단계] 학교 응답(승인/반려) 처리 헬퍼 ────────────────────
def school_reply_contract(contract_id: int, new_status: str, reply_memo: str = ""):
    """학교(행정실)에서 하영자원의 계약 초안에 응답(승인/수정요청)하는 기능"""
    conn = get_conn()
    row = conn.execute("SELECT 비고 FROM contract_master WHERE id=?", (contract_id,)).fetchone()
    current_memo = row[0] if row and row[0] else ""

    # 반려(수정 요청)일 경우 사유를 비고란에 추가 기록
    if reply_memo:
        current_memo = f"{current_memo} | [학교응답]: {reply_memo}"

    conn.execute(
        """UPDATE contract_master SET
           계약서_상태=?, 비고=?, updated_at=?
           WHERE id=?""",
        (new_status, current_memo, datetime.now().strftime("%Y-%m-%d"), contract_id)
    )
    
    # 학교가 '서명완료(승인)'를 누르면, 학교 마스터 정보의 상태도 '계약중'으로 자동 변경
    if new_status == "서명완료":
        c_row = conn.execute("SELECT 학교명 FROM contract_master WHERE id=?", (contract_id,)).fetchone()
        if c_row:
            conn.execute(
                "UPDATE school_prices SET 계약_상태='계약중', updated_at=? WHERE 학교명=?",
                (datetime.now().strftime("%Y-%m-%d"), c_row[0])
            )
    conn.commit()


# ── [6단계] 담당자 고도화 + 워크플로우 헬퍼 ────────────────────

def a_update_contacts(학교명: str, 담당자2명: str, 담당자2연락처: str,
                      행정실팩스: str, 계약_사인_여부: bool):
    """부담당자 + 팩스 + 사인여부 업데이트"""
    conn = get_conn()
    conn.execute(
        """UPDATE school_prices SET
           담당자2명=?, 담당자2연락처=?, 행정실팩스=?,
           계약_사인_여부=?, updated_at=?
           WHERE 학교명=?""",
        (담당자2명, 담당자2연락처, 행정실팩스,
         int(계약_사인_여부), datetime.now().strftime("%Y-%m-%d"), 학교명)
    )
    conn.commit()


def wf_get_steps(contract_id: int) -> dict:
    """계약 체결 워크플로우 5단계 현황 조회
    외부 스크립트·코워크 자동화에서 단계별 상태를 딕셔너리로 조회할 때 사용.
    UI에서는 직접 SQL 조회 방식을 사용함.
    """
    conn = get_conn()
    row  = conn.execute(
        """SELECT 단계1_완료, 단계2_완료, 단계3_완료,
                  단계4_완료, 단계5_완료, 학교명, 계약_상태
           FROM contract_master WHERE id=?""",
        (contract_id,)
    ).fetchone()
    if not row:
        return {}
    return {
        "steps":     [bool(row[i]) for i in range(5)],
        "학교명":    row[5],
        "계약_상태": row[6],
        "완료수":    sum(bool(row[i]) for i in range(5)),
    }


def wf_toggle_step(contract_id: int, step: int):
    """워크플로우 단계 토글 (step: 1~5)"""
    col  = f"단계{step}_완료"
    conn = get_conn()
    conn.execute(
        f"UPDATE contract_master SET {col} = CASE WHEN {col}=1 THEN 0 ELSE 1 END,"
        f" updated_at=? WHERE id=?",
        (datetime.now().strftime("%Y-%m-%d"), contract_id)
    )
    # 5단계 모두 완료 시 계약_상태 자동 → 계약중
    steps = conn.execute(
        "SELECT 단계1_완료,단계2_완료,단계3_완료,단계4_완료,단계5_완료,학교명 "
        "FROM contract_master WHERE id=?", (contract_id,)
    ).fetchone()
    if steps and all(steps[:5]):
        conn.execute(
            "UPDATE contract_master SET 계약_상태='계약중', updated_at=? WHERE id=?",
            (datetime.now().strftime("%Y-%m-%d"), contract_id)
        )
        conn.execute(
            "UPDATE school_prices SET 계약_상태='계약중', updated_at=? WHERE 학교명=?",
            (datetime.now().strftime("%Y-%m-%d"), steps[5])
        )
    conn.commit()


def get_all_contact_csv() -> bytes:
    """전체 담당자 명부 CSV 생성"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT 학교명, 교육청, 담당자명, 담당자연락처, 담당자이메일,
                  담당자2명, 담당자2연락처, 행정실팩스, 계약_종료일
           FROM school_prices ORDER BY 교육청, 학교명"""
    ).fetchall()
    cols = ["학교명","교육청","주담당자명","주담당자연락처","주담당자이메일",
            "부담당자명","부담당자연락처","행정실팩스","계약만료일"]
    buf = io.StringIO()
    import csv
    writer = csv.writer(buf)
    writer.writerow(cols)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


# ── [섹션4] 월별 정산 자동 계산 헬퍼 ──────────────────────────

def calc_monthly_settlement(year: int, month: int) -> dict:
    """
    학교별 월별 정산 자동 계산
    반환: {학교명: {수거량_L, 단가, 정산금액, 세금구분, 수거건수}}
    """
    conn     = get_conn()
    ym_str   = f"{year}-{month:02d}"   # ex) "2026-03"

    # 해당 월 수거 데이터 조회 (음식물만)
    rows = conn.execute(
        """SELECT c.학교명,
                  SUM(c.음식물_kg)  AS 수거량,
                  COUNT(*)          AS 건수,
                  p.음식물단가
           FROM collections c
           LEFT JOIN school_prices p ON c.학교명 = p.학교명
           WHERE substr(c.날짜, 1, 7) = ?
           GROUP BY c.학교명
           ORDER BY c.학교명""",
        (ym_str,)
    ).fetchall()

    result = {}
    for 학교명, 수거량, 건수, 단가 in rows:
        단가      = 단가 or 150          # 단가 미설정 시 기본값
        수거량    = round(수거량 or 0, 1)
        정산금액  = int(수거량 * 단가)
        result[학교명] = {
            "수거량_L":  수거량,
            "단가":      단가,
            "정산금액":  정산금액,
            "세금구분":  "면세",          # 음식물폐기물 = 면세
            "수거건수":  건수,
        }
    return result


def create_settlement_excel(settlement: dict, year: int, month: int) -> bytes:
    """
    월별 정산표 엑셀 생성
    반환: bytes (다운로드용)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month:02d}월_정산"

    # ── 스타일 정의 ──
    hdr_fill  = PatternFill("solid", fgValue="1a73e8")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    tot_fill  = PatternFill("solid", fgValue="e8f0fe")
    tot_font  = Font(bold=True, size=11)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    right_a   = Alignment(horizontal="right",  vertical="center")

    # ── 제목 ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"하영자원 {year}년 {month:02d}월 음식물 정산표"
    ws["A1"].font      = Font(bold=True, size=14, color="1a3a5c")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"작성일: {date.today().strftime('%Y년 %m월 %d일')}"
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws["F2"] = "하영자원 / 대표 정석완"
    ws["F2"].font = Font(bold=True, size=10)

    # ── 헤더 행 ──
    headers = ["No", "학교명", "수거건수", "수거량(L)", "단가(원/L)", "정산금액(원)", "세금구분"]
    col_widths = [5, 22, 10, 12, 12, 16, 10]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell            = ws.cell(row=3, column=i, value=h)
        cell.fill       = hdr_fill
        cell.font       = hdr_font
        cell.alignment  = center
        cell.border     = border
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 데이터 행 ──
    total_vol  = 0.0
    total_amt  = 0
    total_cnt  = 0

    for seq, (school, data) in enumerate(sorted(settlement.items()), 1):
        row = seq + 3
        vals = [
            seq,
            school,
            data["수거건수"],
            data["수거량_L"],
            data["단가"],
            data["정산금액"],
            data["세금구분"],
        ]
        for col, val in enumerate(vals, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = right_a if col in (3,4,5,6) else center if col in (1,7) else cell.alignment
        # 금액 숫자 서식
        ws.cell(row=row, column=6).number_format = "#,##0"
        ws.cell(row=row, column=4).number_format = "#,##0.0"

        total_vol += data["수거량_L"]
        total_amt += data["정산금액"]
        total_cnt += data["수거건수"]

    # ── 합계 행 ──
    tot_row = len(settlement) + 4
    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    ws[f"A{tot_row}"] = "합 계"
    for col in range(1, 8):
        cell        = ws.cell(row=tot_row, column=col)
        cell.fill   = tot_fill
        cell.font   = tot_font
        cell.border = border
        cell.alignment = center

    ws.cell(row=tot_row, column=3, value=total_cnt).alignment = right_a
    ws.cell(row=tot_row, column=4, value=round(total_vol, 1)).number_format = "#,##0.0"
    ws.cell(row=tot_row, column=4).alignment = right_a
    ws.cell(row=tot_row, column=6, value=total_amt).number_format = "#,##0"
    ws.cell(row=tot_row, column=6).alignment = right_a

    # ── 안내 ──
    note_row = tot_row + 2
    ws[f"A{note_row}"] = "※ 음식물류폐기물 수거운반은 부가가치세 면세 품목입니다."
    ws[f"A{note_row}"].font = Font(italic=True, color="888888", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hy_font() -> str:
    """한글 TTF 폰트 등록 — 나눔고딕(Cloud) → 맑은고딕(Windows) 순"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # ⚠️ NotoSansCJK(.ttc/.otf)는 PostScript 아웃라인이라 reportlab 로드 불가
    # Streamlit Cloud: packages.txt에 fonts-nanum 추가로 나눔고딕 TTF 설치
    candidates = [
        # Streamlit Cloud / Ubuntu (fonts-nanum 패키지 설치 후)
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothicRegular.ttf"),
        # fonts-nanum-extra 경로
        ("NanumBarunGothic", "/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf"),
        # Windows
        ("MalgunGothic", "C:/Windows/Fonts/malgun.ttf"),
        ("NanumGothic",  "C:/Windows/Fonts/NanumGothic.ttf"),
        # macOS
        ("AppleGothic",  "/Library/Fonts/AppleGothic.ttf"),
        # 기타 Linux 경로
        ("NanumGothic",  "/usr/share/fonts/nanum/NanumGothic.ttf"),
    ]
    for fid, fpath in candidates:
        if os.path.exists(fpath):
            try:
                pdfmetrics.registerFont(TTFont(fid, fpath))
                return fid
            except Exception:
                continue
    return "Helvetica"  # 최후 fallback (한글 깨짐 — packages.txt 확인 필요)


def _hy_font_bold() -> str:
    """볼드 TTF 폰트 등록"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    candidates = [
        # Streamlit Cloud / Ubuntu
        ("NanumGothicBold", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ("NanumBarunGothicBold", "/usr/share/fonts/truetype/nanum/NanumBarunGothicBold.ttf"),
        # Windows
        ("MalgunGothicBold", "C:/Windows/Fonts/malgunbd.ttf"),
        ("NanumGothicBold",  "C:/Windows/Fonts/NanumGothicBold.ttf"),
        ("NanumGothicBold",  "/usr/share/fonts/nanum/NanumGothicBold.ttf"),
    ]
    for fid, fpath in candidates:
        if os.path.exists(fpath):
            try:
                pdfmetrics.registerFont(TTFont(fid, fpath))
                return fid
            except Exception:
                continue
    return _hy_font()  # 볼드 없으면 일반으로 대체


def _out_dir(sub: str) -> str:
    """출력 폴더 생성 후 경로 반환"""
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base = os.getcwd()
    path = os.path.join(base, sub)
    os.makedirs(path, exist_ok=True)
    return path


# ── E-1: 음식물 견적서 PDF (원본 레이아웃 완전 재현) ──────────
def generate_estimate_pdf(school_name: str, school_biz_no: str,
                          volume_l: float, unit_price: int,
                          contract_period: str,
                          year: str = None) -> str:
    """음식물 견적서 PDF — 원본 PDF(음식물견적서.pdf) 레이아웃 100% 재현"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units    import mm
    from reportlab.lib          import colors
    from reportlab.platypus     import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph,
                                        Spacer, HRFlowable)
    from reportlab.lib.styles   import ParagraphStyle

    FONT  = _hy_font()
    FONTB = _hy_font_bold()
    today = date.today()
    yr    = year or str(today.year)[2:]
    supply_amount = int(volume_l * unit_price) if volume_l else 0

    out   = _out_dir("estimates_pdf")
    fname = f"음식물견적서_{school_name}_{today.strftime('%Y%m%d')}.pdf"
    fpath = os.path.join(out, fname)

    def ps(name, size, align=0, bold=False, color=colors.black, leading_mult=1.5):
        fn = FONTB if bold else FONT
        return ParagraphStyle(name, fontName=fn, fontSize=size,
                               alignment=align, leading=size * leading_mult,
                               textColor=color, spaceAfter=1)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                             leftMargin=18*mm, rightMargin=18*mm,
                             topMargin=18*mm, bottomMargin=15*mm)
    story = []
    W = A4[0] - 36*mm  # 유효 너비

    # ① 제목
    story.append(Paragraph("음식폐기물 처리비용 견적서",
                            ps("title", 20, align=1, bold=True)))
    story.append(Spacer(1, 6*mm))

    # ② 상단 2단 — 고객정보(좌) | 공급자정보(우)
    cell_s = ps("cell", 9)
    cell_b = ps("cellb", 9, bold=True)
    left_col = [
        [Paragraph("고  객  명", ps("lh", 9, align=1)),
         Paragraph(f"<b>{school_name}</b>", ps("lv", 12, align=1, bold=True))],
        [Paragraph("견  적  일", ps("lh", 9, align=1)),
         Paragraph(f"{yr}.{today.month:02d}.{today.day:02d}",
                   ps("lv2", 9, align=1))],
    ]
    left_tbl = Table(left_col, colWidths=[22*mm, 38*mm],
                     rowHeights=[13*mm, 10*mm])
    left_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f5f5f5")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",      (0,0), (0,-1),  "CENTER"),
    ]))

    right_data = [
        ["사업자등록번호", HY["biz_no"],   "허가번호", HY["permit_no"]],
        ["상  호",         HY["name"],      "대  표  자", HY["ceo"]],
        ["주  소",         HY["address"],   "",          ""],
        ["업  태",         HY["biz_type"],  "업  종",  HY["biz_item"]],
    ]
    right_tbl = Table(right_data, colWidths=[20*mm, 50*mm, 18*mm, 38*mm])
    right_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f0f0f0")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#f0f0f0")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",      (0,0), (0,-1), "CENTER"),
        ("ALIGN",      (2,0), (2,-1), "CENTER"),
        ("SPAN",       (1,2), (3,2)),
    ]))

    top_tbl = Table([[left_tbl, right_tbl]],
                    colWidths=[62*mm, W - 62*mm])
    top_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(top_tbl)
    story.append(Spacer(1, 3*mm))

    # ③ 연도 부제
    story.append(Paragraph(f"<b>{yr}년도 음식물류폐기물견적서</b>",
                            ps("sub", 10, align=1, bold=True)))
    story.append(Spacer(1, 3*mm))

    # ④ 공급가액 합계 행
    sum_w = [55*mm, W - 55*mm]
    sum_tbl = Table([
        [Paragraph("<b>공급가액 합계</b>", ps("sh", 10, bold=True)),
         Paragraph(f"<b>{supply_amount:,}</b>" if supply_amount else "",
                   ps("sv", 10, bold=True))]
    ], colWidths=sum_w, rowHeights=10*mm)
    sum_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("BACKGROUND", (0,0), (0,0), colors.HexColor("#d9e1f2")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",      (0,0), (0,0), "CENTER"),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 1*mm))

    # ⑤ 품목 테이블 헤더 + 데이터 행 12줄 + 합계
    h_style = TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1, 0), colors.HexColor("#e8eaf6")),
        ("BACKGROUND", (0,-1),(-1,-1), colors.HexColor("#f5f5f5")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.lightgrey),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ])
    col_w = [55*mm, 22*mm, 20*mm, 22*mm, 30*mm, W-149*mm]
    # 헤더
    header = [["품  명", "규  격", "수  량", "단가(원)", "공급가액", "비  고"]]
    # 데이터: 1번째 행만 채움
    data_row1 = ["음식폐기물수거운반처리",
                 "L(리터)",
                 f"{volume_l:,.0f}" if volume_l else "",
                 f"{unit_price:,}",
                 f"{supply_amount:,}" if supply_amount else f"{unit_price}",
                 "면세"]
    rows = header + [data_row1] + [["", "", "", "", "", ""]]*11
    # 합계 행
    rows.append([Paragraph("<b>합  계</b>", ps("tot", 9, bold=True, align=1)),
                 "", "",
                 Paragraph(f"<b>{supply_amount:,}</b>" if supply_amount else "",
                            ps("tv", 9, bold=True, align=1)),
                 Paragraph(f"<b>{supply_amount:,}</b>" if supply_amount else "",
                            ps("tv2", 9, bold=True, align=1)),
                 ""])
    item_tbl = Table(rows, colWidths=col_w, rowHeights=[8*mm] + [7*mm]*12 + [8*mm])
    item_tbl.setStyle(h_style)
    story.append(item_tbl)
    story.append(Spacer(1, 4*mm))

    # ⑥ 특기사항
    note_data = [
        [Paragraph("<b>특기사항</b>", ps("nk", 9, align=1, bold=True)),
         Paragraph(
             "1. 음식물쓰레기수거용기는 수집운반업체(하영자원)에서 부담한다.<br/>"
             "2. 음식물쓰레기수거 때에 배출자는 수집운반업체가 수거를 원활히 할 수 있게 해야한다.<br/>"
             "3. 천재지변(눈,비)으로 인하여 수거를 할 수 없을 경우 수집운반업체는 배출자에게 지체없이 통보하고"
             " 수거 가능일자를 협의할 수 있다.",
             ps("nv", 8, leading_mult=1.6))]
    ]
    note_tbl = Table(note_data, colWidths=[20*mm, W-20*mm], rowHeights=22*mm)
    note_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "TOP"),
        ("ALIGN",      (0,0), (0,0),   "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",(1,0), (1,0),   4),
    ]))
    story.append(note_tbl)
    story.append(Spacer(1, 3*mm))

    # ⑦ 연락처 하단
    contact_data = [[
        "연락처", HY["tel"],
        "FAX",    HY["fax"],
        "이메일", HY["email"],
        "담당자", f"{HY['ceo']}\n{HY['mobile']}"
    ]]
    ct_w = [16*mm, 28*mm, 12*mm, 28*mm, 16*mm, 40*mm, 16*mm, W-156*mm]
    contact_tbl = Table(contact_data, colWidths=ct_w, rowHeights=12*mm)
    contact_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("BACKGROUND", (0,0), (0,0), colors.HexColor("#f0f0f0")),
        ("BACKGROUND", (2,0), (2,0), colors.HexColor("#f0f0f0")),
        ("BACKGROUND", (4,0), (4,0), colors.HexColor("#f0f0f0")),
        ("BACKGROUND", (6,0), (6,0), colors.HexColor("#f0f0f0")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.grey),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(contact_tbl)

    doc.build(story)
    return fpath


# ── E-2a: 음식물류폐기물 위수탁계약서 (HWP 원본 완전 재현) ───
def generate_contract_doc_pdf(school_name: str, school_biz_no: str,
                               school_addr: str, school_tel: str,
                               start_date: str, end_date: str,
                               waste_type: str = "음식물류폐기물",
                               volume_str: str = "",
                               unit_price: int = 180,
                               contract_amount: str = "") -> str:
    """폐기물 위수탁 운반 처리 계약서 — HWP 원본 구조 100% 재현"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units    import mm
    from reportlab.lib          import colors
    from reportlab.platypus     import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer,
                                        HRFlowable)
    from reportlab.lib.styles   import ParagraphStyle

    FONT  = _hy_font()
    FONTB = _hy_font_bold()
    today = date.today()
    out   = _out_dir("contract_pdf")
    fname = f"위수탁계약서_{school_name}_{today.strftime('%Y%m%d')}.pdf"
    fpath = os.path.join(out, fname)

    def ps(n, sz, align=0, bold=False, color=colors.black, leading_mult=1.6):
        fn = FONTB if bold else FONT
        return ParagraphStyle(n, fontName=fn, fontSize=sz,
                               alignment=align, leading=sz*leading_mult,
                               textColor=color, spaceAfter=1)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                             leftMargin=25*mm, rightMargin=25*mm,
                             topMargin=25*mm, bottomMargin=20*mm)
    story = []
    W = A4[0] - 50*mm

    # ① 제목
    story.append(Paragraph(
        "폐기물 위·수탁 운반·처리 계약서(안)",
        ps("t", 17, align=1, bold=True)))
    story.append(Spacer(1, 8*mm))

    # ② 계약 기본 항목 1~5
    yr = str(today.year)[2:]
    items_top = [
        ("1. 계    약    명", "음식물류폐기물수집,운반 처리"),
        ("2. 배  출  장  소", school_name),
        ("3. 처  리  장  소", HY["processor"]),
        ("4. 결  제  조  건", "계좌이체"),
        ("5. 위·수탁 계약기간",
         f"{start_date}부터  {end_date}까지"),
        ("6. 위·수탁 폐기물 및 처리금액", "(단위 : 원)"),
    ]
    for label, value in items_top:
        row = Table([[
            Paragraph(f"<b>{label}</b>", ps("lbl", 10, bold=True)),
            Paragraph(f": {value}", ps("val", 10))
        ]], colWidths=[55*mm, W - 55*mm])
        row.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), FONT),
            ("VALIGN",   (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 1),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
        ]))
        story.append(row)
        story.append(Spacer(1, 1.5*mm))

    # ③ 폐기물 처리금액 표
    story.append(Spacer(1, 3*mm))
    header = [["폐기물 종류", "단 위", "물 량\n(예상배출량)", "단   가",
               "계 약 금 액", "처 리 방 법",
               "운반비", "처리비"]]
    # 데이터 행
    data_rows = [
        [waste_type, "kg", volume_str, str(unit_price),
         contract_amount, "위탁", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
    ]
    total_row = [Paragraph("<b>총   계</b>",
                            ps("tot", 9, bold=True, align=1)),
                 "", "", "",
                 Paragraph(f"<b>{contract_amount}</b>",
                            ps("tv", 9, bold=True, align=1)),
                 "", "", ""]
    tbl_data = header + data_rows + [total_row]
    c_w = [28*mm, 14*mm, 28*mm, 20*mm, 28*mm, 22*mm, 15*mm, W-155*mm]
    waste_tbl = Table(tbl_data, colWidths=c_w,
                      rowHeights=[10*mm] + [8*mm]*3 + [9*mm])
    waste_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("BACKGROUND", (0,0), (-1, 0), colors.HexColor("#dce6f1")),
        ("BACKGROUND", (0,-1),(-1,-1), colors.HexColor("#f2f2f2")),
        ("BOX",        (0,0), (-1,-1), 0.6, colors.black),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(waste_tbl)
    story.append(Spacer(1, 5*mm))

    # ④ 계약 본문
    body_text = (
        "위 계약을 증명하기 위하여 계약서 2통을 작성하여 배출자와 수집·운반자가 서명(날인)"
        "한 후 각 1통씩 보관한다."
    )
    story.append(Paragraph(body_text, ps("body", 9)))
    story.append(Spacer(1, 2*mm))

    conditions = [
        "【계약 조건】",
        "① 수집ㆍ운반자는 계약기간 중 계약서에 명시된 폐기물을 적정처리 기준에 의거 수집ㆍ운반하여야 한다.",
        "② 수집ㆍ운반자는 배출자의 동의 없이 수집ㆍ운반업무를 제3자에게 위탁할 수 없다.",
        "③ 처리금액의 변동이 있을 때에는 상호 협의하여 계약을 변경할 수 있다.",
        "④ 계약을 해지하고자 할 때에는 상호 30일 전에 서면으로 통보하여야 한다.",
        "⑤ 이 계약에 명시되지 않은 사항은 관계법령 및 일반 관례에 따른다.",
    ]
    for c in conditions:
        story.append(Paragraph(c, ps("cond", 8.5,
                     bold=(c == "【계약 조건】"))))
        story.append(Spacer(1, 0.8*mm))

    story.append(Spacer(1, 6*mm))

    # ⑤ 계약일
    story.append(Paragraph(
        f"{today.year}년  {today.month:02d}월  {today.day:02d}일",
        ps("date", 11, align=1)))
    story.append(Spacer(1, 6*mm))

    # ⑥ 서명란 — 배출자(학교) | 수집·운반자(하영자원)
    sign_data = [
        ["구  분",  "배 출 자 (학  교)", "수집·운반자 (하영자원)"],
        ["기관명",  school_name,         HY["name"]],
        ["대표자",  "",                   HY["ceo"] + "  (인)"],
        ["사업자번호", school_biz_no,     HY["biz_no"]],
        ["주  소",  school_addr,          HY["address"]],
        ["연락처",  school_tel,           HY["tel"]],
    ]
    sign_tbl = Table(sign_data, colWidths=[30*mm, 70*mm, 60*mm])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1, 0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR",  (0,0), (-1, 0), colors.white),
        ("BACKGROUND", (0,1), (0,-1), colors.HexColor("#f0f4ff")),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.grey),
        ("ROWHEIGHT",  (0,0), (-1,-1), 9*mm),
        ("ROWHEIGHT",  (0,2), (-1, 2), 14*mm),
        ("FONTNAME",   (0,0), (-1, 0), FONTB),
    ]))
    story.append(sign_tbl)

    doc.build(story)
    return fpath


# ── E-2b: 계약이행 통합 서약서 (HWP 원본 완전 재현) ──────────
def generate_pledge_pdf(school_name: str, unit_price: int,
                        start_date: str, end_date: str) -> str:
    """계약이행 통합 서약서 — HWP 원본 7개 항목 100% 재현"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units    import mm
    from reportlab.lib          import colors
    from reportlab.platypus     import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles   import ParagraphStyle

    FONT  = _hy_font()
    FONTB = _hy_font_bold()
    today = date.today()
    yr    = str(today.year)[2:]
    out   = _out_dir("contract_pdf")
    fname = f"계약이행서약서_{school_name}_{today.strftime('%Y%m%d')}.pdf"
    fpath = os.path.join(out, fname)

    def ps(n, sz, align=0, bold=False, color=colors.black, leading_mult=1.6):
        fn = FONTB if bold else FONT
        return ParagraphStyle(n, fontName=fn, fontSize=sz,
                               alignment=align, leading=sz*leading_mult,
                               textColor=color, spaceAfter=1)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                             leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=15*mm)
    story = []
    W = A4[0] - 40*mm

    # ① 제목
    story.append(Paragraph("계약이행 통합 서약서",
                            ps("t", 17, align=1, bold=True)))
    story.append(Spacer(1, 5*mm))

    # ② 기본 정보 표
    info_rows = [
        ["계 약 명", f"{yr}년 음식물폐기물처리용역",
         "금  액",   f"{unit_price}원/L"],
        ["발주기관", school_name,
         "계약기간", f"{start_date} ~ {end_date}"],
        ["업 체 명", HY["name"],
         "대 표 자", HY["ceo"] + "  (인)"],
        ["사업자번호", HY["biz_no"],
         "연 락 처", HY["tel"]],
        ["주  소", HY["address"], "", ""],
    ]
    info_tbl = Table(info_rows, colWidths=[22*mm, 66*mm, 22*mm, 60*mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#dce6f1")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#dce6f1")),
        ("BOX",        (0,0), (-1,-1), 0.6, colors.black),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (0,-1), "CENTER"),
        ("ALIGN",      (2,0), (2,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("SPAN",       (1,4), (3,4)),
        ("ROWHEIGHT",  (0,0), (-1,-1), 9*mm),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 5*mm))

    # ③ 이행 내용 체크 표 — 원본 7개 항목 완전 재현
    chk_header = [["순", "구  분", "이 행 내 용 (세부내용)", "확인"]]

    chk_items = [
        ("1", "계약일반조건",
         "상기 본인(법인)은 「지방자치단체 입찰 및 계약 집행기준」 제9장\n계약 일반조건을 준수합니다.",
         "[✓] 예"),
        ("2", "수의계약\n각서",
         "귀 기관과 수의계약을 체결함에 있어서 [붙임1] 수의계약 배제사유 중 어느 사유에도\n"
         "해당되지 않으며 차후에 이러한 사실이 발견된 경우 계약의 해제·해지 및 부정당업자\n"
         "제재 처분을 받아도 하등의 이유를 제기하지 않겠습니다.\n"
         "[붙임1] 수의계약 배제사유 1부",
         "[✓] 예"),
        ("3", "수의계약\n체결 제한\n여부\n확인서",
         "①발주기관 소속 고위공직자, 배우자, 직계존속·비속에 해당하는가?\n"
         "  → [  ] 예  [  ] 아니오  [✓] 해당없음\n"
         "②계약 업무 담당 공직자, 배우자, 직계존속·비속에 해당하는가?\n"
         "  → [  ] 예  [  ] 아니오  [✓] 해당없음\n"
         "③감독기관 소속 고위공직자, 배우자, 직계존속·비속에 해당하는가?\n"
         "  → [  ] 예  [  ] 아니오  [✓] 해당없음\n"
         "④모회사 소속 고위공직자, 배우자, 직계존속·비속에 해당하는가?\n"
         "  → [  ] 예  [  ] 아니오  [✓] 해당없음\n"
         "⑤상임위원회 위원의 국회의원, 배우자, 직계존속에 해당하는가?\n"
         "  → [  ] 예  [  ] 아니오  [✓] 해당없음",
         "해당없음"),
        ("4", "청렴\n서약서",
         "계약의 체결·이행과정에서 금품·향응·편의 등을 제공하거나 요구하지 않을 것이며,\n"
         "이를 위반할 경우 발생하는 모든 법적 책임을 감수하겠습니다.",
         "[✓] 예"),
        ("5", "개인정보\n처리 동의",
         "입찰·계약 업무 처리를 위한 개인정보(성명, 사업자번호, 연락처 등)의 수집·이용에\n"
         "동의합니다.",
         "[✓] 동의"),
        ("6", "중대재해\n처벌법\n준수\n서약",
         "「중대재해 처벌 등에 관한 법률」 및 「산업안전보건법」에 따라 안전·보건\n"
         "조치를 성실히 이행할 것을 서약합니다.\n"
         "(붙임: 공사/용역 안전보건 점검표 제출 대상)",
         "[✓] 예"),
        ("7", "계약이행\n능력\n확인서",
         "상기 계약의 이행에 필요한 인력·장비·자격을 보유하고 있음을 확인하며,\n"
         "계약기간 내 성실하게 용역을 수행할 것을 서약합니다.\n"
         f"[폐기물수집운반업 허가번호: {HY['permit_no']}]",
         "[✓] 확인"),
    ]

    chk_rows = chk_header.copy()
    for num, cat, content, result in chk_items:
        chk_rows.append([
            Paragraph(num,     ps(f"n{num}", 8, align=1)),
            Paragraph(cat,     ps(f"c{num}", 8, align=1)),
            Paragraph(content.replace('\n', '<br/>'),
                      ps(f"d{num}", 7.5, leading_mult=1.55)),
            Paragraph(result,  ps(f"r{num}", 8, align=1)),
        ])

    # 행 높이: 헤더 8mm, 나머지 각 항목 가변
    row_heights = [8*mm, 12*mm, 22*mm, 38*mm, 12*mm, 10*mm, 16*mm, 16*mm]
    chk_tbl = Table(chk_rows,
                    colWidths=[10*mm, 22*mm, W - 52*mm, 20*mm],
                    rowHeights=row_heights)
    chk_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("BACKGROUND", (0,0), (-1, 0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR",  (0,0), (-1, 0), colors.white),
        ("FONTNAME",   (0,0), (-1, 0), FONTB),
        ("BOX",        (0,0), (-1,-1), 0.6, colors.black),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (0,-1), "CENTER"),
        ("ALIGN",      (1,0), (1,-1), "CENTER"),
        ("ALIGN",      (3,0), (3,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (2,1), (2,-1), 3),
        ("LEFTPADDING",(2,1), (2,-1), 4),
    ]))
    story.append(chk_tbl)
    story.append(Spacer(1, 8*mm))

    # ④ 날짜
    story.append(Paragraph(
        f"{today.year}년  {today.month:02d}월  {today.day:02d}일",
        ps("date", 10, align=1)))
    story.append(Spacer(1, 4*mm))

    # ⑤ 서명란
    sign_rows = [
        ["발주기관", school_name, "업체명", HY["name"]],
        ["기관장",   "",          "대표자", HY["ceo"] + "  (인)"],
    ]
    sign_tbl = Table(sign_rows, colWidths=[22*mm, 68*mm, 22*mm, 58*mm])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), FONT),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("BACKGROUND",(0,0),(0,-1), colors.HexColor("#eeeeee")),
        ("BACKGROUND",(2,0),(2,-1), colors.HexColor("#eeeeee")),
        ("BOX",       (0,0),(-1,-1), 0.5, colors.grey),
        ("INNERGRID", (0,0),(-1,-1), 0.3, colors.grey),
        ("ALIGN",     (0,0),(0,-1), "CENTER"),
        ("ALIGN",     (2,0),(2,-1), "CENTER"),
        ("VALIGN",    (0,0),(-1,-1), "MIDDLE"),
        ("ROWHEIGHT", (0,0),(-1,-1), 12*mm),
    ]))
    story.append(sign_tbl)
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        f"{school_name}장  귀중",
        ps("rcp", 10, align=2)))

    doc.build(story)
    return fpath


# ── E-2c: 중대재해 안전보건 점검표 PDF (신규 추가) ──────────
def generate_safety_check_pdf(school_name: str,
                               check_date: str = "") -> str:
    """공사(용역) 안전보건 점검표 — HWP 원본 7개 항목 완전 재현"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units    import mm
    from reportlab.lib          import colors
    from reportlab.platypus     import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles   import ParagraphStyle

    FONT  = _hy_font()
    FONTB = _hy_font_bold()
    today = date.today()
    chk_dt = check_date or f"{today.year}.   .   ."
    out    = _out_dir("contract_pdf")
    fname  = f"안전보건점검표_{school_name}_{today.strftime('%Y%m%d')}.pdf"
    fpath  = os.path.join(out, fname)

    def ps(n, sz, align=0, bold=False, color=colors.black):
        fn = FONTB if bold else FONT
        return ParagraphStyle(n, fontName=fn, fontSize=sz,
                               alignment=align, leading=sz*1.6,
                               textColor=color, spaceAfter=1)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                             leftMargin=18*mm, rightMargin=18*mm,
                             topMargin=18*mm, bottomMargin=15*mm)
    story = []
    W = A4[0] - 36*mm

    # ① 보관 안내
    story.append(Paragraph(
        "붙임4   공사(용역) 안전보건 점검표",
        ps("title", 14, align=1, bold=True)))
    story.append(Paragraph(
        "[학교(기관), 교육(지원)청에서 확인하여 자체 보관]",
        ps("sub", 9, align=1)))
    story.append(Spacer(1, 4*mm))

    # ② 기본 정보
    story.append(Paragraph(
        f"<b>공사(용역) 안전보건 점검표</b>",
        ps("hd", 11, bold=True)))
    info_rows = [
        [f"■ 기  관  명: {school_name}",
         f"■ 확  인  자: 주무관              (서명)",
         f"■ 점검일자: {chk_dt}"],
    ]
    info_tbl = Table(info_rows, colWidths=[W/3, W/3, W/3])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), FONT),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("VALIGN",   (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0),(-1,-1), 2),
    ]))
    story.append(info_tbl)

    # ③ 결재란
    sign_rows = [["결재", "담  당", "행정실장", "학 교 장"],
                 ["", "", "", ""]]
    sign_tbl = Table(sign_rows, colWidths=[18*mm, 28*mm, 28*mm, 28*mm],
                     rowHeights=[7*mm, 14*mm])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.black),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f5f5f5")),
    ]))
    story.append(Spacer(1, 2*mm))
    story.append(sign_tbl)
    story.append(Spacer(1, 4*mm))

    # ④ 점검 항목 7개
    chk_header = [["번호", "점  검  내  용", "확인결과\n예", "아니오"]]
    check_items = [
        ("1",
         "과업지시서(또는 특수조건) 또는 계약서에 '안전관리 및 예방조치 후 작업' 실시 내용\n"
         "포함하였는지 확인  ※계약서가 없을 경우 본 점검표로 갈음",
         "✓", ""),
        ("2",
         "공사(용역)업체에서 근로자에 대한 안전보건교육 실시하였는지 확인",
         "✓", ""),
        ("3",
         "공사(용역)업체에 안전보호구(안전모, 안전대, 안전화 등)를\n"
         "착용하고 작업하도록 주지하였는지 확인 (필요한 경우)",
         "✓", ""),
        ("4",
         "공사(용역)업체에 위험사항(위험성평가 등)과 기계·기구·설비의 안전점검에 관한 사항,\n"
         "공사 전 유의사항에 대해 안내하여 주었는지 확인",
         "✓", ""),
        ("5",
         "공사(용역)업체에 최초 학교(기관)의 현장(업체 근로자가 작업하는 공간)으로 이동할\n"
         "때나 현장 이외 장소 이동 시 교육행정실(담당자)의 안내를 받도록 주지시켰는지 확인",
         "✓", ""),
        ("6",
         "고소, 전기, 화기, 밀폐공간 등의 작업 시 공사(용역)업체에서 (붙임4-2~5)\n"
         "유해·위험 작업 시 안전보건 점검표를 제출하였는지 여부",
         "✓", ""),
        ("7",
         "안전·보건에 관한 종사자의 의견청취를 하였는지 여부\n"
         "- 의견제시사항: 없음",
         "✓", ""),
    ]
    rows = chk_header.copy()
    for num, content, yes, no in check_items:
        rows.append([
            Paragraph(num, ps(f"n{num}", 9, align=1)),
            Paragraph(content.replace('\n', '<br/>'), ps(f"c{num}", 8.5)),
            Paragraph(yes, ps(f"y{num}", 10, align=1, bold=True)),
            Paragraph(no,  ps(f"no{num}", 10, align=1)),
        ])
    row_h = [8*mm, 16*mm, 8*mm, 12*mm, 8*mm, 12*mm, 12*mm, 12*mm]
    chk_tbl = Table(rows, colWidths=[12*mm, W-42*mm, 15*mm, 15*mm],
                    rowHeights=row_h)
    chk_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("BACKGROUND", (0,0), (-1, 0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR",  (0,0), (-1, 0), colors.white),
        ("FONTNAME",   (0,0), (-1, 0), FONTB),
        ("BOX",        (0,0), (-1,-1), 0.6, colors.black),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0), (0,-1), "CENTER"),
        ("ALIGN",      (2,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (1,1), (1,-1), 3),
        ("LEFTPADDING",(1,1), (1,-1), 4),
    ]))
    story.append(chk_tbl)
    story.append(Spacer(1, 5*mm))

    # ⑤ 공사업체 확인서
    story.append(Paragraph("공사업체 확인서", ps("cfm", 11, bold=True)))
    story.append(Spacer(1, 2*mm))
    confirm_text = (
        "위 점검사항에 대해 안내를 받았으며 산업안전보건법규에 따라 작업자에게 "
        "안전보건보호구 지급 및 안전수칙을 준수하여 작업할 것을 확인합니다."
    )
    story.append(Paragraph(confirm_text, ps("ct", 9)))
    story.append(Spacer(1, 4*mm))

    cfm_rows = [
        ["소속(회사)", HY["name"],    "공사(용역)업체 책임자", HY["ceo"]],
        ["",           "",             "서  명",                "(서명)"],
    ]
    cfm_tbl = Table(cfm_rows, colWidths=[24*mm, 50*mm, 38*mm, 58*mm],
                    rowHeights=[9*mm, 14*mm])
    cfm_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (-1,-1), FONT),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0),(0,-1), colors.HexColor("#eeeeee")),
        ("BACKGROUND", (2,0),(2,-1), colors.HexColor("#eeeeee")),
        ("BOX",        (0,0),(-1,-1), 0.5, colors.black),
        ("INNERGRID",  (0,0),(-1,-1), 0.3, colors.grey),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("SPAN",       (0,0),(0,1)),
        ("SPAN",       (1,0),(1,1)),
    ]))
    story.append(cfm_tbl)
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "※ (작성대상) 금액에 상관없이 1회성 소규모 수선 등 모든 공사, 용역(각종 유지·보수 용역 포함)",
        ps("note", 7.5, color=colors.grey)))

    doc.build(story)
    return fpath


# ── E-3: 계약서류 패키지 ZIP 생성 ──────────────────────────
def generate_contract_package(
    school_name: str,
    school_biz_no: str   = "",
    school_addr: str     = "",
    school_tel: str      = "",
    start_date: str      = "",
    end_date: str        = "",
    volume_l: float      = 0,
    unit_price: int      = 180,
    contract_amount: str = "",
) -> bytes:
    """
    학교명 입력 → PDF 3종 자동생성 + 기존 서류 포함 → ZIP 반환 (bytes)
    포함 서류:
      자동생성: 음식물견적서, 위수탁계약서, 계약이행서약서
      기존파일: 사업자등록증, 허가증, 소상공인, 창업기업, 재해율확인서, 계좌
    """
    contract_period = (f"{start_date} ~ {end_date}"
                       if start_date and end_date else "계약기간 기재")

    generated = {}
    errors    = {}

    # ① PDF 3종 생성
    try:
        generated["견적서"] = generate_estimate_pdf(
            school_name, school_biz_no,
            volume_l, unit_price, contract_period
        )
    except Exception as e:
        errors["견적서"] = str(e)

    try:
        generated["위수탁계약서"] = generate_contract_doc_pdf(
            school_name, school_biz_no, school_addr, school_tel,
            start_date, end_date,
            volume_str=f"{volume_l:,.0f}L" if volume_l else "",
            unit_price=unit_price,
            contract_amount=contract_amount,
        )
    except Exception as e:
        errors["위수탁계약서"] = str(e)

    try:
        generated["계약이행서약서"] = generate_pledge_pdf(
            school_name, unit_price, start_date, end_date
        )
    except Exception as e:
        errors["계약이행서약서"] = str(e)

    try:
        generated["안전보건점검표"] = generate_safety_check_pdf(school_name)
    except Exception as e:
        errors["안전보건점검표"] = str(e)

    # ② 기존 서류 파일 경로 목록
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base = os.getcwd()

    static_docs = {
        "사업자등록증.jpg":    os.path.join(base, "사업자등록증.jpg"),
        "허가증.jpg":          os.path.join(base, "허가증.jpg"),
        "소상공인확인서.pdf":  os.path.join(base, "소상공인.pdf"),
        "창업기업확인서.pdf":  os.path.join(base, "창업기업.pdf"),
        "재해율확인서.pdf":    os.path.join(base, "재해율확인서.pdf"),
        "사업자계좌.jpg":      os.path.join(base, "사업자계좌.jpg"),
    }

    # ③ ZIP 메모리 내 생성
    buf = io.BytesIO()
    today_str = date.today().strftime("%Y%m%d")
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 자동 생성 PDF
        for name, path in generated.items():
            if os.path.exists(path):
                zf.write(path, f"01_자동생성/{os.path.basename(path)}")
        # 기존 서류
        for arc_name, fpath in static_docs.items():
            if os.path.exists(fpath):
                zf.write(fpath, f"02_기존서류/{arc_name}")
        # 오류 목록 있으면 txt로 포함
        if errors:
            err_txt = "\n".join(f"{k}: {v}" for k, v in errors.items())
            zf.writestr("오류목록.txt", err_txt)
        # 체크리스트
        checklist = f"""하영자원 계약서류 패키지 체크리스트
생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}
대상학교: {school_name}
계약기간: {contract_period}
단  가: {unit_price}원/L
==============================================
[자동생성 서류]
{'✅' if '견적서' in generated else '❌'} 음식물 견적서 PDF
{'✅' if '위수탁계약서' in generated else '❌'} 위수탁계약서 PDF
{'✅' if '계약이행서약서' in generated else '❌'} 계약이행 통합 서약서 PDF

[기존 서류]
{'✅' if os.path.exists(static_docs['사업자등록증.jpg']) else '⚠️ 없음'} 사업자등록증
{'✅' if os.path.exists(static_docs['허가증.jpg']) else '⚠️ 없음'} 폐기물수집운반업 허가증
{'✅' if os.path.exists(static_docs['소상공인확인서.pdf']) else '⚠️ 없음'} 소상공인 확인서 (만료: 2026-03-31 ⚠️)
{'✅' if os.path.exists(static_docs['창업기업확인서.pdf']) else '⚠️ 없음'} 창업기업 확인서
{'✅' if os.path.exists(static_docs['재해율확인서.pdf']) else '⚠️ 없음'} 재해율 확인서
{'✅' if os.path.exists(static_docs['사업자계좌.jpg']) else '⚠️ 없음'} 사업자 계좌 통장사본
==============================================
⚠️ 주의: 소상공인 확인서는 2026-03-31 만료 예정입니다.
   갱신 주소: sminfo.mss.go.kr
"""
        zf.writestr("00_서류체크리스트.txt", checklist)

    return buf.getvalue(), errors


# ============================================================
# DB 초기화 + 데이터 로드
# ============================================================
init_db()
df_all = load_data()

# ============================================================
# 로그인 화면 렌더링 함수
# ============================================================
def render_login_page():
    """S2B 스타일 랜딩 + 로그인 카드"""
    # 선택된 그룹 state
    if "login_group" not in st.session_state:
        st.session_state.login_group = None

    st.markdown("""
    <div class="login-header">
        <div style="display:flex;align-items:center;justify-content:center;gap:12px;margin-bottom:10px;">
            <span style="font-size:2.8rem;">♻️</span>
            <div>
                <div style="font-size:2rem;font-weight:900;color:#1a3a5c;line-height:1.1;">하영자원 데이터 플랫폼</div>
                <div style="font-size:0.95rem;color:#4a6b8a;margin-top:4px;">투명하고 효율적인 공공 폐기물 관리 솔루션</div>
            </div>
        </div>
        <div style="height:3px;width:80px;background:linear-gradient(90deg,#1a73e8,#34a853);border-radius:2px;margin:16px auto 0;"></div>
    </div>
    """, unsafe_allow_html=True)

    # ── 3개 카드 버튼 ──
    col_l, col_c, col_r = st.columns([1,3,1])
    with col_c:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("""
            <div class="login-card" style="border-top:5px solid #1a73e8;">
                <div class="login-card-icon">🏢</div>
                <div class="login-card-title">관리자</div>
                <div class="login-card-desc">본사 통합 관제<br>전체 권한 접근</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("관리자 로그인", key="btn_admin", use_container_width=True, type="primary"):
                st.session_state.login_group = "관리자"
        with c2:
            st.markdown("""
            <div class="login-card" style="border-top:5px solid #34a853;">
                <div class="login-card-icon">🏫</div>
                <div class="login-card-title">교육청 / 학교</div>
                <div class="login-card-desc">교육청·행정실<br>담당자 전용</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("교육청·학교 로그인", key="btn_school", use_container_width=True):
                st.session_state.login_group = "학교_교육청"
        with c3:
            st.markdown("""
            <div class="login-card" style="border-top:5px solid #fbbc05;">
                <div class="login-card-icon">🚚</div>
                <div class="login-card-title">수거업체</div>
                <div class="login-card-desc">기사·현장 관리자<br>전용 앱</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("수거업체 로그인", key="btn_driver", use_container_width=True):
                st.session_state.login_group = "수거기사"

    # ── 로그인 폼 ──
    if st.session_state.login_group:
        st.write("")
        col_fl, col_fc, col_fr = st.columns([1,2,1])
        with col_fc:
            group = st.session_state.login_group
            icons = {"관리자": "🏢", "학교_교육청": "🏫", "수거기사": "🚚"}
            titles = {"관리자": "관리자 로그인", "학교_교육청": "교육청 / 학교(행정실) 로그인", "수거기사": "수거업체(기사) 로그인"}
            st.markdown(f"""
            <div style="background:white;border-radius:16px;padding:36px;box-shadow:0 4px 20px rgba(0,0,0,.1);">
                <div style="text-align:center;font-size:1.3rem;font-weight:800;color:#1a3a5c;margin-bottom:24px;">
                    {icons[group]} {titles[group]}
                </div>
            </div>
            """, unsafe_allow_html=True)
            with st.form(f"login_form_{group}"):
                uid = st.text_input("아이디", placeholder="아이디를 입력하세요")
                pwd = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
                submitted = st.form_submit_button("🔐 로그인", use_container_width=True, type="primary")
                if submitted:
                    if uid in USER_ACCOUNTS:
                        acc = USER_ACCOUNTS[uid]
                        if acc["password"] == pwd:
                            st.session_state.logged_in = True
                            st.session_state.user_id   = uid
                            st.session_state.user_role = acc["role"]
                            st.session_state.user_name = acc["display_name"]
                            st.session_state.user_org  = acc["org"]
                            st.rerun()
                        else:
                            st.error("❌ 비밀번호가 올바르지 않습니다.")
                    else:
                        st.error("❌ 존재하지 않는 아이디입니다.")

    st.markdown("""
    <div class="login-footer" style="margin-top:40px;">
        ⓒ 2025 하영자원 | 하영자원 데이터 플랫폼 Pro v3.0 | 문의: 하영자원 본사
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# 세션 상태 초기화
# ============================================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# ── 미로그인 → 로그인 화면 표시 후 중단 ──
if not st.session_state.logged_in:
    render_login_page()
    st.stop()

# ============================================================
# 사이드바 (로그인 후)
# ============================================================
with st.sidebar:
    st.markdown("## ♻️ 하영자원 Pro v3")
    st.caption("공공기관(B2G) 맞춤 데이터 플랫폼")
    st.write("---")
    st.markdown(f"""
    **👤 {st.session_state.user_name}**
    🏷️ 역할: `{st.session_state.user_role}`
    🏢 소속: {st.session_state.user_org}
    """)
    st.write("---")
    if st.button("🚪 로그아웃", use_container_width=True, type="secondary"):
        for k in ["logged_in","user_id","user_role","user_name","user_org","login_group"]:
            st.session_state.pop(k, None)
        st.rerun()
    st.write("---")
    st.success("✅ SQLite DB (WAL모드)")
    st.caption("v3: 재활용시세·스쿨존·캘린더·교육청모드")

# ── 로그인 정보 기반 role 매핑 ──
_role_map = {
    "관리자":   "🏢 관리자 (본사 관제)",
    "학교":     "🏫 학교 담당자 (행정실)",
    "수거기사": "🚚 수거 기사 (현장 앱)",
    "교육청":   "🏛️ 교육청 관제 (신규)",
}

# 관리자는 사이드바 radio로 추가 탭 선택
if st.session_state.user_role == "관리자":
    _tabs_admin = [
        "🏢 관리자 (본사 관제)",
        "🏫 학교 마스터 관리",
        "📋 서류 유효기간 관리",
        "💰 견적서 작성",
        "📄 위수탁계약서 작성",
        "📦 계약서류 패키지 생성",
    ]
    role = st.sidebar.radio("메뉴", _tabs_admin, label_visibility="collapsed")
else:
    role = _role_map.get(st.session_state.user_role, "🏢 관리자 (본사 관제)")

# ============================================================
# [모드 1] 관리자 (본사 관제)
# ============================================================
if role == "🏢 관리자 (본사 관제)":
    st.markdown("<h1>🏢 본사 통합 관제 및 정산 센터</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#5f6368;font-size:16px;'>음식물·사업장·재활용 통계를 분리하여 수익/비용 관리가 가능합니다.</p>", unsafe_allow_html=True)

    # KPI 카드
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1: st.markdown(f'<div class="custom-card custom-card-red"><div class="metric-title">🗑️ 음식물 총 수거</div><div class="metric-value-food">{df_all["음식물(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="custom-card custom-card-purple"><div class="metric-title">🗄️ 사업장 총 수거</div><div class="metric-value-biz">{df_all["사업장(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="custom-card custom-card-green"><div class="metric-title">♻️ 재활용 총 수거</div><div class="metric-value-recycle">{df_all["재활용(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="custom-card"><div class="metric-title">💰 누적 청구 금액</div><div class="metric-value-total">{df_all["최종정산액"].sum():,.0f} 원</div></div>', unsafe_allow_html=True)
    # [추가5] 공공예산 절감 지표
    school_count = df_all["학교명"].nunique()
    budget_saving = school_count * int(get_setting("budget_saving_per_school") or 5200000)
    with c5: st.markdown(f'<div class="custom-card custom-card-teal"><div class="metric-title">💵 공공예산 절감(연)</div><div class="metric-value-total" style="color:#00897b;">{budget_saving:,.0f}원</div></div>', unsafe_allow_html=True)

    # ESG 배너 (소나무 + 업무절감 + 예산절감)
    co2 = df_all["탄소감축량(kg)"].sum()
    trees = int(co2 / 6.6)
    work_hours_saved = school_count * 52 * 6  # 학교수 × 52주 × 주당 6시간 절감
    st.markdown(f"""
    <div style="background:#61b346;padding:25px;border-radius:12px;color:white;display:flex;justify-content:space-around;align-items:center;margin-bottom:10px;">
        <div style="text-align:center;">
            <p style="margin:0;font-size:13px;opacity:.9;">🌍 누적 CO₂ 감축량</p>
            <h2 style="margin:0;color:white;font-weight:900;">{co2:,.0f} kg</h2>
            <p style="margin:0;font-size:12px;opacity:.8;">🌲 소나무 {trees:,}그루 효과</p>
        </div>
        <div style="text-align:center;">
            <p style="margin:0;font-size:13px;opacity:.9;">⏱️ 행정시간 절감(연)</p>
            <h2 style="margin:0;color:white;font-weight:900;">{work_hours_saved:,} 시간</h2>
            <p style="margin:0;font-size:12px;opacity:.8;">담당자 90% 업무 단축</p>
        </div>
        <div style="text-align:center;">
            <p style="margin:0;font-size:13px;opacity:.9;">💰 공공예산 절감(연)</p>
            <h2 style="margin:0;color:white;font-weight:900;">{budget_saving/100000000:.1f}억 원</h2>
            <p style="margin:0;font-size:12px;opacity:.8;">{school_count}개교 기준</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.write("---")
    # ── [7단계] 연간 계약 현황 요약 대시보드 ─────────────────
    st.subheader("📊 연간 계약 현황 대시보드")

    try:
        _dash_schools = a_get_all_schools()
        _dash_active  = [s for s in _dash_schools if s["계약상태"] == "계약중"]
        _dash_none    = [s for s in _dash_schools if s["계약상태"] == "미계약"]
        _dash_expire  = [s for s in _dash_schools
                         if "⛔" in s["계약D-day"] or "🔴" in s["계약D-day"]]

        # 연간 예상 총 매출 계산 (계약중 학교 단가 × 월예상량 × 12)
        _dash_contracts = a_get_contracts()
        _revenue_map    = {}
        for c_ in _dash_contracts:
            if c_["계약_상태"] == "계약중":
                sn = c_["학교명"]
                monthly = float(c_.get("월_예상량_L") or 0) * int(c_.get("단가") or 0)
                if sn not in _revenue_map or _revenue_map[sn] < monthly:
                    _revenue_map[sn] = monthly
        _annual_revenue = sum(_revenue_map.values()) * 12

        # 이번 달 정산 예정 (collections 기반)
        _this_ym      = datetime.now().strftime("%Y-%m")
        _month_settle = calc_monthly_settlement(datetime.now().year, datetime.now().month)
        _month_total  = sum(v["정산금액"] for v in _month_settle.values())

        # 올해 신규/갱신 계약 수
        _this_year = str(datetime.now().year)
        _new_cnt   = sum(1 for c_ in _dash_contracts
                         if (c_.get("계약_시작일","") or "").startswith(_this_year))

        # 4열 카드
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("🏫 계약중 학교",   f"{len(_dash_active)}개교",
                  delta=f"미계약 {len(_dash_none)}개교", delta_color="off")
        d2.metric("💰 연간 예상 매출",
                  f"{int(_annual_revenue/10000):,}만원" if _annual_revenue else "데이터 없음")
        d3.metric("📅 이번달 정산 예정",
                  f"{_month_total:,}원" if _month_total else "수거 데이터 없음")
        d4.metric("📝 올해 신규 계약",  f"{_new_cnt}건",
                  delta=f"만료임박 {len(_dash_expire)}교", delta_color="inverse")

        # 계약 만료 임박 경고 배너
        if _dash_expire:
            exp_names = ", ".join(s["학교명"] for s in _dash_expire[:5])
            extra     = f" 외 {len(_dash_expire)-5}개교" if len(_dash_expire) > 5 else ""
            st.warning(f"⏰ **계약 만료임박:** {exp_names}{extra} — 🏫 학교 마스터 관리에서 갱신하세요")
    except Exception as _e:
        st.info(f"대시보드 데이터 로드 중... ({_e})")

    st.write("---")
    st.subheader("📑 통합 정산 관리")

    tab_total, tab_food, tab_biz, tab_recycle, tab_map, tab_sub, tab_price, tab_recycle_price, tab_notify = st.tabs([
        "전체 통합 정산", "음식물 정산", "사업장 정산", "재활용 정산",
        "📍 차량 관제", "🤝 외주업체",
        "💰 단가 설정", "♻️ 재활용 시세 ✨", "📱 알림 설정"
    ])

    with tab_total:
        cur_year = str(datetime.now().year)
        cur_month = datetime.now().strftime("%Y-%m")
        prev_month_n = datetime.now().month - 1 or 12
        prev_month = f"{datetime.now().year if datetime.now().month > 1 else datetime.now().year-1}-{prev_month_n:02d}"

        # ── 상단 필터: 연도 / 월 / 학교 선택 ──
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            all_years = sorted(df_all["년도"].unique(), reverse=True)
            sel_year_admin = st.selectbox("📅 연도", all_years, key="admin_year")
        with fc2:
            months_in_year = sorted(df_all[df_all["년도"]==sel_year_admin]["월별"].unique(), reverse=True)
            sel_month_admin = st.selectbox("🗓️ 월", ["전체"] + list(months_in_year), key="admin_month")
        with fc3:
            sel_school_admin = st.selectbox("🏫 학교", ["전체"] + SCHOOL_LIST, key="admin_school")

        # 필터 적용
        df_filtered = df_all[df_all["년도"] == sel_year_admin]
        if sel_month_admin != "전체":
            df_filtered = df_filtered[df_filtered["월별"] == sel_month_admin]
        if sel_school_admin != "전체":
            df_filtered = df_filtered[df_filtered["학교명"] == sel_school_admin]

        s1, s2, s3, s4 = st.tabs([
            f"📋 필터결과",
            f"📅 {cur_year}년 전체",
            f"🗓️ 이번달({cur_month})",
            f"🗓️ 지난달({prev_month})"
        ])
        with s1:
            st.caption(f"조건: {sel_year_admin}년 / {sel_month_admin} / {sel_school_admin} — {len(df_filtered)}건")
            st.dataframe(df_filtered[["날짜","학교명","학생수","음식물(kg)","최종정산액","탄소감축량(kg)","상태"]], use_container_width=True)
            if not df_filtered.empty:
                m1, m2, m3 = st.columns(3)
                m1.metric("총 음식물", f"{df_filtered['음식물(kg)'].sum():,.0f} kg")
                m2.metric("총 정산액", f"{df_filtered['최종정산액'].sum():,.0f} 원")
                m3.metric("CO₂ 감축", f"{df_filtered['탄소감축량(kg)'].sum():,.1f} kg")
                # 학교별 소계 테이블
                if sel_school_admin == "전체":
                    st.markdown("**📊 학교별 소계**")
                    school_sum = df_filtered.groupby("학교명").agg(
                        수거건수=("id","count"),
                        음식물_합계=("음식물(kg)","sum"),
                        정산액_합계=("최종정산액","sum"),
                        CO2감축=("탄소감축량(kg)","sum")
                    ).reset_index().sort_values("정산액_합계", ascending=False)
                    st.dataframe(school_sum, use_container_width=True)
                dl_f = create_secure_excel(df_filtered[["날짜","학교명","음식물(kg)","사업장(kg)","재활용(kg)","최종정산액","탄소감축량(kg)","상태"]], "필터 정산서")
                st.download_button("📥 필터결과 정산서 다운로드", data=dl_f,
                                   file_name=f"하영자원_{sel_year_admin}_{sel_month_admin}_{sel_school_admin}.xlsx", use_container_width=True)
        with s2:
            df_cur = df_all[df_all["년도"]==cur_year]
            st.dataframe(df_cur[["날짜","학교명","학생수","최종정산액","탄소감축량(kg)","상태"]], use_container_width=True)
        with s3:
            df_cm = df_all[df_all["월별"]==cur_month]
            st.dataframe(df_cm[["날짜","학교명","학생수","최종정산액","상태"]], use_container_width=True)
            if not df_cm.empty:
                pending = df_cm[df_cm["상태"]=="정산대기"]
                if not pending.empty:
                    if st.button(f"✅ 이번달 미정산 {len(pending)}건 → 정산완료 처리", type="primary"):
                        update_collection_status(pending["id"].tolist(), "정산완료")
                        st.success(f"{len(pending)}건이 정산완료로 변경되었습니다.")
                        st.rerun()
        with s4:
            df_pm = df_all[df_all["월별"]==prev_month]
            st.dataframe(df_pm[["날짜","학교명","학생수","최종정산액","상태"]], use_container_width=True)

        # 정산서 다운로드
        b1, b2 = st.columns(2)
        with b1:
            dl_data = create_secure_excel(
                df_all[["날짜","학교명","음식물(kg)","사업장(kg)","재활용(kg)","최종정산액","탄소감축량(kg)","상태"]],
                "전체 통합 정산서"
            )
            st.download_button("📥 전체 통합정산서 다운로드", data=dl_data,
                               file_name=f"하영자원_통합정산서_{cur_month}.xlsx", use_container_width=True)
        with b2:
            month_dl = create_secure_excel(
                df_all[df_all["월별"]==cur_month][["날짜","학교명","음식물(kg)","사업장(kg)","재활용(kg)","최종정산액","상태"]],
                f"{cur_month} 월간 정산서"
            )
            st.download_button(f"📥 이번달 정산서 다운로드", data=month_dl,
                               file_name=f"하영자원_월간정산서_{cur_month}.xlsx", use_container_width=True)

        # ── [4단계] 월별 정산 자동 계산 ─────────────────────
        st.divider()
        st.subheader("📊 월별 학교별 정산 자동 계산")
        st.caption("school_prices 단가 × collections 수거량 → 학교별 정산금액 자동 계산")

        ms1, ms2 = st.columns(2)
        with ms1:
            ms_year  = st.selectbox("정산 연도", list(range(2025, date.today().year + 2)),
                                    index=list(range(2025, date.today().year + 2)).index(date.today().year),
                                    key="ms_year")
        with ms2:
            ms_month = st.selectbox("정산 월", list(range(1, 13)),
                                    index=date.today().month - 1,
                                    key="ms_month",
                                    format_func=lambda m: f"{m:02d}월")

        if st.button("🔢 정산 계산 실행", type="primary",
                     use_container_width=True, key="ms_calc"):
            with st.spinner("계산 중..."):
                settlement = calc_monthly_settlement(int(ms_year), int(ms_month))
                st.session_state["ms_result"]    = settlement
                st.session_state["ms_year_sel"]  = int(ms_year)
                st.session_state["ms_month_sel"] = int(ms_month)

        ms_result = st.session_state.get("ms_result")
        if ms_result:
            ms_y = st.session_state.get("ms_year_sel", ms_year)
            ms_m = st.session_state.get("ms_month_sel", ms_month)

            if not ms_result:
                st.info(f"⚠️ {ms_y}년 {ms_m:02d}월 수거 데이터가 없습니다.")
            else:
                # 합계
                total_vol = sum(v["수거량_L"]  for v in ms_result.values())
                total_amt = sum(v["정산금액"]  for v in ms_result.values())
                total_cnt = sum(v["수거건수"]  for v in ms_result.values())

                # 요약 카드
                sc1, sc2, sc3 = st.columns(3)
                sc1.metric("🏫 정산 학교 수",    f"{len(ms_result)}개교")
                sc2.metric("🗑️ 총 수거량",       f"{total_vol:,.1f} L")
                sc3.metric("💰 총 정산금액",      f"{total_amt:,} 원")

                # 정산표
                df_ms = pd.DataFrame([{
                    "학교명":       s,
                    "수거건수":     d["수거건수"],
                    "수거량(L)":    d["수거량_L"],
                    "단가(원/L)":   d["단가"],
                    "정산금액(원)": d["정산금액"],
                    "세금구분":     d["세금구분"],
                } for s, d in ms_result.items()])

                # 합계 행 추가
                total_row = pd.DataFrame([{
                    "학교명": "【 합 계 】",
                    "수거건수": total_cnt,
                    "수거량(L)": round(total_vol, 1),
                    "단가(원/L)": "-",
                    "정산금액(원)": total_amt,
                    "세금구분": "면세",
                }])
                df_ms_show = pd.concat([df_ms, total_row], ignore_index=True)

                st.dataframe(
                    df_ms_show,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "수거량(L)":    st.column_config.NumberColumn(format="%.1f"),
                        "정산금액(원)": st.column_config.NumberColumn(format="%d"),
                        "단가(원/L)":   st.column_config.TextColumn(width="small"),
                        "세금구분":     st.column_config.TextColumn(width="small"),
                    }
                )

                # 엑셀 다운로드
                try:
                    excel_bytes = create_settlement_excel(ms_result, ms_y, ms_m)
                    st.download_button(
                        label     = f"📥 {ms_y}년 {ms_m:02d}월 정산표 엑셀 다운로드",
                        data      = excel_bytes,
                        file_name = f"하영자원_{ms_y}년{ms_m:02d}월_정산표.xlsx",
                        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key       = "ms_dl",
                        use_container_width=True,
                        type      = "primary",
                    )
                except Exception as e:
                    st.warning(f"엑셀 생성 오류: {e}")
                    st.info("openpyxl 설치 필요: pip install openpyxl")

        # ── [7단계] 수익성 분석 차트 ──────────────────────────
        st.divider()
        st.subheader("📈 학교별 수익성 분석")

        try:
            # ① 학교별 연간 수거량 + 정산금액 집계
            _ch_grp = (
                df_all.groupby("학교명")
                .agg(
                    연간수거량=("음식물(kg)", "sum"),
                    연간정산액=("최종정산액",  "sum"),
                    수거건수  =("id",          "count"),
                )
                .reset_index()
                .sort_values("연간정산액", ascending=False)
            )
            _ch_top = _ch_grp.head(10)

            ch1, ch2 = st.columns(2)

            with ch1:
                st.markdown("**🏆 연간 정산금액 상위 10개교**")
                if not _ch_top.empty:
                    st.bar_chart(
                        _ch_top.set_index("학교명")["연간정산액"],
                        use_container_width=True,
                        height=280,
                        color="#1a73e8",
                    )
                    st.caption("단위: 원 (음식물 기준)")
                else:
                    st.info("수거 데이터 없음")

            with ch2:
                st.markdown("**🗑️ 연간 수거량 상위 10개교**")
                if not _ch_top.empty:
                    st.bar_chart(
                        _ch_top.set_index("학교명")["연간수거량"],
                        use_container_width=True,
                        height=280,
                        color="#34a853",
                    )
                    st.caption("단위: kg (= L)")
                else:
                    st.info("수거 데이터 없음")

            # ② 단가 구간별 학교 분포
            try:
                _sp_data = get_conn().execute(
                    "SELECT 학교명, 음식물단가 FROM school_prices ORDER BY 음식물단가"
                ).fetchall()
                if _sp_data:
                    def _price_band(p):
                        if p <= 100:   return "① 100원 이하"
                        if p <= 150:   return "② 101~150원"
                        if p <= 200:   return "③ 151~200원"
                        return          "④ 201원 이상"

                    _band_cnt = {}
                    for _, p in _sp_data:
                        b = _price_band(p or 150)
                        _band_cnt[b] = _band_cnt.get(b, 0) + 1

                    ch3, ch4 = st.columns(2)

                    with ch3:
                        st.markdown("**💴 단가 구간별 학교 분포**")
                        _band_df = pd.DataFrame(
                            list(_band_cnt.items()),
                            columns=["단가 구간", "학교 수"]
                        ).sort_values("단가 구간")
                        st.bar_chart(
                            _band_df.set_index("단가 구간")["학교 수"],
                            use_container_width=True,
                            height=250,
                            color="#fbbc04",
                        )
                        # 구간별 학교 목록
                        with st.expander("📋 구간별 학교 상세"):
                            for b_name, b_count in sorted(_band_cnt.items()):
                                schools_in = [nm for nm, p in _sp_data if _price_band(p or 150) == b_name]
                                st.markdown(f"**{b_name}** ({b_count}개교): {', '.join(schools_in)}")

                    with ch4:
                        st.markdown("**📅 월별 전체 수거량 추이**")
                        _monthly_vol = (
                            df_all.groupby("월별")["음식물(kg)"]
                            .sum()
                            .reset_index()
                            .sort_values("월별")
                        )
                        if not _monthly_vol.empty:
                            st.line_chart(
                                _monthly_vol.set_index("월별")["음식물(kg)"],
                                use_container_width=True,
                                height=250,
                                color="#ea4335",
                            )
                            st.caption("단위: kg / 전체 학교 합산")
                        else:
                            st.info("월별 데이터 없음")
            except Exception as _e2:
                st.warning(f"단가/월별 차트 오류: {_e2}")

            # ③ 수익성 요약 테이블
            st.markdown("**📋 학교별 수익성 요약 테이블**")
            if not _ch_grp.empty:
                _ch_grp["평균단가(원/L)"] = (_ch_grp["연간정산액"] / _ch_grp["연간수거량"]).round(1)
                _ch_grp["연간정산액(만원)"] = (_ch_grp["연간정산액"] / 10000).round(1)
                st.dataframe(
                    _ch_grp[["학교명","수거건수","연간수거량","평균단가(원/L)","연간정산액(만원)"]].reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "연간수거량":       st.column_config.NumberColumn(format="%,.0f kg"),
                        "연간정산액(만원)": st.column_config.NumberColumn(format="%.1f 만원"),
                        "평균단가(원/L)":   st.column_config.NumberColumn(format="%.1f 원"),
                    }
                )
        except Exception as _e:
            st.warning(f"수익성 분석 오류: {_e}")

    with tab_food:
        f1, f2 = st.tabs([f"📅 이번달", "📅 전체"])
        with f1: st.dataframe(df_all[df_all["월별"]==datetime.now().strftime("%Y-%m")][["날짜","학교명","수거업체","음식물(kg)","단가","음식물비용","상태"]], use_container_width=True)
        with f2: st.dataframe(df_all[["날짜","학교명","수거업체","음식물(kg)","단가","음식물비용","상태"]], use_container_width=True)

    with tab_biz:
        b1, b2 = st.tabs(["📅 이번달", "📅 전체"])
        with b1: st.dataframe(df_all[df_all["월별"]==datetime.now().strftime("%Y-%m")][["날짜","학교명","학생수","사업장(kg)","사업장단가","사업장비용"]], use_container_width=True)
        with b2: st.dataframe(df_all[["날짜","학교명","학생수","사업장(kg)","사업장단가","사업장비용"]], use_container_width=True)

    with tab_recycle:
        r1, r2 = st.tabs(["📅 이번달", "📅 전체"])
        with r1: st.dataframe(df_all[df_all["월별"]==datetime.now().strftime("%Y-%m")][["날짜","학교명","학생수","재활용(kg)","재활용단가","재활용수익"]], use_container_width=True)
        with r2: st.dataframe(df_all[["날짜","학교명","학생수","재활용(kg)","재활용단가","재활용수익"]], use_container_width=True)
        # 품목별 수익 분석
        st.write("---")
        st.subheader("📊 재활용 수익 분석")
        rp_df = get_recycle_prices()
        rr1, rr2 = st.columns(2)
        with rr1:
            st.dataframe(rp_df[["품목명","단가","updated_at"]], use_container_width=True)
        with rr2:
            total_recycle_rev = df_all["재활용수익"].sum()
            st.metric("♻️ 누적 재활용 총 수익", f"{total_recycle_rev:,.0f} 원")
            st.metric("📦 평균 단가", f"{get_avg_recycle_price()} 원/kg")

    with tab_map:
        st.write("📍 **수거 차량 실시간 GPS 관제**")
        # 실제 학교 위치 좌표 (화성/수원/서울 지역)
        school_coords = {
            "화성초등학교": [37.1994, 126.8311],
            "부림초등학교": [37.2134, 126.8901],
            "동탄중학교":   [37.2001, 127.0720],
            "수원고등학교": [37.2636, 127.0286],
            "서초고등학교": [37.4875, 127.0322],
            "국사봉중학교": [37.4810, 126.9201],
        }
        map_df = pd.DataFrame(
            [(v[0], v[1], k) for k, v in school_coords.items()],
            columns=["lat","lon","학교명"]
        )
        st.map(map_df)
        st.caption("🟢 현재 운행 중인 차량 위치 (GPS 연동 시 실시간 업데이트)")
        # 오늘 수거 현황
        today_str = datetime.now().strftime("%Y-%m-%d")
        today_df = df_all[df_all["날짜"].str.startswith(today_str)]
        if not today_df.empty:
            st.success(f"✅ 오늘 수거 완료: {len(today_df)}건")
        else:
            st.info("오늘 수거 데이터가 없습니다.")

    with tab_sub:
        st.subheader("🤝 외주 수거업체 현황")
        st.markdown('<div class="alert-box">🔔 <b>[계약 갱신 알림]</b> \'B자원\' 업체 계약 만료 30일 전입니다. (만료일: 2026-03-25)</div>', unsafe_allow_html=True)
        cc1, cc2, cc3 = st.columns(3)
        with cc1: st.info("🏆 이달 우수 업체: **A환경** (98점)")
        with cc2: st.warning("⚠️ 주의: **B자원** (과속 1회)")
        with cc3: st.success("✅ 스쿨존 위반: 1건")
        vendor_df = pd.DataFrame({
            "외주업체명": ["A환경","B자원"],
            "담당학교": ["동탄중학교","수원고등학교"],
            "안전평가": ["98점(우수)","85점(주의)"],
            "페널티": ["0원","-50,000원"],
            "지급예정액": ["1,350,000원","880,000원"],
            "상태": ["🟢 운행중","🟡 대기중"]
        })
        st.dataframe(vendor_df, use_container_width=True)
        # 실제 다운로드
        vd_excel = create_secure_excel(vendor_df, "외주업체 안전평가 결과서")
        st.download_button("📄 외주업체 안전평가 결과서 다운로드", data=vd_excel,
                           file_name=f"외주업체_안전평가_{datetime.now().strftime('%Y-%m')}.xlsx")

    with tab_price:
        st.subheader("💰 단가 설정 관리")
        st.info("이 화면에서 단가를 변경하면 **즉시 전체 정산에 반영**됩니다.")

        st.markdown("### 🌐 전체 기본 단가")
        gp1, gp2, gp3 = st.columns(3)
        with gp1: g_food    = st.number_input("음식물 기본단가 (원/kg)", value=int(get_setting("default_food_price")),    min_value=0, step=10, key="g_food")
        with gp2: g_recycle = st.number_input("재활용 기본단가 (원/kg)", value=int(get_setting("default_recycle_price")), min_value=0, step=10, key="g_recycle")
        with gp3: g_biz     = st.number_input("사업장 기본단가 (원/kg)", value=int(get_setting("default_biz_price")),     min_value=0, step=10, key="g_biz")
        if st.button("💾 기본 단가 저장", type="primary", key="save_global"):
            set_setting("default_food_price",    g_food)
            set_setting("default_recycle_price", g_recycle)
            set_setting("default_biz_price",     g_biz)
            st.success("✅ 저장 완료. 페이지 새로고침 후 반영됩니다.")

        st.write("---")
        st.markdown("### 🏫 학교별 개별 단가 + 담당자")
        sel_school = st.selectbox("설정할 학교", SCHOOL_LIST, key="price_sel")
        conn = get_conn()
        ex = conn.execute(
            "SELECT 음식물단가,재활용단가,사업장단가,담당자명,담당자연락처,담당자이메일 FROM school_prices WHERE 학교명=?",
            (sel_school,)
        ).fetchone()
        conn.close()
        ef, er, eb, en, et, ee = ex if ex else (150,300,200,"","","")
        sp1, sp2, sp3 = st.columns(3)
        with sp1: sp_food    = st.number_input("음식물 단가", value=int(ef), min_value=0, step=10, key="sp_food")
        with sp2: sp_recycle = st.number_input("재활용 단가", value=int(er), min_value=0, step=10, key="sp_recycle")
        with sp3: sp_biz     = st.number_input("사업장 단가", value=int(eb), min_value=0, step=10, key="sp_biz")
        sc1, sc2, sc3 = st.columns(3)
        with sc1: sp_name  = st.text_input("담당자 이름",   value=en or "", key="sp_name")
        with sc2: sp_tel   = st.text_input("담당자 연락처", value=et or "", placeholder="010-0000-0000", key="sp_tel")
        with sc3: sp_email = st.text_input("담당자 이메일", value=ee or "", placeholder="admin@school.kr", key="sp_email")
        if st.button(f"💾 {sel_school} 저장", type="primary", key="save_school"):
            update_school_price(sel_school, sp_food, sp_recycle, sp_biz, sp_name, sp_tel, sp_email)
            st.success(f"✅ {sel_school} 저장 완료")

        st.write("---")
        st.markdown("### 📋 전체 학교 단가 현황")
        price_overview = get_school_prices()[["학교명","음식물단가","재활용단가","사업장단가","담당자명","담당자연락처","updated_at"]]
        st.dataframe(price_overview, use_container_width=True)
        st.download_button("📥 단가 현황 다운로드", data=create_secure_excel(price_overview, "학교별 계약 단가 현황"),
                           file_name="학교별단가현황.xlsx")

    # [추가2] 재활용 시세 탭
    with tab_recycle_price:
        st.subheader("♻️ 재활용품 23종 시세 관리")
        st.markdown('<span class="badge-v3">✨ v3 신규</span>', unsafe_allow_html=True)
        st.info("품목별 단가를 실시간으로 수정하면 재활용 수익 계산에 즉시 반영됩니다.")

        rp_df = get_recycle_prices()
        rp1, rp2 = st.columns([2, 1])
        with rp1:
            st.markdown("### 📋 현재 시세 현황")
            st.dataframe(rp_df[["품목명","단가","updated_at"]].rename(columns={"단가":"단가(원/kg)","updated_at":"최종수정"}),
                         use_container_width=True)
        with rp2:
            st.markdown("### ✏️ 시세 수정")
            sel_item = st.selectbox("품목 선택", rp_df["품목명"].tolist(), key="sel_recycle_item")
            cur_price = int(rp_df[rp_df["품목명"]==sel_item]["단가"].values[0])
            new_price = st.number_input("새 단가 (원/kg)", value=cur_price, min_value=0, step=10, key="new_recycle_price")
            if st.button("💾 시세 저장", type="primary", key="save_recycle"):
                update_recycle_price(sel_item, new_price)
                st.success(f"✅ {sel_item}: {new_price}원/kg 저장 완료")
                st.rerun()

        st.write("---")
        st.markdown("### 📊 품목별 수익 기여도")
        chart_df = rp_df.sort_values("단가", ascending=False)
        st.bar_chart(chart_df.set_index("품목명")["단가"])

        # 시세 일괄 다운로드
        st.download_button(
            "📥 재활용 시세표 다운로드",
            data=create_secure_excel(rp_df[["품목명","단가"]], "재활용품 시세표"),
            file_name=f"재활용시세표_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )

    with tab_notify:
        st.subheader("📱 알림 설정")
        st.info("카카오 알림톡 API 미연동 상태입니다. 아래에서 시뮬레이션 테스트가 가능합니다.")
        cur_notify = get_setting("kakao_notify_enabled") == "true"
        new_notify = st.toggle("📱 알림톡 활성화 (시뮬레이션)", value=cur_notify)
        if new_notify != cur_notify:
            set_setting("kakao_notify_enabled", "true" if new_notify else "false")
            st.success("설정 저장 완료")
        st.write("---")
        nt1, nt2 = st.columns(2)
        with nt1: test_school = st.selectbox("테스트 학교", SCHOOL_LIST, key="noti_school")
        with nt2: test_phone  = st.text_input("수신 번호", placeholder="010-0000-0000", key="noti_phone")
        if st.button("📱 알림톡 테스트", type="primary"):
            total_est = int(df_all[df_all["학교명"]==test_school]["최종정산액"].sum())
            send_kakao_alimtalk(test_phone or "010-0000-0000", test_school, 100.0, total_est)


# ============================================================
# [모드 2] 학교 담당자 (행정실)
# ============================================================
elif role == "🏫 학교 담당자 (행정실)":
    st.title("🏫 학교 폐기물 통합 대시보드")
    # 로그인된 학교 자동 설정 (selectbox 제거)
    school = st.session_state.user_org
    df_school = df_all[df_all["학교명"] == school]

    if not df_school.empty:
        conn = get_conn()
        pr = conn.execute("SELECT 음식물단가,재활용단가,사업장단가 FROM school_prices WHERE 학교명=?", (school,)).fetchone()
        conn.close()
        if pr:
            st.caption(f"📋 계약 단가 — 음식물: {pr[0]}원/kg | 사업장: {pr[2]}원/kg | 재활용: {pr[1]}원/kg")

        co2s = df_school["탄소감축량(kg)"].sum()
        trees_s = int(co2s / 6.6)
        work_saved = 52 * 6
        budget_s = int(get_setting("budget_saving_per_school") or 5200000)
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#11998e,#38ef7d);padding:20px;border-radius:12px;color:white;margin-bottom:20px;display:flex;justify-content:space-around;">
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;">🌱 누적 CO₂ 감축</p>
                <h3 style="margin:0;color:white;">{co2s:,.0f} kg</h3>
                <p style="margin:0;font-size:12px;">🌲 소나무 {trees_s}그루</p>
            </div>
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;">⏱️ 연간 업무절감</p>
                <h3 style="margin:0;color:white;">{work_saved}시간</h3>
                <p style="margin:0;font-size:12px;">담당자 90% 단축</p>
            </div>
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;">💰 연간 예산절감</p>
                <h3 style="margin:0;color:white;">{budget_s:,}원</h3>
                <p style="margin:0;font-size:12px;">교육청 제출용</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        school_tab1, school_tab2, school_tab3, school_tab4 = st.tabs([
            "📊 수거량 통계", "📅 수거 일정 캘린더 ✨", "🛡️ 안전 현황", "🖨️ 서류 출력"
        ])

        with school_tab1:
            t_daily, t_monthly = st.tabs(["🗓️ 일별 배출량", "🗓️ 연도별/월별 추이"])
            with t_daily:
                dg = df_school.copy()
                dg["일자"] = dg["날짜"].astype(str).str[:10]
                dg = dg.groupby("일자")[["음식물(kg)","사업장(kg)","재활용(kg)"]].sum().reset_index()
                dc1, dc2, dc3 = st.columns(3)
                with dc1:
                    st.markdown("<h5 style='text-align:center;color:#ea4335;'>🗑️ 음식물</h5>", unsafe_allow_html=True)
                    st.bar_chart(dg.set_index("일자")["음식물(kg)"], color="#ea4335")
                with dc2:
                    st.markdown("<h5 style='text-align:center;color:#9b59b6;'>🗄️ 사업장</h5>", unsafe_allow_html=True)
                    st.bar_chart(dg.set_index("일자")["사업장(kg)"], color="#9b59b6")
                with dc3:
                    st.markdown("<h5 style='text-align:center;color:#34a853;'>♻️ 재활용</h5>", unsafe_allow_html=True)
                    st.bar_chart(dg.set_index("일자")["재활용(kg)"], color="#34a853")
            with t_monthly:
                years = sorted(df_school["년도"].unique(), reverse=True)
                ytabs = st.tabs([f"📅 {y}년" for y in years])
                for i, y in enumerate(years):
                    with ytabs[i]:
                        df_yr = df_school[df_school["년도"]==y]
                        # 연도 요약 지표
                        ym1, ym2, ym3, ym4 = st.columns(4)
                        ym1.metric("총 음식물", f"{df_yr['음식물(kg)'].sum():,.0f} kg")
                        ym2.metric("총 정산액", f"{df_yr['최종정산액'].sum():,.0f} 원")
                        ym3.metric("수거횟수", f"{len(df_yr)}회")
                        ym4.metric("CO₂ 감축", f"{df_yr['탄소감축량(kg)'].sum():,.1f} kg")
                        st.write("---")
                        # 월별 하위탭
                        months_yr = sorted(df_yr["월별"].unique())
                        if months_yr:
                            mtabs = st.tabs([f"🗓️ {m[5:]}월" for m in months_yr])
                            for j, m in enumerate(months_yr):
                                with mtabs[j]:
                                    df_m = df_yr[df_yr["월별"]==m]
                                    mm1, mm2, mm3 = st.columns(3)
                                    mm1.metric("음식물", f"{df_m['음식물(kg)'].sum():,.0f} kg")
                                    mm2.metric("정산액", f"{df_m['최종정산액'].sum():,.0f} 원")
                                    mm3.metric("CO₂ 감축", f"{df_m['탄소감축량(kg)'].sum():,.1f} kg")
                                    mc1, mc2, mc3 = st.columns(3)
                                    mg = df_m.groupby("월별")[["음식물(kg)","사업장(kg)","재활용(kg)"]].sum().reset_index()
                                    with mc1: st.bar_chart(df_m.set_index(df_m["날짜"].str[:10])["음식물(kg)"], color="#ea4335")
                                    with mc2: st.bar_chart(df_m.set_index(df_m["날짜"].str[:10])["사업장(kg)"], color="#9b59b6")
                                    with mc3: st.bar_chart(df_m.set_index(df_m["날짜"].str[:10])["재활용(kg)"], color="#34a853")
                                    st.dataframe(df_m[["날짜","음식물(kg)","사업장(kg)","재활용(kg)","최종정산액","탄소감축량(kg)","상태"]], use_container_width=True)

        # [추가4] 수거 일정 캘린더
        with school_tab2:
            st.markdown('<span class="badge-v3">✨ v3 신규</span>', unsafe_allow_html=True)
            st.subheader("📅 수거 일정 캘린더")

            cal_col1, cal_col2 = st.columns([3, 2])
            with cal_col1:
                now = datetime.now()
                sel_year  = st.selectbox("연도", [now.year-1, now.year, now.year+1], index=1, key="cal_y_school")
                sel_month = st.selectbox("월", list(range(1,13)), index=now.month-1, key="cal_m_school")

            sched_df = get_schedules_month(sel_year, sel_month)
            school_sched = sched_df[sched_df["학교명"]==school]
            collect_days = set(school_sched["날짜"].str[8:10].astype(str).str.lstrip("0"))
            # 실제 수거 실적일도 달력에 표시
            real_collect = df_school[df_school["월별"]==f"{sel_year}-{sel_month:02d}" if isinstance(sel_month, int) else df_school["월별"]==f"{sel_year}-{str(sel_month).zfill(2)}"]
            real_days_set = set(real_collect["날짜"].astype(str).str[8:10].str.lstrip("0"))

            # 달력 렌더링
            cal_html = f"<h4 style='margin-bottom:10px;'>{sel_year}년 {sel_month}월 수거 일정</h4>"
            cal_html += "<table style='width:100%;border-collapse:collapse;'>"
            cal_html += "<tr>" + "".join(f"<th style='text-align:center;padding:6px;color:#666;'>{d}</th>" for d in ["일","월","화","수","목","금","토"]) + "</tr>"

            first_day = date(sel_year, sel_month, 1)
            _, num_days = calendar.monthrange(sel_year, sel_month)
            start_weekday = first_day.weekday()  # 0=월
            start_offset = (start_weekday + 1) % 7  # 일요일=0 기준

            day = 1
            today_d = date.today()
            cal_html += "<tr>"
            for i in range(start_offset):
                cal_html += "<td></td>"
            col_idx = start_offset

            while day <= num_days:
                d = date(sel_year, sel_month, day)
                day_str = str(day)
                is_today = (d == today_d)
                is_collect = day_str in collect_days       # 일정 등록일
                is_real    = day_str in real_days_set      # 실제 수거 완료일
                is_weekend = (col_idx % 7 == 0)

                if is_today:
                    cls = "cal-day cal-today"
                elif is_real:
                    cls = "cal-day cal-collect"            # 실적 완료 (진한 표시)
                elif is_collect:
                    cls = "cal-day cal-collect"
                elif is_weekend:
                    cls = "cal-day cal-weekend"
                else:
                    cls = "cal-day"

                icon = "✅" if is_real else ("🚛" if is_collect else "")
                cal_html += f"<td><div class='{cls}'>{day_str}{icon}</div></td>"
                col_idx += 1
                if col_idx % 7 == 0 and day < num_days:
                    cal_html += "</tr><tr>"
                day += 1

            while col_idx % 7 != 0:
                cal_html += "<td></td>"
                col_idx += 1
            cal_html += "</tr></table>"
            cal_html += "<p style='margin-top:8px;font-size:12px;color:#666;'>🟢 수거예정일 &nbsp; 🔵 오늘</p>"
            st.markdown(cal_html, unsafe_allow_html=True)

            st.write("---")
            st.markdown("#### 📋 이번달 수거 일정 목록")
            if not school_sched.empty:
                for _, row in school_sched.iterrows():
                    col_a, col_b, col_c = st.columns([3,1,1])
                    status_icon = "✅" if row["완료여부"] else "⏳"
                    with col_a: st.write(f"{status_icon} {row['날짜']} — {row['메모']}")
                    with col_b:
                        if st.button("완료토글", key=f"tog_{row['id']}"):
                            toggle_schedule(row["id"], row["완료여부"])
                            st.rerun()
                    with col_c:
                        if st.button("삭제", key=f"del_{row['id']}"):
                            delete_schedule(row["id"])
                            st.rerun()
            else:
                st.info("이 학교의 이번달 일정이 없습니다.")

            st.write("---")
            st.markdown("#### ➕ 일정 추가")
            with st.form("add_sched_school"):
                ns_date = st.date_input("수거 날짜", value=date.today(), key="ns_date_s")
                ns_memo = st.text_input("메모", placeholder="정기 수거 / 대용량 수거 등", key="ns_memo_s")
                if st.form_submit_button("📅 일정 추가", type="primary"):
                    add_schedule(str(ns_date), school, ns_memo or "정기 수거")
                    st.success("일정이 추가되었습니다.")
                    st.rerun()

        with school_tab3:
            st.markdown("<h5 style='color:#2e7d32;font-weight:bold;'>🛡️ 금일 수거차량 안전 점검 현황</h5>", unsafe_allow_html=True)
            # [추가3] 실제 시간 기반 스쿨존 상태
            restricted, time_range = is_schoolzone_restricted()
            if restricted:
                st.markdown(f'<div class="schoolzone-danger">🚨 현재 등하교 시간({time_range}) — 수거 차량 학교 진입 제한 중</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="schoolzone-safe">✅ 현재 수거 가능 시간 — 스쿨존 안전 운행 중</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="safety-box">✅ 배차 차량: 하영자원 (본사 직영 운행)<br>✅ 스쿨존 규정속도 준수: 정상 (MAX 28km/h 통과)<br>✅ 후방카메라 및 안전요원 동승: 적합<br>🕐 현재 시각: {datetime.now().strftime("%H:%M")}</div>', unsafe_allow_html=True)

        with school_tab4:
            st.subheader("🖨️ 행정 증빙 서류 자동 출력")
            d1, d2, d3, d4 = st.tabs(["📊 월간 정산서", "📈 처리실적보고서(제30호)", "♻️ 재활용 상계증빙", "🔗 올바로시스템"])

            with d1:
                st.info("행정실 회계 처리용 월간 정산서")
                dd1, dd2, dd3, dd4 = st.columns(4)
                with dd1: st.download_button("통합본", data=create_secure_excel(df_school[["날짜","학교명","음식물(kg)","사업장(kg)","최종정산액"]],"통합 정산(청구)서"), file_name=f"{school}_통합_월간정산서.xlsx", use_container_width=True)
                with dd2: st.download_button("🗑️ 음식물", data=create_secure_excel(df_school[["날짜","학교명","음식물(kg)","음식물비용"]],"음식물 정산(청구)서"), file_name=f"{school}_음식물_월간정산서.xlsx", use_container_width=True)
                with dd3: st.download_button("🗄️ 사업장", data=create_secure_excel(df_school[["날짜","학교명","사업장(kg)","사업장비용"]],"사업장 정산(청구)서"), file_name=f"{school}_사업장_월간정산서.xlsx", use_container_width=True)
                with dd4: st.download_button("♻️ 재활용", data=create_secure_excel(df_school[["날짜","학교명","재활용(kg)","재활용수익"]],"재활용 정산(청구)서"), file_name=f"{school}_재활용_월간정산서.xlsx", use_container_width=True)

            with d2:
                st.info("교육청·지자체 제출용 [폐기물관리법 시행규칙 별지 제30호서식]")
                dr1, dr2, dr3 = st.columns(3)
                with dr1: st.download_button("🗑️ 음식물", data=create_secure_excel(df_school[["날짜","학교명","음식물(kg)"]],"음식물 배출 및 처리 실적보고"), file_name=f"{school}_음식물_실적보고서.xlsx", use_container_width=True)
                with dr2: st.download_button("🗄️ 사업장", data=create_secure_excel(df_school[["날짜","학교명","사업장(kg)"]],"사업장 배출 및 처리 실적보고"), file_name=f"{school}_사업장_실적보고서.xlsx", use_container_width=True)
                with dr3: st.download_button("♻️ 재활용", data=create_secure_excel(df_school[["날짜","학교명","재활용(kg)"]],"재활용 배출 및 처리 실적보고"), file_name=f"{school}_재활용_실적보고서.xlsx", use_container_width=True)

            with d3:
                st.info("재활용품 판매 수익으로 처리비용을 상계(차감)한 내역 증빙")
                st.download_button("📄 상계처리 증빙서",
                    data=create_secure_excel(df_school[["날짜","학교명","재활용(kg)","재활용수익"]],"사업장 폐기물 재활용 상계처리 증빙"),
                    file_name=f"{school}_상계증빙.xlsx")
                # 상계 요약
                total_recycle = df_school["재활용수익"].sum()
                total_cost    = df_school["음식물비용"].sum() + df_school["사업장비용"].sum()
                net_cost      = total_cost - total_recycle
                rr1, rr2, rr3 = st.columns(3)
                rr1.metric("총 처리비용", f"{total_cost:,.0f}원")
                rr2.metric("재활용 상계액", f"-{total_recycle:,.0f}원", delta=f"-{total_recycle:,.0f}")
                rr3.metric("실 청구금액", f"{net_cost:,.0f}원")

            with d4:
                st.info("한국환경공단 올바로(Allbaro) 시스템 전자인계서 연동")
                st.markdown("""
                **올바로 시스템 연동 데이터 미리보기:**
                """)
                preview_df = df_school[["날짜","학교명","음식물(kg)","재활용(kg)","사업장(kg)","상태"]].tail(5)
                st.dataframe(preview_df, use_container_width=True)
                if st.button("🔗 올바로시스템 전송 및 자동결재", type="primary", use_container_width=True):
                    with st.spinner("한국환경공단 서버와 통신 중..."):
                        time.sleep(2)
                    st.success(f"✅ {len(df_school)}건의 전자인계서가 올바로시스템에 이관 완료되었습니다.")
                    st.info("📋 인계번호: " + f"HY-{datetime.now().strftime('%Y%m%d')}-{len(df_school):04d}")
    else:
        st.info("해당 학교의 수거 데이터가 아직 없습니다.")


# ============================================================
# [모드 3] 수거 기사 (현장 앱)  — [추가3] 스쿨존 실제 시간 차단
# ============================================================
elif role == "🚚 수거 기사 (현장 앱)":
    _, mid, _ = st.columns([1, 2, 1])
    with mid:
        st.markdown('<div class="mobile-app-header"><h2 style="margin:0;font-size:22px;">🚚 하영자원 기사 전용 앱</h2></div>', unsafe_allow_html=True)

        # [추가3] 등하교 시간대 자동 차단
        restricted, time_range = is_schoolzone_restricted()
        if restricted:
            st.markdown(f'<div class="schoolzone-danger">🚨 등하교 시간 ({time_range})<br>수거 작업이 제한됩니다<br><span style="font-size:16px;">학교 구역 진입 금지</span></div>', unsafe_allow_html=True)
            st.error("⛔ 현재 등하교 시간대입니다. 수거 데이터 입력이 잠깁니다.")
            st.warning(f"다음 수거 가능 시간: 오전 09:00 이후 또는 오후 16:00 이후")
            # 잠금 상태에서도 안전 점검은 가능
        else:
            st.markdown('<div class="schoolzone-safe">✅ 수거 가능 시간 — 안전 운행하세요</div>', unsafe_allow_html=True)

        # 현재 시간 표시
        st.caption(f"🕐 현재 시각: {datetime.now().strftime('%H:%M:%S')}")

        # 안전 점검 체크리스트
        with st.expander("📋 [필수] 운행 전 안전 점검 리스트", expanded=True):
            st.warning("어린이 안전을 위해 아래 항목을 확인해 주세요.")
            c1 = st.checkbox("차량 후방 카메라 정상 작동 확인")
            c2 = st.checkbox("조수석 안전 요원 탑승 여부 확인")
            c3 = st.checkbox("스쿨존 서행(30km 이하) 운전 숙지")
            c4 = st.checkbox("등하교 시간대 학교 구역 진입 금지 숙지")
            all_checked = c1 and c2 and c3 and c4
            if all_checked:
                st.success("✅ 안전 점검 완료! 오늘도 안전 운행하세요.")

        st.write("---")

        # [추가3] GPS 스쿨존 진입 시뮬레이터
        st.markdown("#### 📍 스쿨존 진입 감지")
        sz_col1, sz_col2 = st.columns(2)
        with sz_col1:
            in_schoolzone = st.toggle("🚨 스쿨존 진입 (GPS 시뮬레이션)")
        with sz_col2:
            current_speed = st.number_input("현재 속도 (km/h)", min_value=0, max_value=100, value=30, step=5)

        if in_schoolzone:
            if current_speed > 30:
                st.markdown(f'<div class="schoolzone-danger">⚠️ 과속 감지! 현재 {current_speed}km/h<br>즉시 감속하세요 → 30km/h 이하</div>', unsafe_allow_html=True)
                # 과속 자동 기록 (실제 운용 시 DB 저장)
                st.error("🚨 과속 기록이 본사에 자동 전송됩니다.")
            else:
                st.markdown(f'<div class="schoolzone-safe">🏫 스쿨존 내 정상 운행 중 ({current_speed}km/h)</div>', unsafe_allow_html=True)

        st.write("---")

        # 오늘 내 배차 일정 표시
        today_str = datetime.now().strftime("%Y-%m-%d")
        today_sched = get_schedules_month(datetime.now().year, datetime.now().month)
        today_sched = today_sched[today_sched["날짜"]==today_str]
        if not today_sched.empty:
            st.markdown("#### 📋 오늘 수거 일정")
            for _, row in today_sched.iterrows():
                done_icon = "✅" if row["완료여부"] else "⏳"
                st.write(f"{done_icon} {row['학교명']} — {row['메모']}")

        st.write("---")
        st.camera_input("📸 현장 증빙 사진 촬영 (선택사항)")

        # 수거 입력 폼 - 등하교 시간 차단
        st.write("---")
        if restricted:
            st.markdown(f'<div class="alert-box">⛔ 등하교 시간({time_range})에는 수거 입력이 제한됩니다.<br>09:00 또는 16:00 이후 입력해 주세요.</div>', unsafe_allow_html=True)
        else:
            with st.form("driver_input"):
                target = st.selectbox("수거 완료한 학교", SCHOOL_LIST)
                fi1, fi2, fi3 = st.columns(3)
                with fi1: food_w = st.number_input("음식물 (kg)", min_value=0, step=10)
                with fi2: biz_w  = st.number_input("사업장 (kg)", min_value=0, step=10)
                with fi3: re_w   = st.number_input("재활용 (kg)", min_value=0, step=10)
                driver_memo = st.text_input("특이사항 메모", placeholder="대용량 배출, 분리수거 불량 등")

                submitted = st.form_submit_button("📤 본사로 수거량 전송", type="primary", use_container_width=True)
                if submitted:
                    if not all_checked:
                        st.error("안전 점검을 먼저 완료해 주세요.")
                    elif food_w > 0 or biz_w > 0 or re_w > 0:
                        save_collection({
                            "날짜":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "학교명":   target,
                            "학생수":   STUDENT_COUNTS[target],
                            "수거업체": "하영자원(본사 직영)",
                            "음식물_kg": food_w,
                            "재활용_kg": re_w,
                            "사업장_kg": biz_w,
                            "상태":     "정산대기",
                            "현장사진": driver_memo or "",
                        })
                        # 일정 완료 처리
                        conn = get_conn()
                        conn.execute(
                            "UPDATE schedules SET 완료여부=1 WHERE 날짜=? AND 학교명=?",
                            (today_str, target)
                        )
                        conn.commit()
                        conn.close()

                        if get_setting("kakao_notify_enabled") == "true":
                            conn2 = get_conn()
                            pr = conn2.execute(
                                "SELECT 담당자연락처,음식물단가,사업장단가 FROM school_prices WHERE 학교명=?", (target,)
                            ).fetchone()
                            conn2.close()
                            if pr and pr[0]:
                                fp = pr[1] or int(get_setting("default_food_price"))
                                bp = pr[2] or int(get_setting("default_biz_price"))
                                send_kakao_alimtalk(pr[0], target, food_w, int(food_w*fp + biz_w*bp))

                        st.success(f"✅ {target} 수거 실적이 기록되었습니다.")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.warning("수거한 중량(kg)을 먼저 입력해 주세요.")


# ============================================================
# [모드 4] 교육청 통합 관제  [추가5]
# ============================================================
elif role == "🏛️ 교육청 관제 (신규)":
    st.markdown('<span class="badge-v3">✨ v3 신규</span>', unsafe_allow_html=True)
    st.title("🏛️ 교육청 통합 관제 대시보드")
    st.markdown("<p style='color:#5f6368;font-size:16px;'>관할 학교 전체의 폐기물 현황·ESG 성과·예산절감을 실시간으로 모니터링합니다.</p>", unsafe_allow_html=True)

    # 로그인된 교육청 자동 설정 (selectbox 제거)
    # 관리자가 교육청 모드로 접근하면 첫 번째 교육청을 기본값으로 사용
    _login_org = st.session_state.user_org
    if _login_org in EDU_OFFICES:
        sel_edu = _login_org
    else:
        sel_edu = list(EDU_OFFICES.keys())[0]
    edu_schools = EDU_OFFICES[sel_edu]
    df_edu = df_all[df_all["학교명"].isin(edu_schools)]

    if df_edu.empty:
        st.info("해당 교육청 데이터가 없습니다.")
    else:
        # KPI
        e1, e2, e3, e4, e5 = st.columns(5)
        with e1: st.markdown(f'<div class="custom-card custom-card-red"><div class="metric-title">🗑️ 음식물 수거</div><div class="metric-value-food">{df_edu["음식물(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
        with e2: st.markdown(f'<div class="custom-card custom-card-green"><div class="metric-title">♻️ 재활용 수거</div><div class="metric-value-recycle">{df_edu["재활용(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
        with e3: st.markdown(f'<div class="custom-card custom-card-purple"><div class="metric-title">🗄️ 사업장 수거</div><div class="metric-value-biz">{df_edu["사업장(kg)"].sum():,.0f} kg</div></div>', unsafe_allow_html=True)
        co2_edu = df_edu["탄소감축량(kg)"].sum()
        trees_edu = int(co2_edu / 6.6)
        with e4: st.markdown(f'<div class="custom-card custom-card-green"><div class="metric-title">🌍 CO₂ 감축</div><div class="metric-value-recycle">{co2_edu:,.0f} kg</div></div>', unsafe_allow_html=True)
        budget_edu = len(edu_schools) * int(get_setting("budget_saving_per_school") or 5200000)
        with e5: st.markdown(f'<div class="custom-card custom-card-teal"><div class="metric-title">💰 예산절감(연)</div><div class="metric-value-total" style="color:#00897b;">{budget_edu:,.0f}원</div></div>', unsafe_allow_html=True)

        # ESG 종합 배너
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#1a73e8,#0d47a1);padding:20px;border-radius:12px;color:white;margin:10px 0;display:flex;justify-content:space-around;">
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;opacity:.9;">관할 학교 수</p>
                <h2 style="margin:0;color:white;">{len(edu_schools)}개교</h2>
            </div>
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;opacity:.9;">🌲 소나무 식재 효과</p>
                <h2 style="margin:0;color:white;">{trees_edu:,}그루</h2>
            </div>
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;opacity:.9;">어린이사고예방(예상)</p>
                <h2 style="margin:0;color:white;">50% 감소</h2>
            </div>
            <div style="text-align:center;">
                <p style="margin:0;font-size:13px;opacity:.9;">담당자 업무시간 절감</p>
                <h2 style="margin:0;color:white;">90%</h2>
            </div>
        </div>
        """, unsafe_allow_html=True)

        edu_tab1, edu_tab2, edu_tab3, edu_tab4 = st.tabs([
            "📊 학교별 현황", "📈 월별 추이", "🛡️ 안전 현황", "📄 교육청 보고서"
        ])

        with edu_tab1:
            # 학교별 집계
            school_summary = df_edu.groupby("학교명").agg(
                음식물=("음식물(kg)", "sum"),
                재활용=("재활용(kg)", "sum"),
                사업장=("사업장(kg)", "sum"),
                청구금액=("최종정산액", "sum"),
                탄소감축=("탄소감축량(kg)", "sum"),
                수거횟수=("id", "count")
            ).reset_index()
            school_summary["예산절감(연)"] = int(get_setting("budget_saving_per_school") or 5200000)
            st.dataframe(school_summary, use_container_width=True)

            # 학교별 수거량 막대 차트
            st.bar_chart(school_summary.set_index("학교명")[["음식물","재활용","사업장"]])

        with edu_tab2:
            monthly_edu = df_edu.groupby("월별").agg(
                음식물=("음식물(kg)", "sum"),
                재활용=("재활용(kg)", "sum"),
                사업장=("사업장(kg)", "sum"),
                청구금액=("최종정산액", "sum")
            ).reset_index()
            st.line_chart(monthly_edu.set_index("월별")[["음식물","재활용","사업장"]])
            st.dataframe(monthly_edu, use_container_width=True)

        with edu_tab3:
            st.subheader("🛡️ 스쿨존 안전 관제 현황")
            restricted, time_range = is_schoolzone_restricted()
            if restricted:
                st.markdown(f'<div class="schoolzone-danger">🚨 현재 등하교 시간({time_range}) — 전체 {len(edu_schools)}개교 수거 차량 진입 제한 중</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="schoolzone-safe">✅ 현재 수거 가능 시간 — 전체 학교 정상 운행 가능</div>', unsafe_allow_html=True)

            # 안전 통계 (학교안전사고현황 기반 데이터)
            safety_data = pd.DataFrame({
                "연도": ["2021년","2022년","2023년","2024년"],
                "사고건수": [523, 514, 523, 526],
                "부상자": [563, 529, 523, 556],
                "사망자": [2, 3, 2, 2]
            })
            st.markdown("**📊 전국 학교 폐기물 차량 관련 안전사고 현황 (참고)**")
            st.dataframe(safety_data, use_container_width=True)
            st.bar_chart(safety_data.set_index("연도")["사고건수"])
            st.caption("출처: 정을호 국회의원실 자료 / 본 플랫폼 도입 시 50% 감소 목표")

        with edu_tab4:
            st.subheader("📄 교육청 제출용 통합 보고서")
            st.info("아래 보고서를 다운로드하여 교육청 행정 제출에 사용하세요.")

            rp1, rp2, rp3 = st.columns(3)
            with rp1:
                esg_report = school_summary[["학교명","탄소감축","예산절감(연)","수거횟수"]]
                esg_report = esg_report.copy()
                esg_report["소나무환산(그루)"] = (esg_report["탄소감축"] / 6.6).astype(int)
                st.download_button(
                    "🌍 ESG 성과 보고서",
                    data=create_secure_excel(esg_report, f"{sel_edu} ESG 성과 보고서"),
                    file_name=f"{sel_edu}_ESG보고서_{datetime.now().strftime('%Y-%m')}.xlsx",
                    use_container_width=True
                )
            with rp2:
                budget_report = school_summary[["학교명","청구금액","수거횟수"]].copy()
                budget_report["예산절감(연)"] = int(get_setting("budget_saving_per_school") or 5200000)
                budget_report["합계절감"] = budget_report["예산절감(연)"]
                st.download_button(
                    "💰 예산절감 효과 보고서",
                    data=create_secure_excel(budget_report, f"{sel_edu} 예산절감 효과 보고서"),
                    file_name=f"{sel_edu}_예산절감보고서_{datetime.now().strftime('%Y-%m')}.xlsx",
                    use_container_width=True
                )
            with rp3:
                collect_report = df_edu[["날짜","학교명","음식물(kg)","재활용(kg)","사업장(kg)","최종정산액","상태"]]
                st.download_button(
                    "📊 수거 실적 통합 보고서",
                    data=create_secure_excel(collect_report, f"{sel_edu} 수거실적 통합 보고서"),
                    file_name=f"{sel_edu}_수거실적보고서_{datetime.now().strftime('%Y-%m')}.xlsx",
                    use_container_width=True
                )


# ============================================================
# [섹션 E] 📦 계약서류 패키지 생성 UI (관리자 전용)
# ============================================================
elif role == "📦 계약서류 패키지 생성":
    st.title("📦 계약서류 패키지 자동 생성")
    st.markdown(
        "<p style='color:#5f6368;'>학교 정보 입력 → PDF 3종 자동생성 + 기존 서류 포함 → ZIP 다운로드</p>",
        unsafe_allow_html=True
    )

    # ── 서류 유효기간 D-day 알림 ───────────────────────────
    st.subheader("📋 서류 유효기간 현황")
    today_dt = date.today()
    alert_rows = []
    for doc in DOC_EXPIRE:
        exp = date.fromisoformat(doc["expire"])
        if exp.year == 9999:
            dday_str = "무기한"
            status   = "🟢 정상"
        else:
            diff     = (exp - today_dt).days
            dday_str = f"D-{diff}" if diff >= 0 else f"D+{abs(diff)} 만료"
            if diff < 0:
                status = "⛔ 만료됨"
            elif diff <= 30:
                status = "🔴 만료임박"
            elif diff <= 60:
                status = "🟡 주의"
            else:
                status = "🟢 정상"
        alert_rows.append({
            "서류명": doc["name"],
            "만료일": doc["expire"] if doc["expire"] != "9999-12-31" else "-",
            "D-day": dday_str,
            "상태":  status,
            "갱신처": doc["renew_url"],
        })

    df_doc = pd.DataFrame(alert_rows)
    # 만료임박/만료 우선 정렬
    sort_key = {"⛔ 만료됨": 0, "🔴 만료임박": 1, "🟡 주의": 2, "🟢 정상": 3}
    df_doc["_sort"] = df_doc["상태"].map(sort_key)
    df_doc = df_doc.sort_values("_sort").drop(columns="_sort")
    st.dataframe(df_doc, use_container_width=True, hide_index=True)

    # 만료임박 경고
    warn_docs = [r for r in alert_rows if "🔴" in r["상태"] or "⛔" in r["상태"]]
    if warn_docs:
        for w in warn_docs:
            st.warning(f"⚠️ **{w['서류명']}** {w['D-day']} — 갱신 필요: {w['갱신처']}")

    st.divider()

    # ── 계약 정보 입력 폼 ──────────────────────────────────
    st.subheader("🏫 계약 정보 입력")

    # 자주 쓰는 학교 빠른선택
    SCHOOLS_PRESET = {
        "직접입력": {},
        "당곡고등학교": {
            "biz_no": "", "addr": "서울특별시 관악구",
            "tel": "", "contract_no": "R26TA01543339 00"
        },
        "서초고등학교": {
            "biz_no": "210-83-00086",
            "addr": "서울특별시 서초구 반포대로27길 29",
            "tel": "02-580-3891", "contract_no": ""
        },
    }

    col_l, col_r = st.columns([1, 2])
    with col_l:
        preset = st.selectbox("빠른선택 학교", list(SCHOOLS_PRESET.keys()))

    preset_data = SCHOOLS_PRESET.get(preset, {})

    col1, col2 = st.columns(2)
    with col1:
        school_nm   = st.text_input("학교명 *",
            value=preset if preset != "직접입력" else "")
        school_bno  = st.text_input("학교 사업자번호",
            value=preset_data.get("biz_no",""))
        school_addr = st.text_input("학교 주소",
            value=preset_data.get("addr",""))
        school_tel  = st.text_input("학교 전화",
            value=preset_data.get("tel",""))
    with col2:
        start_dt    = st.text_input("계약 시작일 (YYYY-MM-DD)", "2026-03-01")
        end_dt      = st.text_input("계약 종료일 (YYYY-MM-DD)", "2027-02-28")
        unit_p      = st.number_input("단가 (원/L)", value=180, step=10)
        volume      = st.number_input("월 예상 수거량 (L)", value=0.0, step=100.0,
                                       help="0이면 견적서 수량란 공백 처리")
        amt_str     = st.text_input("계약 총금액 (표시용)", "")

    # ── 생성 버튼 ──────────────────────────────────────────
    st.divider()
    if st.button("🚀 계약서류 패키지 ZIP 생성", type="primary",
                 use_container_width=True):
        if not school_nm.strip():
            st.error("❌ 학교명을 입력해주세요.")
        else:
            with st.spinner("📄 PDF 생성 중... (3~10초 소요)"):
                try:
                    zip_bytes, errs = generate_contract_package(
                        school_name      = school_nm.strip(),
                        school_biz_no    = school_bno.strip(),
                        school_addr      = school_addr.strip(),
                        school_tel       = school_tel.strip(),
                        start_date       = start_dt.strip(),
                        end_date         = end_dt.strip(),
                        volume_l         = float(volume),
                        unit_price       = int(unit_p),
                        contract_amount  = amt_str.strip(),
                    )
                    st.session_state["pkg_zip"]    = zip_bytes
                    st.session_state["pkg_school"] = school_nm.strip()
                    st.session_state["pkg_errors"] = errs
                    st.success(f"✅ 패키지 생성 완료! ({len(zip_bytes)/1024:.1f} KB)")
                except Exception as e:
                    st.error(f"❌ 생성 오류: {e}")

    # ── 다운로드 버튼 (session_state 유지) ────────────────
    pkg = st.session_state.get("pkg_zip")
    pkg_school = st.session_state.get("pkg_school", "학교")
    pkg_errors = st.session_state.get("pkg_errors", {})

    if pkg:
        fname_zip = f"계약서류패키지_{pkg_school}_{date.today().strftime('%Y%m%d')}.zip"
        st.download_button(
            label    = f"📥 {fname_zip} 다운로드",
            data     = pkg,
            file_name= fname_zip,
            mime     = "application/zip",
            key      = "dl_contract_pkg",
            use_container_width=True,
            type     = "primary",
        )

        # 포함 서류 목록 표시
        with st.expander("📂 포함 서류 목록 보기"):
            st.markdown("""
**[자동 생성 PDF]**
- 📄 음식물 견적서
- 📄 폐기물 위수탁 계약서
- 📄 계약이행 통합 서약서

**[기존 서류]**
- 🖼️ 사업자등록증
- 🖼️ 폐기물수집운반업 허가증
- 📄 소상공인 확인서 ⚠️ 2026-03-31 만료
- 📄 창업기업 확인서
- 📄 재해율 확인서
- 🖼️ 사업자 계좌 통장사본
- 📋 서류 체크리스트 (자동생성)
            """)

        if pkg_errors:
            with st.expander("⚠️ PDF 생성 오류 목록"):
                for k, v in pkg_errors.items():
                    st.error(f"**{k}**: {v}")
                st.info("💡 한글 폰트(malgun.ttf)가 없으면 일부 PDF에서 한글이 깨질 수 있습니다.")

    # ── 하영자원 정보 요약 ─────────────────────────────────
    with st.expander("ℹ️ 하영자원 고정 정보 확인"):
        df_hy = pd.DataFrame([
            {"항목": k, "내용": v} for k, v in HY.items()
        ])
        st.dataframe(df_hy, use_container_width=True, hide_index=True)
        st.caption("※ 위 정보는 모든 서류에 자동으로 입력됩니다.")

    # ── [2단계] 계약중 학교 전체 일괄 생성 ───────────────────
    st.divider()
    st.subheader("🚀 계약중 학교 전체 패키지 일괄 생성")
    st.caption("DB에서 계약상태 = '계약중' 학교를 자동 조회하여 일괄 생성합니다.")

    # 계약중 학교 목록 미리보기
    try:
        active_schools = [s for s in a_get_all_schools() if s["계약상태"] == "계약중"]
    except Exception:
        active_schools = []

    if not active_schools:
        st.info("⚠️ 현재 계약상태가 '계약중'인 학교가 없습니다. 🏫 학교 마스터 관리에서 계약 상태를 먼저 설정해주세요.")
    else:
        with st.expander(f"📋 일괄 생성 대상 학교 ({len(active_schools)}개) 확인"):
            df_active = pd.DataFrame([{
                "학교명":    s["학교명"],
                "단가(원/L)": s["음식물단가"],
                "계약기간":  f"{s['계약시작']} ~ {s['계약종료']}",
                "교육청":    s["교육청"],
            } for s in active_schools])
            st.dataframe(df_active, use_container_width=True, hide_index=True)

        if st.button(
            f"🚀 계약중 {len(active_schools)}개 학교 전체 패키지 일괄 생성",
            type="primary", use_container_width=True, key="bulk_gen"
        ):
            master_buf  = io.BytesIO()
            results     = []   # (학교명, 성공여부, 크기, 오류)
            today_str   = date.today().strftime("%Y%m%d")

            prog_bar  = st.progress(0, text="준비 중...")
            status_ph = st.empty()

            with zipfile.ZipFile(master_buf, "w", zipfile.ZIP_DEFLATED) as master_zip:
                for idx, school in enumerate(active_schools):
                    sname = school["학교명"]
                    status_ph.info(f"⏳ [{idx+1}/{len(active_schools)}] {sname} 처리 중...")
                    try:
                        zip_bytes, errs = generate_contract_package(
                            school_name     = sname,
                            school_biz_no   = school.get("사업자번호", ""),
                            school_addr     = school.get("주소", ""),
                            school_tel      = school.get("전화", ""),
                            start_date      = school.get("계약시작", "2026-03-01"),
                            end_date        = school.get("계약종료", "2027-02-28"),
                            volume_l        = float(school.get("월_예상량_L", 0) or 0),
                            unit_price      = int(school.get("음식물단가", 150)),
                            contract_amount = "",
                        )
                        arc_name = f"{sname}_{today_str}.zip"
                        master_zip.writestr(arc_name, zip_bytes)
                        results.append((sname, True, len(zip_bytes), errs))
                    except Exception as e:
                        results.append((sname, False, 0, {"오류": str(e)}))

                # 전체 결과 요약 TXT
                summary_lines = [
                    f"하영자원 계약서류 일괄 생성 결과",
                    f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"총 {len(active_schools)}개 학교",
                    "=" * 50,
                ]
                for sname, ok, sz, errs in results:
                    status_icon = "✅" if ok else "❌"
                    sz_str      = f"{sz/1024:.1f}KB" if ok else "실패"
                    err_str     = f" ({list(errs.keys())})" if errs else ""
                    summary_lines.append(f"{status_icon} {sname} — {sz_str}{err_str}")

                ok_cnt   = sum(1 for _, ok, _, _ in results if ok)
                fail_cnt = len(results) - ok_cnt
                summary_lines += ["=" * 50,
                    f"성공: {ok_cnt}개  실패: {fail_cnt}개"]
                master_zip.writestr("00_일괄생성결과.txt",
                                    "\n".join(summary_lines))

            prog_bar.progress(1.0, text="✅ 완료!")
            status_ph.empty()

            # session_state 저장
            st.session_state["bulk_zip"]     = master_buf.getvalue()
            st.session_state["bulk_results"] = results
            st.session_state["bulk_ok"]      = ok_cnt
            st.session_state["bulk_fail"]    = fail_cnt

            if fail_cnt == 0:
                st.success(f"✅ 전체 {ok_cnt}개 학교 패키지 생성 완료!")
            else:
                st.warning(f"⚠️ 성공 {ok_cnt}개 / 실패 {fail_cnt}개")

        # 다운로드 버튼
        bulk_zip = st.session_state.get("bulk_zip")
        if bulk_zip:
            bulk_fname = f"하영자원_전체계약서류_{date.today().strftime('%Y%m%d')}.zip"
            st.download_button(
                label         = f"📥 {bulk_fname} 다운로드 ({len(bulk_zip)//1024}KB)",
                data          = bulk_zip,
                file_name     = bulk_fname,
                mime          = "application/zip",
                key           = "bulk_dl",
                use_container_width=True,
                type          = "primary",
            )

            # 결과 상세 테이블
            bulk_res = st.session_state.get("bulk_results", [])
            if bulk_res:
                with st.expander("📋 학교별 생성 결과 상세"):
                    df_res = pd.DataFrame([{
                        "학교명": r[0],
                        "결과":   "✅ 성공" if r[1] else "❌ 실패",
                        "파일크기": f"{r[2]//1024}KB" if r[1] else "-",
                        "오류":   str(list(r[3].keys())) if r[3] else "-",
                    } for r in bulk_res])
                    st.dataframe(df_res, use_container_width=True, hide_index=True)


# ============================================================
# [섹션 D] 📄 위수탁계약서 단독 작성 UI (관리자 전용)
# ============================================================
elif role == "📄 위수탁계약서 작성":
    st.title("📄 폐기물 위수탁 운반 처리 계약서")
    st.markdown(
        "<p style='color:#5f6368;'>학교 정보 입력 → 계약서 PDF 즉시 생성 · 다운로드</p>",
        unsafe_allow_html=True
    )

    # ── 빠른 선택 ────────────────────────────────────────────
    D_SCHOOLS = {
        "직접입력":     {"biz_no": "", "addr": "", "tel": "", "waste": "음식물류폐기물"},
        "서초고등학교": {
            "biz_no": "210-83-00086",
            "addr":   "서울특별시 서초구 반포대로27길 29",
            "tel":    "02-580-3891",
            "waste":  "음식물류폐기물",
        },
        "당곡고등학교": {
            "biz_no": "",
            "addr":   "서울특별시 관악구",
            "tel":    "",
            "waste":  "음식물류폐기물",
        },
    }

    d_preset = st.selectbox("🏫 빠른선택 학교", list(D_SCHOOLS.keys()), key="d_preset")
    d_data   = D_SCHOOLS[d_preset]

    st.divider()

    # ── 입력 폼 ──────────────────────────────────────────────
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("##### 📌 배출자 (갑) 정보")
        d_school   = st.text_input("학교명 *",
            value=d_preset if d_preset != "직접입력" else "",
            key="d_school")
        d_biz_no   = st.text_input("사업자번호",
            value=d_data["biz_no"], key="d_biz_no")
        d_addr     = st.text_input("학교 주소",
            value=d_data["addr"], key="d_addr")
        d_tel      = st.text_input("학교 전화번호",
            value=d_data["tel"], key="d_tel")

    with col2:
        st.markdown("##### 📋 계약 내용")
        d_start    = st.text_input("계약 시작일 (YYYY-MM-DD)", "2026-03-01", key="d_start")
        d_end      = st.text_input("계약 종료일 (YYYY-MM-DD)", "2027-02-28", key="d_end")
        d_waste    = st.text_input("폐기물 종류", value=d_data["waste"], key="d_waste")
        d_volume   = st.text_input("물량 (예: 월 500L 내외)", "", key="d_volume")
        d_unit     = st.number_input("단가 (원/L)", value=180, step=10, key="d_unit")
        d_amount   = st.text_input("계약 총금액 (표시용, 예: 1,080,000원)", "", key="d_amount")

    st.divider()

    # ── 미리보기 요약 ─────────────────────────────────────────
    with st.expander("📋 입력 내용 미리보기"):
        prev = {
            "배출자(갑)": d_school, "사업자번호": d_biz_no,
            "주소": d_addr, "전화": d_tel,
            "계약기간": f"{d_start} ~ {d_end}",
            "폐기물 종류": d_waste, "물량": d_volume,
            "단가": f"{d_unit}원/L", "계약금액": d_amount,
            "운반자(을)": HY["name"], "처리업체": HY["processor"],
        }
        df_prev = pd.DataFrame(list(prev.items()), columns=["항목", "내용"])
        st.dataframe(df_prev, use_container_width=True, hide_index=True)

    # ── 생성 버튼 ─────────────────────────────────────────────
    if st.button("📄 위수탁계약서 PDF 생성", type="primary",
                 use_container_width=True, key="d_gen"):
        if not d_school.strip():
            st.error("❌ 학교명을 입력해주세요.")
        else:
            with st.spinner("📄 PDF 생성 중..."):
                try:
                    pdf_path = generate_contract_doc_pdf(
                        school_name    = d_school.strip(),
                        school_biz_no  = d_biz_no.strip(),
                        school_addr    = d_addr.strip(),
                        school_tel     = d_tel.strip(),
                        start_date     = d_start.strip(),
                        end_date       = d_end.strip(),
                        waste_type     = d_waste.strip() or "음식물류폐기물",
                        volume_str     = d_volume.strip(),
                        unit_price     = int(d_unit),
                        contract_amount= d_amount.strip(),
                    )
                    with open(pdf_path, "rb") as f:
                        st.session_state["d_pdf_bytes"] = f.read()
                    st.session_state["d_pdf_name"] = (
                        f"위수탁계약서_{d_school.strip()}_{date.today().strftime('%Y%m%d')}.pdf"
                    )
                    st.success("✅ PDF 생성 완료!")
                except Exception as e:
                    st.error(f"❌ 오류: {e}")
                    st.info("💡 한글 폰트(malgun.ttf)가 설치되어 있는지 확인하세요.")

    # ── 다운로드 ──────────────────────────────────────────────
    d_pdf  = st.session_state.get("d_pdf_bytes")
    d_name = st.session_state.get("d_pdf_name", "위수탁계약서.pdf")

    if d_pdf:
        st.download_button(
            label         = f"📥 {d_name} 다운로드",
            data          = d_pdf,
            file_name     = d_name,
            mime          = "application/pdf",
            key           = "d_dl",
            use_container_width=True,
            type          = "primary",
        )
        st.caption("💡 다운로드 후 출력 → 학교장 / 하영자원 대표자 각 1부 보관")

# ── [신규 2단계] 학교로 초안 전송 기능 ──────────────────────
        st.divider()
        st.markdown("##### 📤 학교 행정실로 전자문서(초안) 전송")
        st.info(f"작성된 '{d_school}' 위수탁계약서 초안을 해당 학교 시스템으로 보냅니다.")
        if st.button(f"📩 {d_school} 행정실로 전송하기", type="primary", use_container_width=True, key="send_draft"):
            try:
                conn = get_conn()
                # 학교 행정실이 볼 수 있도록 DB에 '초안 수신' 상태로 기록
                conn.execute(
                    """INSERT INTO contract_master 
                       (학교명, 계약_시작일, 계약_종료일, 폐기물_종류, 단가, 계약_상태, 계약서_상태, 비고, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (d_school.strip(), d_start.strip(), d_end.strip(), d_waste.strip(), int(d_unit), 
                     "협의중", "수신", "하영자원 본사에서 초안 발송함", 
                     datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%Y-%m-%d"))
                )
                conn.commit()
                st.success(f"✅ {d_school} 행정실 [전자 계약함]으로 발송이 완료되었습니다!")
            except Exception as e:
                st.error(f"❌ 전송 오류: {e}")
        # ────────────────────────────────────────────────────────

    # ── 하영자원(을) 정보 요약 ────────────────────────────────
    with st.expander("ℹ️ 운반자(을) 하영자원 고정 정보"):
        df_hy2 = pd.DataFrame([{"항목": k, "내용": v} for k, v in HY.items()])
        st.dataframe(df_hy2, use_container_width=True, hide_index=True)
        st.caption("※ 위 정보는 계약서에 자동으로 입력됩니다.")


# ============================================================
# [섹션 C] 📋 서류 유효기간 관리 UI (관리자 전용)
# ============================================================
elif role == "📋 서류 유효기간 관리":
    st.title("📋 서류 유효기간 관리")
    st.markdown(
        "<p style='color:#5f6368;'>계약 서류 만료일 추적 · 갱신 관리 · 알림 센터</p>",
        unsafe_allow_html=True
    )

    # ── 상단 요약 카드 ────────────────────────────────────────
    all_docs = c_get_all_docs()
    n_exp    = sum(1 for d in all_docs if d["상태"] == "⛔ 만료됨")
    n_red    = sum(1 for d in all_docs if d["상태"] == "🔴 만료임박")
    n_yel    = sum(1 for d in all_docs if d["상태"] == "🟡 주의")
    n_ok     = sum(1 for d in all_docs if d["상태"] == "🟢 정상")

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("⛔ 만료됨",   n_exp,  delta=None)
    mc2.metric("🔴 만료임박", n_red,  delta=None)
    mc3.metric("🟡 주의",     n_yel,  delta=None)
    mc4.metric("🟢 정상",     n_ok,   delta=None)

    # 긴급 배너
    urgent = [d for d in all_docs if d["상태"] in ("⛔ 만료됨", "🔴 만료임박") and not d["갱신완료"]]
    if urgent:
        st.error(f"🚨 **즉시 조치 필요한 서류 {len(urgent)}건**")
        for u in urgent:
            st.warning(f"**{u['서류명']}** — {u['상태']}  |  만료일: {u['만료일']}  |  {u['D-day']}  |  갱신처: {u['갱신처']}")
    else:
        st.success("✅ 만료임박 서류 없음")

    st.divider()

    # ── 서류 목록 테이블 ──────────────────────────────────────
    st.subheader("📑 전체 서류 목록")

    # 필터
    cf1, cf2 = st.columns([2, 1])
    with cf1:
        filter_status = st.multiselect(
            "상태 필터", ["⛔ 만료됨", "🔴 만료임박", "🟡 주의", "🟢 정상"],
            default=["⛔ 만료됨", "🔴 만료임박", "🟡 주의", "🟢 정상"],
            key="c_filter"
        )
    with cf2:
        hide_renewed = st.checkbox("갱신완료 서류 숨기기", value=False, key="c_hide_renewed")

    display_docs = [
        d for d in all_docs
        if d["상태"] in filter_status
        and not (hide_renewed and d["갱신완료"])
    ]

    if not display_docs:
        st.info("해당 조건의 서류가 없습니다.")
    else:
        df_c = pd.DataFrame([{
            "ID":    d["id"],
            "서류명": d["서류명"],
            "만료일": d["만료일"],
            "D-day": d["D-day"],
            "상태":  d["상태"],
            "갱신처": d["갱신처"],
            "갱신완료": "✅" if d["갱신완료"] else "□",
            "메모":  d["메모"],
        } for d in display_docs])

        st.dataframe(df_c, use_container_width=True, hide_index=True,
                     column_config={
                         "ID":     st.column_config.NumberColumn(width="small"),
                         "서류명": st.column_config.TextColumn(width="medium"),
                         "D-day":  st.column_config.TextColumn(width="small"),
                         "상태":   st.column_config.TextColumn(width="small"),
                         "갱신완료": st.column_config.TextColumn(width="small"),
                     })

    st.divider()

    # ── 서류 수정 폼 ──────────────────────────────────────────
    st.subheader("✏️ 서류 정보 수정")

    if all_docs:
        doc_names = {d["서류명"]: d for d in all_docs}
        sel_doc_name = st.selectbox("수정할 서류 선택", list(doc_names.keys()), key="c_sel")
        sel_doc = doc_names[sel_doc_name]

        ec1, ec2 = st.columns(2)
        with ec1:
            e_issued  = st.text_input("발급일 (YYYY-MM-DD)",
                value=sel_doc["발급일"] or "", key="c_e_issued")
            e_expire  = st.text_input("만료일 (YYYY-MM-DD, 무기한=9999-12-31)",
                value=sel_doc["만료일"] if sel_doc["만료일"] != "무기한" else "9999-12-31",
                key="c_e_expire")
            e_renewed = st.checkbox("갱신 완료 처리",
                value=sel_doc["갱신완료"], key="c_e_renewed")
        with ec2:
            e_renew_url = st.text_input("갱신처 URL/기관",
                value=sel_doc["갱신처"], key="c_e_url")
            e_file_note = st.text_input("파일 비고",
                value=sel_doc["비고"], key="c_e_note")
            e_memo      = st.text_area("메모 (자유 입력)",
                value=sel_doc["메모"], height=80, key="c_e_memo")

        cb1, cb2 = st.columns(2)
        with cb1:
            if st.button("💾 저장", type="primary", use_container_width=True, key="c_save"):
                try:
                    c_update_doc(
                        doc_id=sel_doc["id"],
                        issued=e_issued.strip(),
                        expire=e_expire.strip(),
                        renew_url=e_renew_url.strip(),
                        file_note=e_file_note.strip(),
                        renewed=e_renewed,
                        memo=e_memo.strip(),
                    )
                    st.success(f"✅ **{sel_doc_name}** 정보가 업데이트되었습니다.")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 저장 오류: {e}")
        with cb2:
            if st.button("🔄 갱신완료 토글", use_container_width=True, key="c_toggle"):
                c_toggle_renewed(sel_doc["id"])
                st.rerun()

    st.divider()

    # ── 서류 신규 추가 ────────────────────────────────────────
    with st.expander("➕ 서류 신규 추가"):
        na1, na2 = st.columns(2)
        with na1:
            n_name   = st.text_input("서류명 *", key="c_n_name")
            n_issued = st.text_input("발급일 (YYYY-MM-DD)", key="c_n_issued")
            n_expire = st.text_input("만료일 (YYYY-MM-DD)", key="c_n_expire")
        with na2:
            n_url    = st.text_input("갱신처", key="c_n_url")
            n_note   = st.text_input("파일 비고", key="c_n_note")
            n_memo   = st.text_area("메모", height=68, key="c_n_memo")

        if st.button("➕ 추가", type="primary", use_container_width=True, key="c_add"):
            if not n_name.strip():
                st.error("❌ 서류명을 입력하세요.")
            elif not n_expire.strip():
                st.error("❌ 만료일을 입력하세요.")
            else:
                try:
                    c_add_doc(n_name.strip(), n_issued.strip(), n_expire.strip(),
                              n_url.strip(), n_note.strip(), n_memo.strip())
                    st.success(f"✅ **{n_name.strip()}** 추가 완료")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 추가 오류: {e}")

    # ── 서류 삭제 ─────────────────────────────────────────────
    with st.expander("🗑️ 서류 삭제 (기본 7종 제외 추천)"):
        del_names = {d["서류명"]: d["id"] for d in all_docs}
        del_sel   = st.selectbox("삭제할 서류", list(del_names.keys()), key="c_del_sel")
        if st.button(f"🗑️ '{del_sel}' 삭제", type="secondary",
                     use_container_width=True, key="c_del_btn"):
            c_delete_doc(del_names[del_sel])
            st.success(f"🗑️ **{del_sel}** 삭제 완료")
            st.rerun()

    st.divider()

    # ── D-day 캘린더 뷰 ──────────────────────────────────────
    st.subheader("📅 만료일 타임라인")
    timeline_docs = [d for d in all_docs if d["만료일"] != "무기한"]
    if timeline_docs:
        today_dt = date.today()
        tl_rows  = []
        for d in sorted(timeline_docs, key=lambda x: x["만료일"]):
            exp   = date.fromisoformat(d["만료일"])
            diff  = (exp - today_dt).days
            bar_w = max(0, min(100, int((1 - diff / 365) * 100))) if diff <= 365 else 0
            tl_rows.append({
                "서류명":  d["서류명"],
                "만료일":  d["만료일"],
                "D-day":  d["D-day"],
                "상태":    d["상태"],
                "진행률":  bar_w,
            })
        df_tl = pd.DataFrame(tl_rows)
        st.dataframe(
            df_tl,
            use_container_width=True,
            hide_index=True,
            column_config={
                "진행률": st.column_config.ProgressColumn(
                    "⏱️ 만료 진행률", min_value=0, max_value=100
                )
            }
        )
    else:
        st.info("만료일이 설정된 서류가 없습니다.")

    # ── 갱신 가이드 ──────────────────────────────────────────
    with st.expander("📖 주요 서류 갱신 가이드"):
        st.markdown("""
| 서류명 | 갱신 방법 | 소요시간 |
|--------|-----------|---------|
| 소상공인 확인서 | [sminfo.mss.go.kr](https://sminfo.mss.go.kr) 접속 → 공동인증서 로그인 → 즉시 발급 | 5분 |
| 창업기업 확인서 | 중소벤처기업부 창업지원포털 → 개업일 7년 이내 발급 | 1일 |
| 재해율 확인서 | 안전보건공단 kosha.or.kr → 사업장 재해율 확인서 발급 | 당일 |
| 사업자등록증 | 홈택스 hometax.go.kr → 사업자등록증 재발급 | 즉시 |
| 허가증 | 화성시청 환경부서 방문 또는 팩스 신청 | 3~5일 |
        """)


# ============================================================
# [섹션 B] 💰 견적서 작성 UI (관리자 전용)
# ============================================================
elif role == "💰 견적서 작성":
    st.title("💰 음식물 견적서 작성")
    st.markdown(
        "<p style='color:#5f6368;'>학교별 단가·수거량 입력 → 견적서 PDF 즉시 생성 · 다운로드</p>",
        unsafe_allow_html=True
    )

    # ── 빠른선택 학교 ─────────────────────────────────────────
    B_SCHOOLS = {
        "직접입력": {"biz_no": "", "unit": 180, "volume": 0.0, "period": "2026.03.01 ~ 2027.02.28"},
        "서초고등학교":  {"biz_no": "210-83-00086", "unit": 180, "volume": 0.0,   "period": "2026.03.01 ~ 2027.02.28"},
        "당곡고등학교":  {"biz_no": "",             "unit": 150, "volume": 0.0,   "period": "2026.03.01 ~ 2027.02.28"},
        "국제고등학교":  {"biz_no": "",             "unit": 150, "volume": 0.0,   "period": "2026.03.01 ~ 2027.02.28"},
        "부림초등학교":  {"biz_no": "",             "unit": 120, "volume": 0.0,   "period": "2026.03.01 ~ 2027.02.28"},
    }
    # DB 등록 학교 동적 추가
    try:
        db_schools = [r[0] for r in get_conn().execute(
            "SELECT DISTINCT 학교명 FROM school_prices ORDER BY 학교명"
        ).fetchall()]
        for s in db_schools:
            if s not in B_SCHOOLS:
                B_SCHOOLS[s] = {"biz_no": "", "unit": 150, "volume": 0.0,
                                 "period": "2026.03.01 ~ 2027.02.28"}
    except Exception:
        pass

    b_preset = st.selectbox("🏫 빠른선택 학교", list(B_SCHOOLS.keys()), key="b_preset")
    b_data   = B_SCHOOLS[b_preset]

    st.divider()

    # ── 입력 폼 ──────────────────────────────────────────────
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("##### 🏫 고객(학교) 정보")
        b_school  = st.text_input("학교명 *",
            value=b_preset if b_preset != "직접입력" else "",
            key="b_school")
        b_biz_no  = st.text_input("학교 사업자번호",
            value=b_data["biz_no"], key="b_biz_no")
        b_period  = st.text_input("계약기간",
            value=b_data["period"], key="b_period")

    with col2:
        st.markdown("##### 💵 단가 및 수량")
        b_unit    = st.number_input("단가 (원/L) *", value=b_data["unit"],
                                     min_value=1, step=10, key="b_unit")
        b_volume  = st.number_input("연간 예상 수거량 (L)",
                                     value=b_data["volume"],
                                     min_value=0.0, step=100.0,
                                     help="0 입력 시 견적서 수량란 공백 처리",
                                     key="b_volume")
        # 공급가액 실시간 계산
        supply_amt = int(b_volume * b_unit) if b_volume > 0 else 0
        st.metric("공급가액 (자동 계산)", f"{supply_amt:,}원",
                  help="단가 × 수거량")

    # ── 미리보기 요약 ─────────────────────────────────────────
    with st.expander("📋 견적 내용 미리보기"):
        prev_b = {
            "고객명":   b_school,
            "사업자번호": b_biz_no,
            "견적일":   date.today().strftime("%Y.%m.%d"),
            "계약기간": b_period,
            "단가":     f"{b_unit}원/L",
            "수거량":   f"{b_volume:,.0f}L" if b_volume else "(미입력)",
            "공급가액": f"{supply_amt:,}원" if supply_amt else "(수량 입력 후 자동계산)",
            "세금":     "면세",
            "공급자":   f"{HY['name']} / {HY['ceo']} / {HY['biz_no']}",
            "연락처":   f"{HY['tel']} / {HY['email']}",
        }
        df_prev_b = pd.DataFrame(list(prev_b.items()), columns=["항목", "내용"])
        st.dataframe(df_prev_b, use_container_width=True, hide_index=True)

    st.divider()

    # ── 생성 버튼 ─────────────────────────────────────────────
    if st.button("💰 견적서 PDF 생성", type="primary",
                 use_container_width=True, key="b_gen"):
        if not b_school.strip():
            st.error("❌ 학교명을 입력해주세요.")
        else:
            with st.spinner("📄 PDF 생성 중..."):
                try:
                    pdf_path = generate_estimate_pdf(
                        school_name    = b_school.strip(),
                        school_biz_no  = b_biz_no.strip(),
                        volume_l       = float(b_volume),
                        unit_price     = int(b_unit),
                        contract_period= b_period.strip(),
                    )
                    with open(pdf_path, "rb") as f:
                        st.session_state["b_pdf_bytes"] = f.read()
                    st.session_state["b_pdf_name"] = (
                        f"음식물견적서_{b_school.strip()}_{date.today().strftime('%Y%m%d')}.pdf"
                    )
                    st.success("✅ 견적서 PDF 생성 완료!")
                except Exception as e:
                    st.error(f"❌ 오류: {e}")
                    st.info("💡 한글 폰트(malgun.ttf)가 설치되어 있는지 확인하세요.")

    # ── 다운로드 ──────────────────────────────────────────────
    b_pdf  = st.session_state.get("b_pdf_bytes")
    b_name = st.session_state.get("b_pdf_name", "음식물견적서.pdf")

    if b_pdf:
        st.download_button(
            label         = f"📥 {b_name} 다운로드",
            data          = b_pdf,
            file_name     = b_name,
            mime          = "application/pdf",
            key           = "b_dl",
            use_container_width=True,
            type          = "primary",
        )
        st.caption("💡 다운로드 후 출력 → 학교 제출 또는 계약서류 패키지에 포함")

    st.divider()

    # ── 일괄 견적 비교 (다학교) ──────────────────────────────
    with st.expander("📊 다학교 단가 비교표"):
        st.markdown("##### 현재 등록 학교별 단가 현황")
        try:
            price_rows = get_conn().execute(
                "SELECT 학교명, 음식물단가 FROM school_prices ORDER BY 음식물단가 DESC"
            ).fetchall()
            if price_rows:
                df_prices = pd.DataFrame(price_rows, columns=["학교명", "단가(원/L)"])
                df_prices["월 평균 견적(500L 기준)"] = df_prices["단가(원/L)"].apply(
                    lambda p: f"{int(p)*500:,}원"
                )
                st.dataframe(df_prices, use_container_width=True, hide_index=True)
            else:
                st.info("등록된 학교 단가 정보가 없습니다.")
        except Exception as e:
            st.warning(f"단가 조회 오류: {e}")

    # ── 하영자원 공급자 정보 ──────────────────────────────────
    with st.expander("ℹ️ 공급자(하영자원) 고정 정보"):
        df_hy3 = pd.DataFrame([{"항목": k, "내용": v} for k, v in HY.items()])
        st.dataframe(df_hy3, use_container_width=True, hide_index=True)
        st.caption("※ 위 정보는 견적서에 자동으로 입력됩니다.")


# ============================================================
# [섹션 A] 🏫 학교 마스터 관리 UI (관리자 전용)
# ============================================================
elif role == "🏫 학교 마스터 관리":
    st.title("🏫 학교 마스터 관리")
    st.markdown(
        "<p style='color:#5f6368;'>전체 계약 학교 정보 · 계약 이력 · 만료 D-day 통합 관리</p>",
        unsafe_allow_html=True
    )

    all_schools = a_get_all_schools()

    # ── 상단 요약 카드 ─────────────────────────────────────
    n_active  = sum(1 for s in all_schools if s["계약상태"] == "계약중")
    n_expire  = sum(1 for s in all_schools if "🔴" in s["계약D-day"] or "⛔" in s["계약D-day"])
    n_none    = sum(1 for s in all_schools if s["계약상태"] == "미계약")
    n_total   = len(all_schools)

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("🏫 전체 학교",   n_total)
    mc2.metric("✅ 계약중",      n_active)
    mc3.metric("🔴 만료임박",    n_expire)
    mc4.metric("⚪ 미계약",      n_none)

    # 만료임박 경고
    exp_schools = [s for s in all_schools
                   if ("🔴" in s["계약D-day"] or "⛔" in s["계약D-day"])
                   and s["계약상태"] == "계약중"]
    if exp_schools:
        st.error(f"🚨 **계약 만료임박 학교 {len(exp_schools)}곳**")
        for s in exp_schools:
            st.warning(
                f"**{s['학교명']}** — {s['계약D-day']}  |  "
                f"만료일: {s['계약종료']}  |  {s['교육청']}"
            )

    st.divider()

    # ── 학교 목록 탭 ──────────────────────────────────────
    tab_list, tab_edit, tab_contract = st.tabs(
        ["📑 전체 목록", "✏️ 학교 정보 수정", "📂 계약 이력"]
    )

    # ▸ 탭1: 전체 목록
    with tab_list:
        # 필터
        fl1, fl2 = st.columns(2)
        with fl1:
            edu_list   = ["전체"] + sorted({s["교육청"] for s in all_schools if s["교육청"]})
            sel_edu_f  = st.selectbox("교육청 필터", edu_list, key="a_edu_f")
        with fl2:
            sta_list   = ["전체", "계약중", "미계약", "계약만료", "협의중"]
            sel_sta_f  = st.selectbox("계약상태 필터", sta_list, key="a_sta_f")

        filtered = [
            s for s in all_schools
            if (sel_edu_f == "전체" or s["교육청"] == sel_edu_f)
            and (sel_sta_f == "전체" or s["계약상태"] == sel_sta_f)
        ]

        # ── [3단계 ①] D-day 색상 강화 ──────────────────────
        def _dday_color(dday_str: str) -> str:
            if "⛔" in dday_str: return "🟥"
            if "🔴" in dday_str: return "🟥"
            if "🟡" in dday_str: return "🟨"
            if "🟢" in dday_str: return "🟩"
            return "⬜"

        df_a = pd.DataFrame([{
            "학교명":       s["학교명"],
            "교육청":       s["교육청"],
            "단가(원/L)":  s["음식물단가"],
            "계약상태":     s["계약상태"],
            "계약만료":     s["계약종료"] or "-",
            "D-day":       s["계약D-day"],
            "⬛상태":       _dday_color(s["계약D-day"]),
            "담당자이메일": s.get("담당자이메일","") or "-",
        } for s in filtered])

        st.dataframe(df_a, use_container_width=True, hide_index=True,
                     column_config={
                         "단가(원/L)":  st.column_config.NumberColumn(format="%d원"),
                         "D-day":      st.column_config.TextColumn(width="small"),
                         "⬛상태":      st.column_config.TextColumn("상태", width="small"),
                         "담당자이메일": st.column_config.TextColumn(width="medium"),
                     })
        st.caption(f"총 {len(filtered)}개 학교 표시 중 (전체 {n_total}개)  "
                   f"🟥 만료/임박  🟨 90일 이내  🟩 정상  ⬜ 미설정")

        # CSV 다운로드
        csv_buf = io.StringIO()
        df_a.to_csv(csv_buf, index=False, encoding="utf-8-sig")
        st.download_button(
            "📥 학교 목록 CSV 다운로드",
            data=csv_buf.getvalue().encode("utf-8-sig"),
            file_name=f"학교마스터_{date.today().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="a_csv_dl"
        )

        # ── [3단계 ②] 갱신 필요 체크리스트 ─────────────────
        st.divider()
        st.markdown("##### 📋 갱신 필요 체크리스트 (D-90 이내)")
        renew_targets = [
            s for s in all_schools
            if s["계약상태"] == "계약중"
            and s["계약종료"]
            and s["계약종료"] not in ("", "9999-12-31")
            and ("⛔" in s["계약D-day"] or "🔴" in s["계약D-day"] or "🟡" in s["계약D-day"])
        ]

        if not renew_targets:
            st.success("✅ D-90 이내 만료 예정 학교 없음")
        else:
            st.warning(f"⚠️ 총 {len(renew_targets)}개 학교 갱신 필요")

            checklist_lines = [
                f"하영자원 계약 갱신 체크리스트",
                f"기준일: {date.today().strftime('%Y년 %m월 %d일')}",
                "=" * 50,
            ]
            for s in renew_targets:
                # a_get_all_schools()에 담당자이메일 포함됨 (8단계 수정)
                email = s.get("담당자이메일", "") or ""

                line = (f"[ ] {s['학교명']}"
                        f"  만료일: {s['계약종료']}"
                        f"  {s['계약D-day']}"
                        f"  이메일: {email or '미등록'}")
                checklist_lines.append(line)

                # 화면에도 표시
                st.markdown(
                    f"- **{s['학교명']}** &nbsp; {s['계약D-day']} &nbsp; "
                    f"만료: `{s['계약종료']}` &nbsp; 📧 {email or '이메일 미등록'}",
                    unsafe_allow_html=True
                )

            checklist_lines.append("=" * 50)
            checklist_txt = "\n".join(checklist_lines)

            st.download_button(
                label     = "📋 갱신 체크리스트 TXT 다운로드",
                data      = checklist_txt.encode("utf-8"),
                file_name = f"갱신체크리스트_{date.today().strftime('%Y%m%d')}.txt",
                mime      = "text/plain",
                key       = "a_checklist_dl",
            )

    # ▸ 탭2: 학교 정보 수정
    with tab_edit:
        school_names = [s["학교명"] for s in all_schools]
        sel_school   = st.selectbox("수정할 학교 선택", school_names, key="a_sel_school")
        sel_s        = next((s for s in all_schools if s["학교명"] == sel_school), {})

        # 담당자 전체 정보 조회 (부담당자 포함)
        try:
            sp_row = get_conn().execute(
                """SELECT 담당자명, 담당자연락처, 담당자이메일,
                          담당자2명, 담당자2연락처, 행정실팩스, 계약_사인_여부
                   FROM school_prices WHERE 학교명=?""",
                (sel_school,)
            ).fetchone() or ("","","","","","",0)
        except Exception:
            sp_row = ("","","","","","",0)

        ec1, ec2 = st.columns(2)
        with ec1:
            st.markdown("##### 🏫 학교 기본정보")
            a_biz_no = st.text_input("사업자번호",  value=sel_s.get("사업자번호",""), key="a_e_bno")
            a_addr   = st.text_input("주소",        value=sel_s.get("주소",""),       key="a_e_addr")
            a_tel    = st.text_input("전화번호",    value=sel_s.get("전화",""),        key="a_e_tel")
            a_unit   = st.number_input("음식물 단가(원/L)",
                                        value=int(sel_s.get("음식물단가", 150)),
                                        min_value=1, step=10, key="a_e_unit")
        with ec2:
            st.markdown("##### 📋 계약 정보")
            a_start  = st.text_input("계약 시작일", value=sel_s.get("계약시작",""), key="a_e_start")
            a_end    = st.text_input("계약 종료일", value=sel_s.get("계약종료",""), key="a_e_end")
            a_status = st.selectbox("계약 상태",
                                    ["계약중","미계약","계약만료","협의중"],
                                    index=["계약중","미계약","계약만료","협의중"].index(
                                        sel_s.get("계약상태","미계약")
                                        if sel_s.get("계약상태","미계약") in ["계약중","미계약","계약만료","협의중"]
                                        else "미계약"),
                                    key="a_e_status")
            a_note = st.text_area("비고", value=sel_s.get("비고",""), height=60, key="a_e_note")

        # ── [6단계] 담당자 고도화 — 주담당자 + 부담당자 ──────
        st.markdown("##### 👤 담당자 정보")
        d1, d2 = st.columns(2)
        with d1:
            st.caption("주담당자 (행정실)")
            a_mgr   = st.text_input("주담당자명",    value=sp_row[0], key="a_e_mgr")
            a_mgr_t = st.text_input("주담당자 연락처", value=sp_row[1], key="a_e_mgr_t")
            a_mgr_e = st.text_input("주담당자 이메일", value=sp_row[2], key="a_e_mgr_e")
            # 문자/메일 바로가기
            if sp_row[1]:
                st.markdown(f"📱 [문자 보내기](sms:{sp_row[1]})", unsafe_allow_html=False)
            if sp_row[2]:
                st.markdown(f"📧 [이메일 보내기](mailto:{sp_row[2]})", unsafe_allow_html=False)
        with d2:
            st.caption("부담당자 (영양사 등)")
            a_mgr2  = st.text_input("부담당자명",    value=sp_row[3], key="a_e_mgr2")
            a_mgr2t = st.text_input("부담당자 연락처", value=sp_row[4], key="a_e_mgr2t")
            a_fax   = st.text_input("행정실 팩스",   value=sp_row[5], key="a_e_fax")
            a_sign  = st.checkbox("계약서 도장/서명 수령 완료",
                                   value=bool(sp_row[6]), key="a_e_sign")

        # 저장 버튼
        if st.button("💾 전체 저장", type="primary", use_container_width=True, key="a_save"):
            try:
                a_update_school(
                    학교명=sel_school, 단가=int(a_unit),
                    사업자번호=a_biz_no.strip(), 주소=a_addr.strip(),
                    전화=a_tel.strip(), 시작일=a_start.strip(),
                    종료일=a_end.strip(), 상태=a_status,
                    비고=a_note.strip(), 담당자명=a_mgr.strip(),
                    담당자연락처=a_mgr_t.strip(), 담당자이메일=a_mgr_e.strip(),
                )
                a_update_contacts(
                    학교명=sel_school,
                    담당자2명=a_mgr2.strip(),
                    담당자2연락처=a_mgr2t.strip(),
                    행정실팩스=a_fax.strip(),
                    계약_사인_여부=a_sign,
                )
                st.success(f"✅ **{sel_school}** 정보 저장 완료")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 저장 오류: {e}")

        # ── [6단계] 담당자 명부 CSV 다운로드 ─────────────────
        st.divider()
        csv_contact = get_all_contact_csv()
        st.download_button(
            label     = "📥 전체 담당자 명부 CSV 다운로드",
            data      = csv_contact,
            file_name = f"하영자원_담당자명부_{date.today().strftime('%Y%m%d')}.csv",
            mime      = "text/csv",
            key       = "a_contact_csv",
        )

        # ── [6단계] 계약 체결 워크플로우 5단계 ───────────────
        st.divider()
        st.markdown("##### 🗂️ 계약 체결 워크플로우")
        st.caption("각 단계를 클릭하면 완료/미완료 토글 — 5단계 모두 완료 시 계약상태 자동 '계약중' 변경")

        WF_STEPS = [
            "📄 1단계: 견적서 발송",
            "📝 2단계: 위수탁계약서 발송",
            "📦 3단계: 서류 제출 완료",
            "🏛️ 4단계: 나라장터 전자계약",
            "✅ 5단계: 계약 완료",
        ]

        # 해당 학교 최신 계약 ID 조회
        try:
            wf_row = get_conn().execute(
                """SELECT id, 단계1_완료, 단계2_완료, 단계3_완료,
                          단계4_완료, 단계5_완료, 계약_상태
                   FROM contract_master WHERE 학교명=?
                   ORDER BY id DESC LIMIT 1""",
                (sel_school,)
            ).fetchone()
        except Exception:
            wf_row = None

        if not wf_row:
            st.info("⚠️ 계약 이력이 없습니다. 📂 계약 이력 탭에서 먼저 계약을 등록하세요.")
        else:
            wf_id    = wf_row[0]
            wf_done  = [bool(wf_row[i+1]) for i in range(5)]
            wf_count = sum(wf_done)
            wf_pct   = wf_count / 5

            # 진행률 바
            st.progress(wf_pct,
                text=f"진행률 {wf_count}/5단계 완료 "
                     f"({'🎉 계약완료!' if wf_count == 5 else f'{int(wf_pct*100)}%'})")

            # 단계별 토글 버튼
            wf_cols = st.columns(5)
            for i, (col, label, done) in enumerate(zip(wf_cols, WF_STEPS, wf_done)):
                with col:
                    icon  = "✅" if done else "☐"
                    color = "#34a853" if done else "#ea4335"
                    st.markdown(
                        f"<div style='text-align:center;font-size:11px;"
                        f"color:{color};font-weight:bold;'>{icon}</div>",
                        unsafe_allow_html=True
                    )
                    btn_label = label.split(":")[0]  # "📄 1단계"만 표시
                    if col.button(btn_label, key=f"wf_{wf_id}_{i+1}",
                                  use_container_width=True,
                                  type="primary" if done else "secondary"):
                        wf_toggle_step(wf_id, i + 1)
                        st.rerun()

            if wf_count == 5:
                st.success(f"🎉 **{sel_school}** 계약 체결 완료! 계약상태가 '계약중'으로 변경되었습니다.")

        st.divider()
        with st.expander("➕ 신규 학교 추가"):
            na1, na2 = st.columns(2)
            with na1:
                n_nm    = st.text_input("학교명 *",  key="a_n_nm")
                n_edu   = st.text_input("교육청",     key="a_n_edu")
                n_unit2 = st.number_input("단가(원/L)", value=150, min_value=1, step=10, key="a_n_unit")
                n_bno   = st.text_input("사업자번호", key="a_n_bno")
            with na2:
                n_addr2 = st.text_input("주소",       key="a_n_addr")
                n_tel2  = st.text_input("전화",        key="a_n_tel")
                n_s2    = st.text_input("계약시작일", key="a_n_start")
                n_e2    = st.text_input("계약종료일", key="a_n_end")
            n_sta2  = st.selectbox("계약상태", ["미계약","계약중","협의중"], key="a_n_sta")
            n_note2 = st.text_input("비고", key="a_n_note")
            if st.button("➕ 학교 추가", type="primary", use_container_width=True, key="a_add"):
                if not n_nm.strip():
                    st.error("❌ 학교명을 입력하세요.")
                else:
                    try:
                        a_add_school(n_nm.strip(), n_edu.strip(), int(n_unit2),
                                     n_bno.strip(), n_addr2.strip(), n_tel2.strip(),
                                     n_s2.strip(), n_e2.strip(), n_sta2, n_note2.strip())
                        st.success(f"✅ **{n_nm.strip()}** 추가 완료")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ 추가 오류: {e}")

    # ▸ 탭3: 계약 이력
    with tab_contract:
        st.markdown("##### 📂 전체 계약 이력")

        # 학교 필터
        c_school_filter = st.selectbox(
            "학교 선택 (전체 조회=전체)",
            ["전체"] + school_names,
            key="a_c_filter"
        )
        contracts = a_get_contracts(
            학교명=None if c_school_filter == "전체" else c_school_filter
        )

        if contracts:
            df_con = pd.DataFrame([{
                "ID":         c["id"],
                "학교명":     c["학교명"],
                "계약기간":   f"{c['계약_시작일']} ~ {c['계약_종료일']}",
                "단가(원/L)": c["단가"],
                "상태":       c["계약_상태"],
                "D-day":     c.get("D-day",""),
                "나라장터":   c["나라장터_번호"] or "-",
                "비고":       c["비고"] or "-",
            } for c in contracts])
            st.dataframe(df_con, use_container_width=True, hide_index=True)
        else:
            st.info("등록된 계약 이력이 없습니다.")

        # ── [5단계] 나라장터 계약번호 파서 ───────────────────
        st.divider()
        st.markdown("##### 🔍 나라장터 계약번호 파서")
        g2b_col1, g2b_col2 = st.columns([2, 1])
        with g2b_col1:
            g2b_input = st.text_input(
                "계약번호 입력 (예: R26TA01543339 00)",
                value="R26TA01543339 00",
                key="g2b_input"
            )
        with g2b_col2:
            st.markdown("<br>", unsafe_allow_html=True)
            g2b_parse_btn = st.button("🔍 파싱", key="g2b_parse", use_container_width=True)

        if g2b_parse_btn and g2b_input.strip():
            parsed = parse_g2b_contract_no(g2b_input.strip())
            if "오류" in parsed:
                st.error(f"❌ {parsed['오류']}")
            else:
                pc1, pc2, pc3, pc4 = st.columns(4)
                pc1.metric("계약유형",  parsed["계약유형"])
                pc2.metric("계약연도",  parsed["계약연도"])
                pc3.metric("기관코드",  parsed["기관코드"])
                pc4.metric("차수",      parsed["차수"])
                st.info(
                    f"**일련번호:** `{parsed['일련번호']}`  |  "
                    f"**차수코드:** `{parsed['차수코드']}`  |  "
                    f"**변경계약:** {'⚠️ 예' if parsed['변경계약여부'] else '✅ 아니오(원계약)'}"
                )
                if parsed["변경계약여부"]:
                    st.warning("⚠️ 변경계약입니다. 원계약과 비교 확인이 필요합니다.")

        # ── [5단계] 전자문서 수신 현황 ───────────────────────
        st.divider()
        st.markdown("##### 📩 전자문서 수신 현황")

        _status_icon = {
            "미수신":   "⚪",
            "수신":     "📩",
            "서명완료": "✍️",
            "계약완료": "✅",
            "반려":     "❌",
        }

        try:
            edoc_list = g2b_get_edoc_list()
        except Exception:
            edoc_list = []

        # 상태 필터
        edoc_status_f = st.selectbox(
            "상태 필터",
            ["전체", "미수신", "수신", "서명완료", "계약완료", "반려"],
            key="edoc_status_f"
        )
        edoc_filtered = [
            e for e in edoc_list
            if edoc_status_f == "전체" or e["계약서_상태"] == edoc_status_f
        ]

        if edoc_filtered:
            df_edoc = pd.DataFrame([{
                "ID":         e["id"],
                "학교명":     e["학교명"],
                "나라장터번호": e["나라장터_번호"] or "-",
                "전자문서번호": (e["전자문서번호"] or "-")[:20] + "..." if len(e.get("전자문서번호","")) > 20 else e.get("전자문서번호","-"),
                "수신일자":   e["수신일자"] or "-",
                "상태":       f"{_status_icon.get(e['계약서_상태'],'⚪')} {e['계약서_상태']}",
                "계약기간":   f"{e['계약_시작일']} ~ {e['계약_종료일']}",
            } for e in edoc_filtered])
            st.dataframe(df_edoc, use_container_width=True, hide_index=True)
        else:
            st.info("해당 조건의 전자문서가 없습니다.")

        # 전자문서 상태 변경 폼
        with st.expander("📩 전자문서 정보 입력 / 상태 변경"):
            ec1, ec2 = st.columns(2)
            with ec1:
                edoc_contract_id = st.number_input(
                    "계약 ID (위 목록의 ID 입력)", min_value=1, step=1, key="edoc_cid"
                )
                edoc_no   = st.text_input(
                    "전자문서번호 (UUID)",
                    value="019c93cb-58c9-7b21-981f-a64746d74cbc",
                    key="edoc_no"
                )
                edoc_sender = st.text_input(
                    "송신자 ID (발주기관)",
                    value="A311030206000",
                    key="edoc_sender"
                )
            with ec2:
                edoc_recv = st.text_input(
                    "수신일자 (YYYY-MM-DD)",
                    value=date.today().strftime("%Y-%m-%d"),
                    key="edoc_recv"
                )
                edoc_new_status = st.selectbox(
                    "계약서 상태",
                    ["미수신", "수신", "서명완료", "계약완료", "반려"],
                    key="edoc_new_status"
                )
                st.markdown(f"""
**상태별 의미:**
- ⚪ 미수신 — 전자문서 미도착
- 📩 수신 — 문서 도착, 검토 중
- ✍️ 서명완료 — 전자서명 완료 → **계약중 자동 변경**
- ✅ 계약완료 — 최종 완료
- ❌ 반려 — 반려, 재발행 필요
""")

            if st.button(
                "💾 전자문서 상태 저장",
                type="primary", use_container_width=True, key="edoc_save"
            ):
                try:
                    g2b_update_edoc(
                        contract_id = int(edoc_contract_id),
                        edoc_no     = edoc_no.strip(),
                        sender_id   = edoc_sender.strip(),
                        recv_date   = edoc_recv.strip(),
                        edoc_status = edoc_new_status,
                    )
                    if edoc_new_status == "서명완료":
                        st.success("✍️ 서명완료 처리! 계약상태가 **계약중**으로 자동 변경되었습니다.")
                    elif edoc_new_status == "계약완료":
                        st.success("✅ 계약완료 처리되었습니다.")
                    else:
                        st.success(f"💾 저장 완료: {edoc_new_status}")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 저장 오류: {e}")
        st.divider()
        st.markdown("##### 🔄 계약 갱신 완료 처리")
        st.caption("갱신 완료 시 계약 종료일 자동 +1년, 새 계약 이력 자동 등록")

        renew_schools = [
            s for s in all_schools
            if s["계약상태"] == "계약중"
            and s["계약종료"]
            and s["계약종료"] not in ("", "9999-12-31")
        ]

        if not renew_schools:
            st.info("갱신 처리 가능한 계약중 학교가 없습니다.")
        else:
            rn_col1, rn_col2 = st.columns(2)
            with rn_col1:
                renew_sel = st.selectbox(
                    "갱신할 학교 선택",
                    [s["학교명"] for s in renew_schools],
                    key="rn_sel"
                )
            renew_s = next((s for s in renew_schools if s["학교명"] == renew_sel), {})

            # 갱신 후 날짜 자동 계산 (+1년)
            try:
                from datetime import timedelta
                old_end   = date.fromisoformat(renew_s.get("계약종료","2027-02-28"))
                new_start = old_end + timedelta(days=1)
                new_end   = date(old_end.year + 1, old_end.month, old_end.day)
            except Exception:
                new_start = date.today()
                new_end   = date(date.today().year + 1, date.today().month, date.today().day)

            with rn_col2:
                st.info(
                    f"**{renew_sel}** 갱신 예정\n\n"
                    f"현재 만료: `{renew_s.get('계약종료','')}` → "
                    f"갱신 후: `{new_start}` ~ `{new_end}`"
                )

            rn_unit = st.number_input(
                "갱신 단가(원/L) — 변경 없으면 현재값 유지",
                value=int(renew_s.get("음식물단가", 150)),
                min_value=1, step=10, key="rn_unit"
            )
            rn_g2b  = st.text_input("갱신 나라장터 번호 (신규 계약번호)", key="rn_g2b")
            rn_note = st.text_input("갱신 비고", value="갱신 계약", key="rn_note")

            if st.button(
                f"✅ {renew_sel} 갱신완료 처리",
                type="primary", use_container_width=True, key="rn_btn"
            ):
                try:
                    # ① 새 계약 이력 등록
                    a_add_contract(
                        학교명      = renew_sel,
                        계약번호    = rn_g2b.strip(),
                        시작일      = str(new_start),
                        종료일      = str(new_end),
                        폐기물종류  = "음식물류폐기물",
                        단가        = int(rn_unit),
                        월예상량    = 0.0,
                        상태        = "계약중",
                        나라장터번호= rn_g2b.strip(),
                        비고        = rn_note.strip(),
                    )
                    # ② school_prices 계약 상태/날짜 업데이트
                    get_conn().execute(
                        """UPDATE school_prices SET
                           계약_시작일=?, 계약_종료일=?,
                           계약_상태=?, 음식물단가=?, updated_at=?
                           WHERE 학교명=?""",
                        (str(new_start), str(new_end),
                         "계약중", int(rn_unit),
                         datetime.now().strftime("%Y-%m-%d"), renew_sel)
                    )
                    get_conn().commit()
                    st.success(
                        f"✅ **{renew_sel}** 갱신 완료!\n\n"
                        f"새 계약기간: {new_start} ~ {new_end}  |  단가: {rn_unit}원/L"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 갱신 처리 오류: {e}")

        st.divider()
        with st.expander("➕ 계약 이력 신규 등록"):
            ci1, ci2 = st.columns(2)
            with ci1:
                ci_school  = st.selectbox("학교명 *", school_names, key="ci_school")
                ci_no      = st.text_input("계약번호", key="ci_no")
                ci_start   = st.text_input("시작일 (YYYY-MM-DD)", "2026-03-01", key="ci_start")
                ci_end     = st.text_input("종료일 (YYYY-MM-DD)", "2027-02-28", key="ci_end")
            with ci2:
                ci_waste   = st.text_input("폐기물 종류", "음식물류폐기물", key="ci_waste")
                ci_unit    = st.number_input("단가(원/L)", value=150, step=10, key="ci_unit")
                ci_vol     = st.number_input("월 예상량(L)", value=0.0, step=100.0, key="ci_vol")
                ci_g2b     = st.text_input("나라장터 번호", key="ci_g2b")
            ci_sta  = st.selectbox("계약상태", ["계약중","미계약","계약만료","협의중"], key="ci_sta")
            ci_note = st.text_input("비고", key="ci_note")

            if st.button("➕ 계약 등록", type="primary", use_container_width=True, key="ci_add"):
                try:
                    a_add_contract(
                        학교명=ci_school, 계약번호=ci_no.strip(),
                        시작일=ci_start.strip(), 종료일=ci_end.strip(),
                        폐기물종류=ci_waste.strip(), 단가=int(ci_unit),
                        월예상량=float(ci_vol), 상태=ci_sta,
                        나라장터번호=ci_g2b.strip(), 비고=ci_note.strip()
                    )
                    st.success(f"✅ {ci_school} 계약 등록 완료")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 등록 오류: {e}")

        if contracts:
            with st.expander("🗑️ 계약 이력 삭제"):
                del_id = st.number_input("삭제할 계약 ID", min_value=1, step=1, key="a_del_id")
                if st.button("🗑️ 삭제", type="secondary", use_container_width=True, key="a_del_btn"):
                    a_delete_contract(int(del_id))
                    st.success(f"ID {del_id} 삭제 완료")
                    st.rerun()
